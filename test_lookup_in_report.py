#!/usr/bin/env python3
"""
Test that LOOKUP values appear in the detailed test results
"""

import json
import os
from pathlib import Path
from openpyxl import load_workbook
from services.validation_service import ValidationPipeline

# Find the latest validation results
results_files = list(Path("output").glob("*_results.json"))
if not results_files:
    print("No results files found!")
    exit(1)

latest_file = max(results_files, key=lambda x: x.stat().st_mtime)
print(f"Using results file: {latest_file}")

# Generate Excel report
pipeline = ValidationPipeline()
excel_path = str(latest_file).replace("_results.json", "_test_report.xlsx")
pipeline.generate_excel_report(str(latest_file), excel_path)

print(f"Generated report: {excel_path}")

# Check the report content
wb = load_workbook(excel_path)

# Find a test tab with LOOKUP
for sheet_name in wb.sheetnames:
    if sheet_name in ["Guide", "IAG Summary Report"]:
        continue
    
    ws = wb[sheet_name]
    print(f"\nChecking sheet: {sheet_name}")
    
    # Find the detailed results section
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and "Detailed Test Results" in str(cell_value):
            print(f"Found detailed results at row {row}")
            
            # Check headers in the next row
            header_row = row + 1
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=header_row, column=col).value
                if header:
                    headers.append(header)
            
            print(f"Headers: {headers}")
            
            # Check if we have JobTitle or other lookup columns
            lookup_cols = [h for h in headers if h not in 
                          ["Item ID", "Audit Leader", "AuditLeader", "Status", 
                           "Failure Reason", "Internal Notes", "Audit Leader Response"]]
            
            if lookup_cols:
                print(f"Found lookup columns: {lookup_cols}")
                
                # Show first few data rows
                print("\nFirst 3 data rows:")
                for data_row in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(len(headers) + 1, ws.max_column + 1)):
                        value = ws.cell(row=data_row, column=col).value
                        row_data.append(str(value) if value else "")
                    print(f"Row {data_row}: {row_data}")
            else:
                print("No lookup columns found in headers")
            
            break

wb.close()
print("\nDone!")