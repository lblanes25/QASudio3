#!/usr/bin/env python3
"""
Test the complete IAG report with all 3 sections
"""

import glob
import json
from services.validation_service import ValidationPipeline
from openpyxl import load_workbook

def test_complete_iag_report():
    """Test generating complete IAG report with all sections"""
    
    # Find recent validation results
    results_files = sorted(glob.glob("output/*_results.json"))
    if not results_files:
        print("ERROR: No validation results files found")
        return False
        
    latest_results = results_files[-1]
    print(f"Using validation results: {latest_results}")
    
    # Load to check contents
    with open(latest_results, 'r') as f:
        results_data = json.load(f)
    
    print(f"Found {len(results_data.get('rule_results', {}))} rules")
    print(f"Found {len(results_data.get('grouped_summary', {}))} audit leaders")
    
    # Create pipeline and generate report
    print("\nGenerating complete IAG report...")
    pipeline = ValidationPipeline()
    
    output_path = "test_complete_iag_report.xlsx"
    pipeline.generate_excel_report(latest_results, output_path)
    
    # Verify the report
    print(f"\nReport generated: {output_path}")
    print("\nVerifying report structure...")
    
    wb = load_workbook(output_path)
    
    # Check main sheet
    assert "IAG Summary Report" in wb.sheetnames
    ws = wb["IAG Summary Report"]
    
    # Find section headers
    section1_row = None
    section2_row = None
    section3_row = None
    
    for row in range(1, 50):
        cell_value = ws[f'A{row}'].value
        if cell_value == "IAG Overall Results and Rating":
            section1_row = row
            print(f"✓ Section 1 found at row {row}")
        elif cell_value == "Audit Leader Overall Results and Ratings":
            section2_row = row
            print(f"✓ Section 2 found at row {row}")
        elif cell_value == "Detailed Analytics Section":
            section3_row = row
            print(f"✓ Section 3 found at row {row}")
    
    # Verify all sections exist
    assert section1_row is not None, "Section 1 not found"
    assert section2_row is not None, "Section 2 not found"
    assert section3_row is not None, "Section 3 not found"
    
    # Verify section spacing
    print(f"\nSection spacing:")
    print(f"  Section 1 to 2: {section2_row - section1_row} rows")
    print(f"  Section 2 to 3: {section3_row - section2_row} rows")
    
    # Check for individual rule tabs
    rule_tabs = [sheet for sheet in wb.sheetnames if sheet != "IAG Summary Report"]
    print(f"\n✓ Found {len(rule_tabs)} individual rule tabs:")
    
    # Check content of first test tab
    if rule_tabs:
        test_ws = wb[rule_tabs[0]]
        print(f"\nChecking first test tab: {rule_tabs[0]}")
        
        # Verify header section
        assert test_ws['A1'].value == "Test Name:", "Missing test name label"
        assert test_ws['A5'].value == "Error Threshold:", "Missing error threshold label"
        print("  ✓ Header section found (rows 1-5)")
        
        # Verify audit leader summary
        assert test_ws['A7'].value == "Audit Leader Summary", "Missing leader summary section"
        print("  ✓ Audit Leader Summary section found")
        
        # Look for detailed results section
        detail_found = False
        for row in range(10, 30):
            if test_ws[f'A{row}'].value and "Detailed Test Results" in str(test_ws[f'A{row}'].value):
                detail_found = True
                print(f"  ✓ Detailed Test Results section found at row {row}")
                break
        
        if not detail_found:
            print("  ! Detailed Test Results section not found (may need result_df data)")
    
    print("\n✅ Complete IAG report generated successfully!")
    print(f"\nOpen {output_path} to review the complete report")
    
    return True

if __name__ == "__main__":
    test_complete_iag_report()