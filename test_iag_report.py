#!/usr/bin/env python3
"""
Test script for IAG report generation - Phase 1
Tests Section 1 (IAG Overall Results) implementation
"""

import glob
import json
import os
from openpyxl import load_workbook
from services.validation_service import ValidationPipeline
from pathlib import Path


def test_iag_report_generation():
    """Test basic IAG report generation with Section 1"""
    try:
        # Find a recent validation results file
        results_files = sorted(glob.glob("output/*_results.json"))
        if not results_files:
            print("ERROR: No validation results files found in output/")
            return False
            
        latest_results = results_files[-1]
        print(f"Using validation results: {latest_results}")
        
        # Load and inspect the results first
        with open(latest_results, 'r') as f:
            results_data = json.load(f)
        
        # Check if we have grouped_summary data
        if 'grouped_summary' not in results_data:
            print("WARNING: No grouped_summary in results - report may be empty")
        else:
            print(f"Found {len(results_data['grouped_summary'])} audit leaders")
            # Print summary counts
            total_gc = total_pc = total_dnc = total_na = 0
            for leader, stats in results_data['grouped_summary'].items():
                print(f"  {leader}: GC={stats.get('GC', 0)}, PC={stats.get('PC', 0)}, "
                      f"DNC={stats.get('DNC', 0)}, NA={stats.get('NA', 0)}")
                total_gc += stats.get('GC', 0)
                total_pc += stats.get('PC', 0)
                total_dnc += stats.get('DNC', 0)
                total_na += stats.get('NA', 0)
            
            print(f"\nTotals: GC={total_gc}, PC={total_pc}, DNC={total_dnc}, NA={total_na}")
            expected_total_score = (total_gc * 5) + (total_pc * 3) + (total_dnc * 1)
            print(f"Expected Total Score: {expected_total_score}")
        
        # Create validation service instance
        print("\nCreating ValidationPipeline instance...")
        service = ValidationPipeline()
        
        # Generate report
        output_path = "test_iag_report.xlsx"
        print(f"Generating Excel report: {output_path}")
        service.generate_excel_report(latest_results, output_path)
        
        # Load and check the generated report
        print("\nVerifying generated report...")
        wb = load_workbook(output_path)
        
        # Check sheet exists
        assert "IAG Summary Report" in wb.sheetnames, "IAG Summary Report sheet not found"
        print("✓ IAG Summary Report sheet exists")
        
        # Verify Section 1 structure
        ws = wb["IAG Summary Report"]
        assert ws['A3'].value == "IAG Overall Results and Rating", "Section 1 header missing"
        print("✓ Section 1 header found")
        
        # Check all required fields - updated for executive-friendly format
        field_checks = [
            ('A4', 'Total Analytics Tested:'),
            ('A5', 'Total Data Points Reviewed:'),
            ('A6', 'Number of Audit Leaders:'),
            ('A7', 'Overall Compliance Rate:'),
            ('A8', 'Overall Rating:'),
            ('A9', 'Override Rating:'),
            ('C9', 'Rationale:')
        ]
        
        for cell, expected in field_checks:
            actual = ws[cell].value
            assert actual == expected, f"Cell {cell} should be '{expected}' but got '{actual}'"
            print(f"✓ {cell}: {expected}")
        
        # Verify values are populated
        assert ws['B4'].value is not None, "Total Analytics Tested not populated"
        assert ws['B5'].value is not None, "Total Data Points Reviewed not populated"
        assert ws['B6'].value is not None, "Number of Audit Leaders not populated"
        assert ws['B7'].value is not None, "Overall Compliance Rate not populated"
        assert ws['B8'].value is not None, "Overall Rating not populated"
        
        # Display calculated values
        print("\nCalculated Values:")
        print(f"  Total Analytics Tested: {ws['B4'].value}")
        print(f"  Total Data Points Reviewed: {ws['B5'].value}")
        print(f"  Number of Audit Leaders: {ws['B6'].value}")
        print(f"  Overall Compliance Rate: {ws['B7'].value}")
        print(f"  Overall Rating: {ws['B8'].value}")
        
        # Check rating color
        rating = ws['B8'].value
        fill = ws['B8'].fill
        if fill and fill.start_color:
            print(f"  Rating Color: {fill.start_color.rgb}")
            
        # Verify override fields are blank
        assert ws['B9'].value == "" or ws['B9'].value is None, "Override Rating should be blank"
        assert ws['D9'].value == "" or ws['D9'].value is None, "Rationale should be blank"
        print("\n✓ Override fields are blank (ready for manual entry)")
        
        # Check column widths
        print("\nColumn Widths:")
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            width = ws.column_dimensions[col].width
            print(f"  Column {col}: {width}")
        
        print("\n✅ All tests passed!")
        print(f"\nReport generated successfully: {output_path}")
        return True
        
    except Exception as e:
        print(f"\n❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    # Run the test
    success = test_iag_report_generation()
    
    # Optional: Open the Excel file for manual inspection
    if success and os.path.exists("test_iag_report.xlsx"):
        print("\nTo manually inspect the report, open: test_iag_report.xlsx")
        # Uncomment the next line if you want to auto-open the file
        # os.system("start test_iag_report.xlsx")  # Windows
        # os.system("open test_iag_report.xlsx")   # macOS
        # os.system("xdg-open test_iag_report.xlsx")  # Linux