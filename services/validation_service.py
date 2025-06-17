# services/validation_service.py

import pandas as pd
import threading
import logging
from typing import Dict, List, Any, Optional, Tuple, Union, Set
import json
import os
import shutil
from pathlib import Path
import datetime
import concurrent.futures
import pythoncom
from collections import defaultdict

# Import our components
from core.formula_engine.excel_formula_processor import ExcelFormulaProcessor
from core.rule_engine.rule_manager import ValidationRule, ValidationRuleManager
from core.rule_engine.rule_evaluator import RuleEvaluator, RuleEvaluationResult
from core.rule_engine.compliance_determiner import ComplianceDeterminer
from data_integration.io.importer import DataImporter
from data_integration.io.data_validator import DataValidator
# from reporting.generation.report_generator import ReportGenerator  # To be implemented

logger = logging.getLogger(__name__)


class ValidationPipeline:
    """
    Orchestrates the validation process by connecting data sources,
    validation rules, and result processing into a complete workflow.
    """

    def __init__(self,
                 rule_manager: Optional[ValidationRuleManager] = None,
                 evaluator: Optional[RuleEvaluator] = None,
                 data_importer: Optional[DataImporter] = None,
                 output_dir: Optional[str] = None,
                 archive_dir: Optional[str] = None,
                 max_workers: int = 4,
                 rule_config_paths: Optional[List[str]] = None,
                 report_config_path: Optional[str] = None,
                 lookup_manager: Optional[Any] = None):
        """
        Initialize the validation pipeline.

        Args:
            rule_manager: ValidationRuleManager for rule access
            evaluator: RuleEvaluator for rule evaluation
            data_importer: DataImporter for loading data
            output_dir: Directory for output files
            archive_dir: Directory for archiving output files
            max_workers: Maximum number of worker threads for parallel processing
            rule_config_paths: List of paths to YAML rule configuration files
            report_config_path: Path to YAML report configuration file
            lookup_manager: Optional SmartLookupManager for LOOKUP operations
        """
        self.rule_manager = rule_manager or ValidationRuleManager()
        self.evaluator = evaluator or RuleEvaluator(rule_manager=self.rule_manager)
        self.data_importer = data_importer or DataImporter()
        self.lookup_manager = lookup_manager
        # ReportGenerator removed - using generate_excel_report method instead

        # Set output directory
        self.output_dir = Path(output_dir) if output_dir else Path("./output")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Set archive directory if provided
        self.archive_dir = Path(archive_dir) if archive_dir else None
        if self.archive_dir:
            self.archive_dir.mkdir(parents=True, exist_ok=True)

        # Initialize data validator for pre-validation
        self.data_validator = DataValidator()

        # Set maximum worker threads for parallel processing
        self.max_workers = max_workers

        # Store rule configuration paths
        self.rule_config_paths = rule_config_paths or []

        # Load rules from configuration files
        if self.rule_config_paths:
            self._load_rule_configurations()

        # Initialize report generator with proper error handling
        # Look for template file for individual rule tabs
        template_path = None
        if self.output_dir.parent / "templates" / "qa_report_template.xlsx":
            template_path = str(self.output_dir.parent / "templates" / "qa_report_template.xlsx")
        
        if report_config_path:
            if os.path.exists(report_config_path):
                logger.info(f"Initializing report generator with configuration: {report_config_path}")
                # ReportGenerator removed - using generate_excel_report method instead
            else:
                logger.warning(f"Report configuration file not found: {report_config_path}")
                logger.info("Using default report configuration")
                # ReportGenerator removed - using generate_excel_report method instead
        else:
            logger.info("Initializing report generator with default configuration")
            # ReportGenerator removed - using generate_excel_report method instead

    def validate_data_source(self,
                             data_source: Union[str, pd.DataFrame],
                             rule_ids: Optional[List[str]] = None,
                             analytic_id: Optional[str] = None,
                             responsible_party_column: Optional[str] = None,
                             data_source_params: Optional[Dict[str, Any]] = None,
                             pre_validation: Optional[Dict[str, Any]] = None,
                             output_formats: Optional[List[str]] = None,
                             min_severity: Optional[str] = None,
                             exclude_rule_types: Optional[List[str]] = None,
                             expected_schema: Optional[Union[List[str], str]] = None,
                             use_parallel: bool = False,
                             report_config: Optional[str] = None,
                             use_all_rules: bool = False,
                             analytic_title: Optional[str] = None) -> Dict[str, Any]:
        """
        Run validation process on a data source.

        Args:
            data_source: Path to data file or DataFrame object
            rule_ids: List of rule IDs to apply (if None, uses all rules)
            analytic_id: Optional analytic ID to filter rules
            responsible_party_column: Column identifying responsible parties
            data_source_params: Parameters for data source loading
            pre_validation: Optional validation rules to apply before main validation
            output_formats: Output formats to generate ('json', 'excel', 'html', etc.)
            min_severity: Minimum severity level to include (e.g., 'critical', 'high')
            exclude_rule_types: List of rule types/categories to exclude
            expected_schema: Expected column list or path to schema file
            use_parallel: Whether to evaluate rules in parallel
            report_config: Optional path to report configuration YAML file
            use_all_rules: If True, use all available rules regardless of analytic_id
            analytic_title: Optional title for the analytic report (used in template reports)

        Returns:
            Dictionary with validation results
        """
        # If report_config specified, update the ReportGenerator
        if report_config:
            # ReportGenerator removed - report_config parameter is no longer used
            pass
            
        # If responsible_party_column specified, set it in metadata for all rules
        if responsible_party_column:
            for rule in self.rule_manager.list_rules():
                rule.metadata['responsible_party_column'] = responsible_party_column

        # Track timing for performance analysis
        start_time = datetime.datetime.now()

        # Initialize results structure
        results = {
            'valid': True,
            'analytic_id': analytic_id,
            'timestamp': start_time.isoformat(),
            'data_source': str(data_source) if isinstance(data_source, str) else "DataFrame",
            'rule_results': {},
            'summary': {},
            'output_files': []
        }

        try:
            # Load data if string path provided
            data_df = self._load_data(data_source, data_source_params)

            # Add basic data metrics to results
            results['data_metrics'] = {
                'row_count': len(data_df),
                'column_count': len(data_df.columns),
                'columns': list(data_df.columns)
            }
            
            # Add entity volume tracking if entity ID column exists
            entity_id_column = 'AuditEntityID'  # Standard entity ID column
            if entity_id_column in data_df.columns:
                results['data_metrics']['total_unique_entities'] = data_df[entity_id_column].nunique()
            else:
                results['data_metrics']['total_unique_entities'] = len(data_df)

            # Validate schema if provided
            if expected_schema:
                schema_valid, schema_errors = self._validate_schema(data_df, expected_schema)
                results['schema_validation'] = {
                    'valid': schema_valid,
                    'errors': schema_errors
                }

                # Exit if schema validation failed
                if not schema_valid:
                    results['valid'] = False
                    results['status'] = 'SCHEMA_VALIDATION_FAILED'
                    return results

            # Perform pre-validation if specified
            if pre_validation:
                pre_validation_results = self.data_validator.validate(
                    data_df, pre_validation, raise_exception=False
                )
                results['pre_validation'] = pre_validation_results

                # Exit if pre-validation failed
                if not pre_validation_results['valid']:
                    results['valid'] = False
                    results['status'] = 'PRE_VALIDATION_FAILED'
                    return results

            # Get rules to apply with filtering
            # If use_all_rules is True, don't filter by analytic_id
            rules = self._get_rules_to_apply(
                rule_ids,
                analytic_id if not use_all_rules else None,
                min_severity=min_severity,
                exclude_rule_types=exclude_rule_types
            )
            
            logger.info(f"Found {len(rules)} rules to apply")
            if rules:
                logger.info(f"Rule IDs: {[r.rule_id for r in rules[:5]]}{'...' if len(rules) > 5 else ''}")

            if not rules:
                results['valid'] = False
                results['status'] = 'NO_RULES_FOUND'
                logger.warning(f"No rules found. rule_ids={rule_ids}, analytic_id={analytic_id}")
                return results

            # Add rule metadata to results
            results['rules_applied'] = [rule.rule_id for rule in rules]

            # Evaluate rules (serially or in parallel)
            if use_parallel and len(rules) > 1:
                rule_results = self._evaluate_rules_parallel(rules, data_df, responsible_party_column)
            else:
                # For serial evaluation, we need to handle lookup_manager differently
                rule_results = {}
                for rule in rules:
                    try:
                        result = self.evaluator.evaluate_rule(
                            rule, data_df, responsible_party_column, self.lookup_manager
                        )
                        rule_results[rule.rule_id] = result
                    except Exception as e:
                        logger.error(f"Error evaluating rule {rule.rule_id}: {str(e)}")
                        # Continue with other rules even if one fails

            # Process evaluation results including grouping by responsible party
            self._process_evaluation_results(rule_results, results, responsible_party_column, data_df)

            # Generate outputs in requested formats
            if output_formats:
                output_paths = self._generate_outputs(results, rule_results, data_df, output_formats, 
                                                     analytic_title, responsible_party_column)
                results['output_files'].extend(output_paths)

                # Archive outputs if archive directory is configured
                if self.archive_dir:
                    archive_paths = self._archive_outputs(output_paths)
                    results['archived_files'] = archive_paths
                    
            # Store rule_results for potential leader pack generation
            results['_rule_evaluation_results'] = rule_results

            # Calculate total execution time
            end_time = datetime.datetime.now()
            results['execution_time'] = (end_time - start_time).total_seconds()

            return results

        except Exception as e:
            logger.error(f"Error in validation pipeline: {str(e)}")
            results['valid'] = False
            results['status'] = 'ERROR'
            results['error'] = str(e)

            # Calculate execution time even if error occurred
            end_time = datetime.datetime.now()
            results['execution_time'] = (end_time - start_time).total_seconds()

            return results

    def _load_data(self,
                   data_source: Union[str, pd.DataFrame],
                   params: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
        """
        Load data from file or use provided DataFrame.

        Args:
            data_source: Path to data file or DataFrame object
            params: Parameters for data loading

        Returns:
            DataFrame with loaded data
        """
        if isinstance(data_source, pd.DataFrame):
            return data_source

        # Use data importer to load file
        return self.data_importer.load_file(data_source, **(params or {}))

    def _validate_schema(self,
                         df: pd.DataFrame,
                         expected_schema: Union[List[str], str]) -> Tuple[bool, List[str]]:
        """
        Validate DataFrame against expected schema.

        Args:
            df: DataFrame to validate
            expected_schema: List of expected columns or path to schema file

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []

        # Load schema from file if string path provided
        if isinstance(expected_schema, str) and os.path.exists(expected_schema):
            try:
                # Try to load schema file (JSON or CSV)
                if expected_schema.lower().endswith('.json'):
                    with open(expected_schema, 'r') as f:
                        schema_data = json.load(f)
                        if isinstance(schema_data, list):
                            expected_columns = schema_data
                        elif isinstance(schema_data, dict) and 'columns' in schema_data:
                            expected_columns = schema_data['columns']
                        else:
                            errors.append(f"Invalid schema file format in {expected_schema}")
                            return False, errors
                elif expected_schema.lower().endswith('.csv'):
                    schema_df = pd.read_csv(expected_schema)
                    if 'column_name' in schema_df.columns:
                        expected_columns = schema_df['column_name'].tolist()
                    else:
                        expected_columns = schema_df.iloc[:, 0].tolist()
                else:
                    errors.append(f"Unsupported schema file format: {expected_schema}")
                    return False, errors
            except Exception as e:
                errors.append(f"Error loading schema file: {str(e)}")
                return False, errors
        else:
            # Use provided list directly
            expected_columns = expected_schema

        # Check that all expected columns exist
        actual_columns = set(df.columns)
        for col in expected_columns:
            if col not in actual_columns:
                errors.append(f"Expected column '{col}' is missing")

        # Optionally, check for unexpected columns
        # This is commented out as sometimes extra columns are acceptable
        # unexpected_columns = actual_columns - set(expected_columns)
        # for col in unexpected_columns:
        #     errors.append(f"Unexpected column '{col}' found")

        return len(errors) == 0, errors

    def _get_rules_to_apply(self,
                            rule_ids: Optional[List[str]] = None,
                            analytic_id: Optional[str] = None,
                            min_severity: Optional[str] = None,
                            exclude_rule_types: Optional[List[str]] = None) -> List[ValidationRule]:
        """
        Get list of rules to apply based on filters.

        Args:
            rule_ids: Optional list of specific rule IDs
            analytic_id: Optional analytic ID to filter rules
            min_severity: Minimum severity level to include
            exclude_rule_types: List of rule types/categories to exclude

        Returns:
            List of ValidationRule objects to apply
        """
        # If specific rule IDs provided, get those rules
        if rule_ids:
            rules = [self.rule_manager.get_rule(rule_id) for rule_id in rule_ids
                     if self.rule_manager.get_rule(rule_id) is not None]
        # If analytic ID provided, filter by that
        elif analytic_id:
            all_rules = self.rule_manager.list_rules()
            rules = [rule for rule in all_rules if rule.analytic_id == analytic_id]
        # Otherwise, return all rules
        else:
            rules = self.rule_manager.list_rules()

        # Apply severity filter if specified
        if min_severity and rules:
            severity_levels = {"critical": 4, "high": 3, "medium": 2, "low": 1, "info": 0}
            min_severity_level = severity_levels.get(min_severity.lower(), 0)

            filtered_rules = []
            for rule in rules:
                rule_severity = rule.severity.lower() if hasattr(rule, "severity") else "medium"
                rule_severity_level = severity_levels.get(rule_severity, 2)  # Default to medium

                if rule_severity_level >= min_severity_level:
                    filtered_rules.append(rule)

            rules = filtered_rules

        # Apply category/type exclusion if specified
        if exclude_rule_types and rules:
            exclude_types = [t.lower() for t in exclude_rule_types]
            rules = [rule for rule in rules
                     if not (hasattr(rule, "category") and rule.category.lower() in exclude_types)]

        return rules

    def validate_rule_configuration_file(file_path: str) -> Tuple[bool, List[str]]:
        """
        Validate a YAML rule configuration file without loading the rules.

        Args:
            file_path: Path to the YAML rule configuration file

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        import yaml

        errors = []

        try:
            path = Path(file_path)

            # Check if file exists
            if not path.exists():
                errors.append(f"File not found: {file_path}")
                return False, errors

            # Verify file is a YAML file
            if not path.suffix.lower() in ['.yaml', '.yml']:
                errors.append(f"File does not have a YAML extension: {file_path}")

            # Load and parse YAML
            with open(path, 'r') as f:
                try:
                    config = yaml.safe_load(f)
                except yaml.YAMLError as e:
                    errors.append(f"Invalid YAML: {str(e)}")
                    return False, errors

            # Check structure
            if not isinstance(config, dict):
                errors.append("Root element must be a dictionary")
                return False, errors

            if 'rules' not in config:
                errors.append("Missing 'rules' key in configuration")
                return False, errors

            if not isinstance(config['rules'], list):
                errors.append("'rules' must be a list")
                return False, errors

            # Validate each rule
            for i, rule in enumerate(config['rules']):
                rule_errors = []

                # Check rule is a dictionary
                if not isinstance(rule, dict):
                    errors.append(f"Rule #{i + 1} is not a dictionary")
                    continue

                # Check required fields
                if 'name' not in rule:
                    rule_errors.append("Missing required field 'name'")
                elif not isinstance(rule['name'], str) or not rule['name']:
                    rule_errors.append("'name' must be a non-empty string")

                if 'formula' not in rule:
                    rule_errors.append("Missing required field 'formula'")
                elif not isinstance(rule['formula'], str) or not rule['formula']:
                    rule_errors.append("'formula' must be a non-empty string")

                # Check optional fields
                if 'threshold' in rule:
                    try:
                        threshold = float(rule['threshold'])
                        if not 0 <= threshold <= 1:
                            rule_errors.append("'threshold' must be between 0 and 1")
                    except (ValueError, TypeError):
                        rule_errors.append(f"Invalid 'threshold' value: {rule['threshold']}")

                if 'severity' in rule and (not isinstance(rule['severity'], str) or
                                           rule['severity'].lower() not in ValidationRule.SEVERITY_LEVELS):
                    rule_errors.append(f"Invalid 'severity' value: {rule['severity']}. "
                                       f"Must be one of: {', '.join(ValidationRule.SEVERITY_LEVELS)}")

                if 'tags' in rule and not isinstance(rule['tags'], list):
                    rule_errors.append("'tags' must be a list")

                # Add rule errors to main errors list
                if rule_errors:
                    errors.append(f"Rule '{rule.get('name', f'#{i + 1}')}' has following issues:")
                    for err in rule_errors:
                        errors.append(f"  - {err}")

            return len(errors) == 0, errors

        except Exception as e:
            errors.append(f"Error validating file: {str(e)}")
            return False, errors

    def _load_rule_configurations(self) -> None:
        """
        Load validation rules from YAML configuration files.
        """
        import yaml

        loaded_rules_count = 0
        errors = []

        for config_path in self.rule_config_paths:
            try:
                # Ensure path exists
                path = Path(config_path)
                if not path.exists():
                    errors.append(f"Rule configuration file not found: {config_path}")
                    continue

                # Load YAML file
                with open(path, 'r') as f:
                    try:
                        config = yaml.safe_load(f)
                    except yaml.YAMLError as e:
                        errors.append(f"Invalid YAML in {config_path}: {str(e)}")
                        continue

                # Validate basic structure
                if not isinstance(config, dict) or 'rules' not in config:
                    errors.append(f"Invalid rule configuration format in {config_path}. 'rules' key not found")
                    continue

                if not isinstance(config['rules'], list):
                    errors.append(f"Invalid rule configuration format in {config_path}. 'rules' must be a list")
                    continue

                # Process each rule definition
                file_rules_count = 0
                file_errors = []

                for i, rule_config in enumerate(config['rules']):
                    try:
                        # Create and add rule to manager
                        rule = self._create_rule_from_config(rule_config, path)
                        self.rule_manager.add_rule(rule)
                        file_rules_count += 1
                    except Exception as e:
                        file_errors.append(f"Error in rule #{i + 1}: {str(e)}")

                # Log results for this file
                if file_errors:
                    error_message = f"Loaded {file_rules_count} rules from {config_path}, but encountered {len(file_errors)} errors:"
                    for err in file_errors:
                        error_message += f"\n  - {err}"
                    logger.warning(error_message)
                    errors.extend(file_errors)
                else:
                    logger.info(f"Successfully loaded {file_rules_count} rules from {config_path}")

                loaded_rules_count += file_rules_count

            except Exception as e:
                errors.append(f"Error processing rule configuration file {config_path}: {str(e)}")

        # Log overall results
        if loaded_rules_count > 0:
            logger.info(
                f"Loaded a total of {loaded_rules_count} rules from {len(self.rule_config_paths)} configuration files")

        if errors:
            error_summary = f"Encountered {len(errors)} errors while loading rule configurations:"
            for err in errors:
                error_summary += f"\n  - {err}"
            logger.error(error_summary)

    def _create_rule_from_config(self, rule_config: Dict[str, Any], config_path: Path) -> ValidationRule:
        """
        Create a ValidationRule instance from a configuration dictionary.

        Args:
            rule_config: Dictionary containing rule configuration
            config_path: Path to the source configuration file (for error context)

        Returns:
            ValidationRule instance

        Raises:
            ValueError: If the rule configuration is invalid
        """
        # Validate required fields
        required_fields = ['name', 'formula']
        for field in required_fields:
            if field not in rule_config:
                raise ValueError(f"Required field '{field}' missing from rule configuration")

        # Extract rule attributes with validation
        name = rule_config['name']
        if not name or not isinstance(name, str):
            raise ValueError(f"Rule name must be a non-empty string")

        formula = rule_config['formula']
        if not formula or not isinstance(formula, str):
            raise ValueError(f"Rule formula must be a non-empty string")

        # Optional fields with defaults
        description = rule_config.get('description', '')
        if not isinstance(description, str):
            raise ValueError(f"Rule description must be a string")

        threshold = rule_config.get('threshold', 1.0)
        try:
            threshold = float(threshold)
            if not 0 <= threshold <= 1:
                raise ValueError(f"Threshold must be between 0 and 1")
        except (ValueError, TypeError):
            raise ValueError(f"Invalid threshold value: {threshold}")

        severity = rule_config.get('severity', 'medium')
        if not isinstance(severity, str) or severity.lower() not in ValidationRule.SEVERITY_LEVELS:
            raise ValueError(
                f"Invalid severity level: {severity}. Must be one of: {', '.join(ValidationRule.SEVERITY_LEVELS)}")

        category = rule_config.get('category', 'data_quality')
        if not isinstance(category, str):
            raise ValueError(f"Category must be a string")

        tags = rule_config.get('tags', [])
        if not isinstance(tags, list):
            # Try to convert to list if it's a string
            if isinstance(tags, str):
                tags = [tags]
            else:
                raise ValueError(f"Tags must be a list of strings")

        # Generate a rule_id if not provided (based on file and rule name)
        rule_id = rule_config.get('rule_id')
        if not rule_id:
            # Create deterministic ID based on file name and rule name
            file_stem = config_path.stem
            rule_id = f"{file_stem}_{name}"

        # Extract metadata fields
        metadata = {}
        for key, value in rule_config.items():
            if key not in ['rule_id', 'name', 'formula', 'description',
                           'threshold', 'severity', 'category', 'tags']:
                metadata[key] = value

        # Create the ValidationRule
        return ValidationRule(
            rule_id=rule_id,
            name=name,
            formula=formula,
            description=description,
            threshold=threshold,
            severity=severity,
            category=category,
            tags=tags,
            metadata=metadata
        )

    def get_rule_configuration_summary(self) -> Dict[str, Any]:
        """
        Get a summary of rule configurations loaded from YAML files.

        Returns:
            Dictionary with summary information
        """
        rules = self.rule_manager.list_rules()

        # Count rules by category
        rules_by_category = {}
        for rule in rules:
            category = rule.category
            if category not in rules_by_category:
                rules_by_category[category] = 0
            rules_by_category[category] += 1

        # Count rules by severity
        rules_by_severity = {}
        for rule in rules:
            severity = rule.severity
            if severity not in rules_by_severity:
                rules_by_severity[severity] = 0
            rules_by_severity[severity] += 1

        # Count rules by tag
        tag_counts = {}
        for rule in rules:
            for tag in rule.tags:
                if tag not in tag_counts:
                    tag_counts[tag] = 0
                tag_counts[tag] += 1

        return {
            'total_rules': len(rules),
            'config_files': len(self.rule_config_paths),
            'rules_by_category': rules_by_category,
            'rules_by_severity': rules_by_severity,
            'tag_counts': tag_counts
        }

    def reload_rule_configurations(self) -> Dict[str, Any]:
        """
        Reload all rule configurations from YAML files.
        Useful when configuration files have been updated.

        Returns:
            Dictionary with reload results
        """
        # Keep track of previous rules
        previous_rules = {rule.rule_id: rule for rule in self.rule_manager.list_rules()}

        # Clear all rules from manager (alternative: replace specific rules)
        for rule_id in list(previous_rules.keys()):
            self.rule_manager.delete_rule(rule_id)

        # Record start time
        start_time = datetime.datetime.now()

        # Load rules from configuration files
        self._load_rule_configurations()

        # Get current rules
        current_rules = {rule.rule_id: rule for rule in self.rule_manager.list_rules()}

        # Calculate changes
        added_rules = [rule_id for rule_id in current_rules if rule_id not in previous_rules]
        removed_rules = [rule_id for rule_id in previous_rules if rule_id not in current_rules]
        kept_rules = [rule_id for rule_id in current_rules if rule_id in previous_rules]

        # Record end time and calculate duration
        end_time = datetime.datetime.now()
        duration = (end_time - start_time).total_seconds()

        return {
            'previous_rules_count': len(previous_rules),
            'current_rules_count': len(current_rules),
            'added_rules': added_rules,
            'added_count': len(added_rules),
            'removed_rules': removed_rules,
            'removed_count': len(removed_rules),
            'kept_rules': kept_rules,
            'kept_count': len(kept_rules),
            'duration_seconds': duration
        }

    def _evaluate_rules_parallel(self,
                                 rules: List[ValidationRule],
                                 data_df: pd.DataFrame,
                                 responsible_party_column: Optional[str] = None) -> Dict[str, RuleEvaluationResult]:
        """
        Evaluate rules in parallel using thread pool, with COM safety measures.
        """
        results = {}

        # Limit the number of worker threads to avoid Excel instance explosion
        max_workers = min(self.max_workers, 4)  # Cap at 4 workers regardless of setting

        logger.info(f"Evaluating {len(rules)} rules with {max_workers} worker threads")

        # Use thread pool to evaluate rules in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit tasks but keep track of future objects
            futures = []
            for rule in rules:
                # Create a copy of the DataFrame for each rule to avoid threading issues
                rule_df = data_df.copy()
                futures.append(executor.submit(self._evaluate_single_rule, rule, rule_df, responsible_party_column, self.lookup_manager))

            # Process futures as they complete
            for future in concurrent.futures.as_completed(futures):
                try:
                    rule_id, result = future.result()
                    if result:
                        results[rule_id] = result
                except Exception as e:
                    logger.error(f"Exception in thread pool task: {str(e)}")

        # Force garbage collection after all threads complete
        try:
            import gc
            gc.collect()
        except:
            pass

        return results

    def _evaluate_single_rule(self, rule, data_df, responsible_party_column, lookup_manager=None):
        """
        Worker function to evaluate a single rule with thread isolation.

        This isolates all COM operations to a single thread context.
        """
        try:
            # Initialize COM for this thread
            pythoncom.CoInitialize()
            logger.debug(f"COM initialized in thread {threading.current_thread().ident} for rule {rule.rule_id}")

            # Use with statement to ensure proper cleanup
            with ExcelFormulaProcessor(visible=False, lookup_manager=lookup_manager) as processor:
                # Create a dedicated evaluator for this thread
                thread_evaluator = RuleEvaluator(
                    rule_manager=self.rule_manager,
                    compliance_determiner=ComplianceDeterminer(),
                    excel_visible=False
                )

                # Evaluate the rule
                result = thread_evaluator.evaluate_rule(
                    rule, data_df, responsible_party_column, lookup_manager
                )

                # Return the result
                return rule.rule_id, result

        except Exception as e:
            logger.error(f"Error evaluating rule {rule.rule_id}: {str(e)}")
            return rule.rule_id, None

        finally:
            # Clean up COM in this thread
            try:
                pythoncom.CoUninitialize()
                logger.debug(f"COM uninitialized in thread {threading.current_thread().ident} for rule {rule.rule_id}")
            except Exception as e:
                logger.error(f"Error uninitializing COM in thread {threading.current_thread().ident}: {str(e)}")

    def _calculate_entities_per_leader(self, 
                                      data_df: pd.DataFrame, 
                                      responsible_party_column: str, 
                                      entity_id_column: str = 'AuditEntityID') -> Dict[str, int]:
        """
        Calculate unique entities per audit leader.
        
        Args:
            data_df: DataFrame with validation data
            responsible_party_column: Column identifying responsible parties
            entity_id_column: Column with entity IDs
            
        Returns:
            Dictionary mapping leader names to unique entity counts
        """
        if entity_id_column in data_df.columns and responsible_party_column in data_df.columns:
            entities_per_leader = data_df.groupby(responsible_party_column)[entity_id_column].nunique().to_dict()
            return entities_per_leader
        return {}

    def _process_evaluation_results(self,
                                    rule_results: Dict[str, RuleEvaluationResult],
                                    results: Dict[str, Any],
                                    responsible_party_column: Optional[str] = None,
                                    data_df: Optional[pd.DataFrame] = None) -> None:
        """
        Process and summarize rule evaluation results.

        Args:
            rule_results: Dictionary of rule evaluation results
            results: Results dictionary to update
            responsible_party_column: Column identifying responsible parties for grouping
            data_df: Original DataFrame for entity counting
        """
        # Track overall compliance
        overall_valid = True
        compliance_counts = {
            'GC': 0,   # Generally Conforms
            'PC': 0,   # Partially Conforms
            'DNC': 0,  # Does Not Conform
            'NA': 0    # Not Applicable
        }

        # Group rules by category and severity for reporting
        rule_stats = {
            'by_category': defaultdict(lambda: {'count': 0, 'GC': 0, 'PC': 0, 'DNC': 0, 'NA': 0}),
            'by_severity': defaultdict(lambda: {'count': 0, 'GC': 0, 'PC': 0, 'DNC': 0, 'NA': 0})
        }

        # Group results by responsible party if specified
        grouped_summary = defaultdict(lambda: {'count': 0, 'GC': 0, 'PC': 0, 'DNC': 0, 'NA': 0})

        # Process each rule result
        for rule_id, result in rule_results.items():
            # Add result summary to results
            # For now, always include details if available (can be optimized later)
            if hasattr(result, 'summary_with_details'):
                # Include detailed results for potential Excel report generation
                results['rule_results'][rule_id] = result.summary_with_details(include_all_rows=True)
            else:
                # Standard summary only
                results['rule_results'][rule_id] = result.summary

            # Update compliance counts
            compliance_status = result.compliance_status
            compliance_counts[compliance_status] += 1

            # Update rule statistics
            rule = result.rule
            category = rule.category if hasattr(rule, 'category') else 'uncategorized'
            severity = rule.severity if hasattr(rule, 'severity') else 'medium'

            # Update category stats
            rule_stats['by_category'][category]['count'] += 1
            rule_stats['by_category'][category][compliance_status] += 1

            # Update severity stats
            rule_stats['by_severity'][severity]['count'] += 1
            rule_stats['by_severity'][severity][compliance_status] += 1

            # Update overall validity (valid only if all rules are GC)
            if compliance_status != 'GC':
                overall_valid = False

            # Process group results if responsible party column specified
            if responsible_party_column and hasattr(result, 'party_results'):
                for party, party_result in result.party_results.items():
                    grouped_summary[party]['count'] += 1
                    grouped_summary[party][party_result['status']] += 1

        # Update summary information
        results['valid'] = overall_valid
        results['summary'] = {
            'total_rules': len(rule_results),
            'compliance_counts': compliance_counts,
            'compliance_rate': compliance_counts['GC'] / len(rule_results) if rule_results else 0,
            'rule_stats': rule_stats
        }

        # Add grouped summary if present
        if grouped_summary:
            # Calculate entities per leader if data_df provided
            entities_per_leader = {}
            if data_df is not None and responsible_party_column:
                entities_per_leader = self._calculate_entities_per_leader(
                    data_df, responsible_party_column
                )
                # Also add to data_metrics
                results['data_metrics']['entities_per_leader'] = entities_per_leader
            
            # Calculate compliance rates for each group
            formatted_groups = {}
            for party, counts in grouped_summary.items():
                if counts['count'] > 0:
                    formatted_groups[party] = {
                        'total_rules': counts['count'],
                        'GC': counts['GC'],
                        'PC': counts['PC'],
                        'DNC': counts['DNC'],
                        'NA': counts.get('NA', 0),
                        'compliance_rate': counts['GC'] / counts['count'] if counts['count'] > 0 else 0,
                        'entity_count': entities_per_leader.get(party, 0)
                    }

            results['grouped_summary'] = formatted_groups

        # Set overall status
        if overall_valid:
            results['status'] = 'FULLY_COMPLIANT'
        elif compliance_counts['DNC'] > 0:
            results['status'] = 'NON_COMPLIANT'
        else:
            results['status'] = 'PARTIALLY_COMPLIANT'
            
        # Add population summary if data_df provided
        if data_df is not None:
            results['population_summary'] = {
                'total_entities': results['data_metrics'].get('total_unique_entities', results['data_metrics']['row_count']),
                'entities_per_leader': results['data_metrics'].get('entities_per_leader', {}),
                'data_source': results.get('data_source', 'Unknown'),
                'filters_applied': [],  # TODO: Implement filter tracking
                'sampling_method': '100% population',
                'validation_date': results.get('timestamp', datetime.datetime.now().isoformat()),
                'total_rows': results['data_metrics']['row_count'],
                'total_columns': results['data_metrics']['column_count']
            }

    def _generate_outputs(self,
                          results: Dict[str, Any],
                          rule_results: Dict[str, RuleEvaluationResult],
                          data_df: pd.DataFrame,
                          output_formats: List[str],
                          analytic_title: Optional[str] = None,
                          responsible_party_column: Optional[str] = None) -> List[str]:
        """
        Generate output files in requested formats.

        Args:
            results: Validation results
            rule_results: Rule evaluation results
            data_df: Input DataFrame
            output_formats: List of output formats to generate

        Returns:
            List of generated output file paths
        """
        # Create timestamp string for filenames
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        analytic_id = results.get('analytic_id', 'validation')

        output_paths = []

        # Generate outputs in each requested format
        for format in output_formats:
            if format.lower() == 'excel_template':
                # Template report generation removed - TemplateBasedReportGenerator no longer available
                logger.info("Excel template format requested but TemplateBasedReportGenerator has been removed.")
                logger.info("Use generate_excel_report() method for IAG Excel reports.")
                # Skip template generation
                continue
                    
            elif format.lower() == 'json':
                # Export results to JSON
                json_path = self.output_dir / f"{analytic_id}_{timestamp}_results.json"
                with open(json_path, 'w') as f:
                    json.dump(results, f, indent=2, default=str)
                output_paths.append(str(json_path))

            elif format.lower() == 'excel':
                # Excel report generation removed - use generate_excel_report() method instead
                logger.info("Excel output format requested but ReportGenerator has been removed.")
                logger.info("Use the generate_excel_report() method for IAG Excel reports.")
                # Save results JSON for use with generate_excel_report
                json_path = self.output_dir / f"{analytic_id}_{timestamp}_results.json"
                if not json_path.exists():
                    with open(json_path, 'w') as f:
                        json.dump(results, f, indent=2, default=str)
                logger.info(f"Results saved to {json_path} - use this with generate_excel_report()")
                # Skip excel generation in this method
                continue

            elif format.lower() == 'html':
                # HTML report generation removed - ReportGenerator no longer available
                logger.info("HTML output format requested but ReportGenerator has been removed.")
                logger.info("HTML reports are not currently supported.")
                # Skip HTML generation
                continue

            elif format.lower() == 'csv':
                # Export summary results to CSV
                csv_path = self.output_dir / f"{analytic_id}_{timestamp}_summary.csv"
                self._export_to_csv(results, csv_path)
                output_paths.append(str(csv_path))

        return output_paths

    def _archive_outputs(self, output_paths: List[str]) -> List[str]:
        """
        Archive output files to the archive directory.

        Args:
            output_paths: List of paths to output files

        Returns:
            List of archived file paths
        """
        if not self.archive_dir:
            return []

        # Create timestamped subdirectory in archive
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_subdir = self.archive_dir / timestamp
        archive_subdir.mkdir(exist_ok=True)

        archived_paths = []

        # Copy each file to archive
        for path in output_paths:
            src_path = Path(path)
            if src_path.exists():
                dst_path = archive_subdir / src_path.name
                shutil.copy2(src_path, dst_path)
                archived_paths.append(str(dst_path))

        return archived_paths

    def _export_to_excel(self,
                         results: Dict[str, Any],
                         rule_results: Dict[str, RuleEvaluationResult],
                         data_df: pd.DataFrame,
                         output_path: Path) -> None:
        """
        Export detailed results to Excel file.

        Args:
            results: Validation results
            rule_results: Rule evaluation results
            data_df: Input DataFrame
            output_path: Path for Excel output file
        """
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                # Create summary sheet
                summary_data = {
                    'Metric': ['Status', 'Total Rules', 'GC Count', 'PC Count', 'DNC Count',
                               'Compliance Rate', 'Execution Time'],
                    'Value': [
                        results['status'],
                        results['summary']['total_rules'],
                        results['summary']['compliance_counts']['GC'],
                        results['summary']['compliance_counts']['PC'],
                        results['summary']['compliance_counts']['DNC'],
                        f"{results['summary']['compliance_rate']:.2%}",
                        f"{results['execution_time']:.2f} seconds"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

                # Create Rules sheet with detailed rule results
                rules_data = []
                for rule_id, rule_result in results['rule_results'].items():
                    rules_data.append({
                        'Rule ID': rule_id,
                        'Rule Name': rule_result['rule_name'],
                        'Status': rule_result['compliance_status'],
                        'Total Items': rule_result['total_items'],
                        'GC Count': rule_result['gc_count'],
                        'PC Count': rule_result['pc_count'],
                        'DNC Count': rule_result['dnc_count'],
                        'Compliance Rate': rule_result['compliance_rate']
                    })

                if rules_data:
                    rules_df = pd.DataFrame(rules_data)
                    rules_df.to_excel(writer, sheet_name='Rules', index=False)

                # Export a sample of the input data
                data_sample = data_df.head(100)  # Just first 100 rows as sample
                data_sample.to_excel(writer, sheet_name='Data Sample', index=False)

                # Create Compliance Breakdown by Category
                if 'rule_stats' in results['summary']:
                    # By Category
                    category_data = []
                    for category, stats in results['summary']['rule_stats']['by_category'].items():
                        compliance_rate = stats['GC'] / stats['count'] if stats['count'] > 0 else 0
                        category_data.append({
                            'Category': category,
                            'Total Rules': stats['count'],
                            'GC': stats['GC'],
                            'PC': stats['PC'],
                            'DNC': stats['DNC'],
                            'Compliance Rate': compliance_rate
                        })

                    if category_data:
                        category_df = pd.DataFrame(category_data)
                        category_df.to_excel(writer, sheet_name='By Category', index=False)

                    # By Severity
                    severity_data = []
                    for severity, stats in results['summary']['rule_stats']['by_severity'].items():
                        compliance_rate = stats['GC'] / stats['count'] if stats['count'] > 0 else 0
                        severity_data.append({
                            'Severity': severity,
                            'Total Rules': stats['count'],
                            'GC': stats['GC'],
                            'PC': stats['PC'],
                            'DNC': stats['DNC'],
                            'Compliance Rate': compliance_rate
                        })

                    if severity_data:
                        severity_df = pd.DataFrame(severity_data)
                        severity_df.to_excel(writer, sheet_name='By Severity', index=False)

                # Add Responsible Party breakdown if available
                if 'grouped_summary' in results and results['grouped_summary']:
                    party_data = []
                    for party, stats in results['grouped_summary'].items():
                        party_data.append({
                            'Responsible Party': party,
                            'Total Rules': stats['total_rules'],
                            'GC': stats['GC'],
                            'PC': stats['PC'],
                            'DNC': stats['DNC'],
                            'Compliance Rate': stats['compliance_rate']
                        })

                    if party_data:
                        party_df = pd.DataFrame(party_data)
                        party_df.to_excel(writer, sheet_name='By Responsible Party', index=False)

                # Export detailed results for each rule
                for rule_id, result in rule_results.items():
                    # Get failures only
                    failure_df = result.get_failing_items()
                    if not failure_df.empty:
                        # Limit to first 1000 rows if very large
                        if len(failure_df) > 1000:
                            failure_df = failure_df.head(1000)

                        # Excel sheet names have 31 char limit
                        sheet_name = f"Rule_{rule_id[-20:]}" if len(rule_id) > 20 else f"Rule_{rule_id}"
                        failure_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Add styling to make the report more readable
                workbook = writer.book
                # Add formats for different compliance statuses
                gc_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
                pc_format = workbook.add_format({'bg_color': '#FFEB9C', 'font_color': '#9C6500'})
                dnc_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})

                # Apply conditional formatting to Rules sheet
                if rules_data:
                    rules_sheet = writer.sheets['Rules']
                    rules_sheet.conditional_format('C2:C1000', {'type': 'text',
                                                                'criteria': 'containing',
                                                                'value': 'GC',
                                                                'format': gc_format})
                    rules_sheet.conditional_format('C2:C1000', {'type': 'text',
                                                                'criteria': 'containing',
                                                                'value': 'PC',
                                                                'format': pc_format})
                    rules_sheet.conditional_format('C2:C1000', {'type': 'text',
                                                                'criteria': 'containing',
                                                                'value': 'DNC',
                                                                'format': dnc_format})

        except Exception as e:
            logger.error(f"Error exporting results to Excel: {str(e)}")
            raise

    def _export_to_csv(self, results: Dict[str, Any], output_path: Path) -> None:
        """
        Export summary results to CSV file.

        Args:
            results: Validation results
            output_path: Path for CSV output file
        """
        try:
            # Extract rule results for CSV export
            csv_data = []
            for rule_id, rule_result in results['rule_results'].items():
                csv_data.append({
                    'Rule ID': rule_id,
                    'Rule Name': rule_result['rule_name'],
                    'Status': rule_result['compliance_status'],
                    'Total Items': rule_result['total_items'],
                    'GC Count': rule_result['gc_count'],
                    'PC Count': rule_result['pc_count'],
                    'DNC Count': rule_result['dnc_count'],
                    'Compliance Rate': rule_result['compliance_rate'],
                    'Analytic ID': results.get('analytic_id', ''),
                    'Timestamp': results.get('timestamp', '')
                })

            if csv_data:
                # Convert to DataFrame and export to CSV
                csv_df = pd.DataFrame(csv_data)
                csv_df.to_csv(output_path, index=False)

        except Exception as e:
            logger.error(f"Error exporting results to CSV: {str(e)}")
            raise

    def aggregate_analytics(
            self,
            result_paths: List[str],
            output_formats: Optional[List[str]] = None,
            weights_config: Optional[str] = None,
            output_dir: Optional[str] = None,
            report_config: Optional[str] = None,
            return_report: bool = False
    ) -> Dict[str, Any]:
        """
        Aggregate results from multiple analytics runs.

        Args:
            result_paths: Paths to JSON result files from previous validation runs
            output_formats: Output formats to generate ('json', 'excel', etc.)
            weights_config: Path to weights configuration file
            output_dir: Directory for output files (defaults to self.output_dir)
            report_config: Path to report configuration YAML file
            return_report: Whether to include the full report structure in the return value

        Returns:
            Dictionary with aggregation results and output file paths
        """
        from business_logic.aggregation.analytics_aggregator import (
            aggregate_analytics_results,
            load_weights_configuration,
            create_summary_report
        )
        from datetime import datetime  # Fixed missing import

        logger.info(f"Aggregating results from {len(result_paths)} analytics runs")

        # Load result files
        result_dicts = []
        for path in result_paths:
            try:
                with open(path, 'r') as f:
                    result = json.load(f)
                    result_dicts.append(result)
                    logger.debug(f"Successfully loaded results from {path}")
            except Exception as e:
                logger.error(f"Error loading results from {path}: {str(e)}")

        if not result_dicts:
            return {
                'success': False,
                'error': 'No valid result files loaded'
            }

        # Load weights configuration if specified
        weights_config_dict = None
        if weights_config:
            weights_config_dict = load_weights_configuration(weights_config)

        # Set output directory
        if output_dir:
            output_dir_path = Path(output_dir)
        else:
            output_dir_path = self.output_dir

        output_dir_path.mkdir(parents=True, exist_ok=True)

        # Perform aggregation
        summary = aggregate_analytics_results(
            result_dicts=result_dicts,
            weights_config=weights_config_dict
        )

        # Add robust logging after aggregation
        logger.info(f"Aggregation complete. Leaders: {len(summary.leader_summary)}, Rules: {len(summary.rule_details)}")

        # Create summary report
        report = create_summary_report(summary)

        # Generate outputs in requested formats
        output_paths = []
        if output_formats:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            for fmt in output_formats:
                if fmt.lower() == 'json':
                    # Export aggregated results to JSON
                    json_path = output_dir_path / f"aggregated_results_{timestamp}.json"
                    summary.export_to_file(str(json_path))
                    output_paths.append(str(json_path))
                    logger.info(f"Exported aggregated results to {json_path}")

                elif fmt.lower() == 'excel':
                    # Create Excel report
                    excel_path = output_dir_path / f"aggregated_report_{timestamp}.xlsx"

                    # Aggregate Excel generation removed - ReportGenerator no longer available
                    logger.warning("Aggregate Excel report generation has been removed.")
                    logger.info("Use generate_excel_report() method for IAG Excel reports.")
                else:
                    # Report format validation - log unsupported formats
                    logger.warning(f"Unsupported output format: {fmt}")

        # Return aggregation results with optional full report
        result = {
            'success': True,
            'timestamp': datetime.now().isoformat(),
            'leader_count': len(summary.leader_summary),
            'rule_count': len(summary.rule_details),
            'department_summary': summary.department_summary,
            'output_files': output_paths
        }

        # Include full report if requested
        if return_report:
            result['report'] = report

        return result

    def generate_leader_packs(self,
                              results: Dict[str, Any],
                              rule_results: Dict[str, RuleEvaluationResult],
                              responsible_party_column: Optional[str] = None,
                              output_dir: Optional[str] = None,
                              selected_leaders: Optional[List[str]] = None,
                              include_only_failures: bool = False,
                              generate_email_content: bool = False,
                              zip_output: bool = True,
                              export_csv_summary: bool = False,
                              suppress_logs: bool = False) -> Dict[str, Any]:
        """
        Generate individual Excel reports for each audit leader containing only their relevant data.
        This method produces per-leader Excel reports summarizing compliance results, for distribution
        or documentation purposes.

        Args:
            results: Validation results dictionary (from validate_data_source)
            rule_results: Dictionary of rule evaluation results
            responsible_party_column: Column name for identifying responsible parties
            output_dir: Directory to save the leader packs (defaults to self.output_dir)
            selected_leaders: Optional list of specific leaders to generate packs for
            include_only_failures: Whether to only include leaders with at least one failed rule
            generate_email_content: Whether to generate email-ready summaries
            zip_output: Whether to create a ZIP file containing all leader packs
            export_csv_summary: Whether to export a CSV summary of leader metrics
            suppress_logs: Whether to suppress detailed logging (for automated workflows)

        Returns:
            Dictionary with generation results including paths to all generated files
        """
        # Configure logging
        log_level = logging.INFO
        if suppress_logs:
            log_level = logging.WARNING

        # Use default output directory if not specified
        if output_dir is None:
            output_dir = self.output_dir

        # Validate rule_results has content
        if not rule_results:
            error_msg = "No rule evaluation results provided"
            if not suppress_logs:
                logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg
            }

        # If responsible_party_column not specified, try to find it in results
        inferred_column = False
        if responsible_party_column is None:
            # Look for responsible_party_column in rule metadata
            for rule_id, result in rule_results.items():
                if (hasattr(result, 'rule') and
                        hasattr(result.rule, 'metadata') and
                        'responsible_party_column' in result.rule.metadata):
                    responsible_party_column = result.rule.metadata['responsible_party_column']
                    inferred_column = True
                    break

        # If we still don't have responsible_party_column, use default
        if responsible_party_column is None:
            # Try a common default based on context
            if 'grouped_summary' in results and results['grouped_summary']:
                # Name the column "Audit Leader" as a default
                responsible_party_column = "Audit Leader"
                inferred_column = True

                if not suppress_logs:
                    logger.warning(f"No responsible_party_column specified, using default: {responsible_party_column}")
            else:
                error_msg = "No responsible_party_column specified and couldn't determine from context"
                if not suppress_logs:
                    logger.error(error_msg)
                return {
                    "success": False,
                    "error": error_msg
                }

        # Log the responsible party column if it was inferred
        if inferred_column and not suppress_logs:
            logger.info(f"Using responsible party column: {responsible_party_column} (inferred from context)")

        # ReportGenerator removed - this method is deprecated
        logger.warning("generate_leader_packs is deprecated. ReportGenerator has been removed.")
        logger.warning("Use generate_excel_report() for IAG reports instead.")
        
        return {
            "success": False,
            "error": "generate_leader_packs is deprecated. Use generate_excel_report() instead.",
            "message": "ReportGenerator has been removed from the project."
        }
    
    def generate_iag_summary_report(
        self,
        results: Dict[str, Any],
        rule_results: Dict[str, Any],
        responsible_party_column: str,
        output_path: Optional[Union[str, Path]] = None,
        **kwargs
    ) -> Union[Path, Dict[str, Any]]:
        """
        Generate IAG and AL Results and Ratings Summary report.
        
        Args:
            results: Validation results from validate_data_source()
            rule_results: Rule evaluation results dictionary
            responsible_party_column: Column for audit leader grouping
            output_path: Optional output path (auto-generated if None)
            **kwargs: Additional options passed to IAGSummaryGenerator:
                - review_year_name: Header text (e.g., "2024 Q3 Compliance Review")
                - analytics_metadata: Dict with error thresholds, risk levels, etc.
                - manual_overrides: Dict for manual rating overrides
                
        Returns:
            Path to generated IAG summary Excel file, or error dict if failed
        """
        try:
            # Generate default output path if not provided
            if not output_path:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                analytic_id = results.get('analytic_id', 'validation')
                output_path = self.output_dir / f"{analytic_id}_{timestamp}_IAG_Summary.xlsx"
            
            # Ensure output_path is a Path object
            output_path = Path(output_path)
            
            # Log generation start
            logger.info(f"Generating IAG summary report for {len(rule_results)} rules")
            
            # ReportGenerator removed - this method is deprecated
            logger.warning("generate_iag_summary_report is deprecated. ReportGenerator has been removed.")
            logger.warning("Use generate_excel_report() for IAG reports instead.")
            
            return {
                "success": False,
                "error": "generate_iag_summary_report is deprecated. Use generate_excel_report() instead.",
                "message": "ReportGenerator has been removed from the project."
            }
            
        except Exception as e:
            error_msg = f"Failed to generate IAG summary report: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "error": error_msg
            }
    
    def generate_comprehensive_iag_report(
        self,
        results: Dict[str, Any],
        rule_results: Dict[str, Any],
        responsible_party_column: str,
        output_path: Optional[Union[str, Path]] = None,
        **kwargs
    ) -> Union[Path, Dict[str, Any]]:
        """
        Generate comprehensive IAG report with summary and individual analytic tabs.
        
        This method generates a complete Excel workbook containing:
        1. IAG Summary tab with overall results and ratings
        2. Individual analytic tabs for each rule (QA-ID format)
        
        Args:
            results: Validation results from validate_data_source()
            rule_results: Rule evaluation results dictionary
            responsible_party_column: Column for audit leader grouping
            output_path: Optional output path (auto-generated if None)
            **kwargs: Additional options for report generation
                
        Returns:
            Path to generated comprehensive IAG Excel file, or error dict if failed
        """
        try:
            # Generate default output path if not provided
            if not output_path:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                analytic_id = results.get('analytic_id', 'validation')
                output_path = self.output_dir / f"{analytic_id}_{timestamp}_Comprehensive_IAG.xlsx"
            
            # Ensure output_path is a Path object
            output_path = Path(output_path)
            
            # Get source data for detailed results
            data_source = results.get('data_source')
            if isinstance(data_source, str):
                # Load the data file
                source_data = self.data_importer.load_file(data_source)
            elif isinstance(data_source, pd.DataFrame):
                source_data = data_source
            else:
                # Try to get from first rule result
                first_result = next(iter(rule_results.values()))
                if hasattr(first_result, 'result_df'):
                    source_data = first_result.result_df
                else:
                    raise ValueError("Could not determine source data for detailed results")
            
            # Log generation start
            logger.info(f"Generating comprehensive IAG report with {len(rule_results)} individual analytic tabs")
            
            # ReportGenerator removed - this method is deprecated
            logger.warning("generate_comprehensive_iag_report is deprecated. ReportGenerator has been removed.")
            logger.warning("Use generate_excel_report() for IAG reports instead.")
            
            return {
                "success": False,
                "error": "generate_comprehensive_iag_report is deprecated. Use generate_excel_report() instead.",
                "message": "ReportGenerator has been removed from the project."
            }
            
        except Exception as e:
            error_msg = f"Failed to generate comprehensive IAG report: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "error": error_msg
            }
    
    def generate_excel_report(self, validation_results_path: str, output_path: str) -> None:
        """
        Generate complete IAG Summary Report with all 3 sections.
        
        Args:
            validation_results_path: Path to validation results JSON file
            output_path: Path where Excel report should be saved
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from core.scoring.iag_scoring_calculator import IAGScoringCalculator
        
        # Ensure rules are loaded from data/rules directory
        if len(self.rule_manager.list_rules()) == 0:
            logger.info("No rules loaded, loading from data/rules directory")
            self.rule_manager.load_rules_from_directory("data/rules")
            logger.info(f"Loaded {len(self.rule_manager.list_rules())} rules")
        
        # Read validation results
        with open(validation_results_path, 'r') as f:
            results = json.load(f)
        
        # Create workbook with Guide as first tab
        wb = Workbook()
        
        # Create Guide tab first
        ws = wb.active
        ws.title = "Guide"
        self._create_guide_tab_content(ws)
        
        # Create IAG Summary Report as second tab
        ws = wb.create_sheet("IAG Summary Report")
        
        # Get responsible party column from first rule
        first_rule_id = list(results['rule_results'].keys())[0] if results['rule_results'] else None
        responsible_party_column = 'AuditLeader'  # Default
        
        if first_rule_id:
            # Try to get rule from manager first
            rule = self.rule_manager.get_rule(first_rule_id)
            if rule:
                responsible_party_column = rule.metadata.get('responsible_party_column', 'AuditLeader')
        
        # Section 1: IAG Overall Results (rows 3-10)
        current_row = self._generate_section1_iag_overall(ws, results, start_row=3)
        
        # Section 2: Audit Leader Results (2 rows below Section 1)
        current_row = self._generate_section2_leader_results(ws, results, 
                                                           responsible_party_column, 
                                                           start_row=current_row+2)
        
        # Section 3: Detailed Analytics (2 rows below Section 2)
        self._generate_section3_detailed_analytics(ws, results, 
                                                 responsible_party_column, 
                                                 start_row=current_row+2)
        
        # Apply column widths
        self._apply_column_widths(ws)
        
        # Generate individual test tabs
        for rule_id, rule_result in results['rule_results'].items():
            self._create_test_tab(wb, rule_id, rule_result, results)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")
    
    def _generate_section1_iag_overall(self, ws, results: Dict[str, Any], start_row: int = 3) -> int:
        """
        Generate Section 1: Executive Summary with severity-weighted IAG scoring.
        
        Args:
            ws: Worksheet object
            results: Validation results dictionary
            start_row: Starting row number
            
        Returns:
            Next available row number
        """
        from core.scoring.iag_scoring_calculator import IAGScoringCalculator
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        calculator = IAGScoringCalculator()
        
        # Debug logging
        logger.info(f"Starting _generate_section1_iag_overall with {len(results.get('rule_results', {}))} rules")
        
        # Prepare rule results organized by leader with risk levels
        rule_results_by_leader = {}
        for rule_id, rule_result in results.get('rule_results', {}).items():
            # Get rule's risk level
            rule = self.rule_manager.get_rule(rule_id)
            if rule:
                logger.debug(f"Rule {rule_id} found with metadata: {rule.metadata}")
            else:
                logger.warning(f"Rule {rule_id} not found in rule_manager")
            
            risk_level = rule.metadata.get('risk_level', 3) if rule else 3
            
            # Get party results if available
            if 'party_results' in rule_result:
                logger.debug(f"Rule {rule_id} has party_results with {len(rule_result['party_results'])} parties")
                for leader, leader_result in rule_result['party_results'].items():
                    if leader not in rule_results_by_leader:
                        rule_results_by_leader[leader] = []
                    rule_results_by_leader[leader].append({
                        'compliance_status': leader_result.get('status', 'NA'),
                        'risk_level': risk_level
                    })
            else:
                logger.warning(f"Rule {rule_id} missing party_results")
        
        logger.info(f"rule_results_by_leader has {len(rule_results_by_leader)} leaders")
        logger.debug(f"rule_results_by_leader structure: {list(rule_results_by_leader.keys())}")
        
        # Calculate severity-weighted score
        weighted_score, counts = calculator.calculate_severity_weighted_score(rule_results_by_leader)
        
        # Fallback if severity-weighted calculation returns N/A
        if weighted_score == "N/A" and 'grouped_summary' in results and results['grouped_summary']:
            logger.info("Severity-weighted calculation failed, falling back to standard IAG scoring")
            # Calculate standard IAG score from grouped_summary
            total_gc = sum(stats.get('GC', 0) for stats in results['grouped_summary'].values())
            total_pc = sum(stats.get('PC', 0) for stats in results['grouped_summary'].values())
            total_dnc = sum(stats.get('DNC', 0) for stats in results['grouped_summary'].values())
            total_na = sum(stats.get('NA', 0) for stats in results['grouped_summary'].values())
            
            total_applicable = total_gc + total_pc + total_dnc
            if total_applicable > 0:
                total_score = (total_gc * 5) + (total_pc * 3) + (total_dnc * 1)
                max_score = total_applicable * 5
                weighted_score = total_score / max_score
                logger.info(f"Fallback calculation: score={total_score}, max={max_score}, weighted={weighted_score}")
            else:
                weighted_score = "N/A"
        
        rating = calculator.assign_rating(weighted_score)
        
        # Get summary metrics
        total_rules = len(results.get('rule_results', {}))
        
        # Get total data points with multiple fallbacks
        total_entities = 0
        if 'population_summary' in results:
            total_entities = results['population_summary'].get('total_entities', 0)
            logger.debug(f"Got total_entities from population_summary: {total_entities}")
        
        if total_entities == 0 and 'data_metrics' in results:
            total_entities = results['data_metrics'].get('total_unique_entities', 0)
            logger.debug(f"Got total_entities from data_metrics.total_unique_entities: {total_entities}")
            
        if total_entities == 0 and 'data_metrics' in results:
            total_entities = results['data_metrics'].get('row_count', 0)
            logger.debug(f"Falling back to row_count: {total_entities}")
            
        # If still 0, try counting from grouped_summary
        if total_entities == 0 and 'grouped_summary' in results:
            # Sum all test counts across leaders
            for leader_stats in results['grouped_summary'].values():
                total_entities += leader_stats.get('total_rules', 0)
            logger.debug(f"Calculated total_entities from grouped_summary: {total_entities}")
        
        num_leaders = len(results.get('grouped_summary', {}))
        
        # Write header
        ws[f'A{start_row}'] = "IAG Overall Results and Rating"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        
        # Write executive metrics
        ws[f'A{start_row+1}'] = "Total Analytics Tested:"
        ws[f'B{start_row+1}'] = total_rules
        ws[f'B{start_row+1}'].alignment = Alignment(horizontal='center')
        
        ws[f'A{start_row+2}'] = "Total Data Points Reviewed:"
        ws[f'B{start_row+2}'] = total_entities
        ws[f'B{start_row+2}'].number_format = '#,##0'  # Number with thousands separator
        ws[f'B{start_row+2}'].alignment = Alignment(horizontal='center')
        
        ws[f'A{start_row+3}'] = "Number of Audit Leaders:"
        ws[f'B{start_row+3}'] = num_leaders
        ws[f'B{start_row+3}'].alignment = Alignment(horizontal='center')
        
        ws[f'A{start_row+4}'] = "Overall Compliance Rate:"
        if weighted_score != "N/A":
            ws[f'B{start_row+4}'] = weighted_score
            ws[f'B{start_row+4}'].number_format = '0.0%'  # Format as percentage with 1 decimal
        else:
            ws[f'B{start_row+4}'] = "N/A"
        ws[f'B{start_row+4}'].alignment = Alignment(horizontal='center')
        
        # Add explanation note
        if rule_results_by_leader:
            note_text = "Compliance rate calculated using IAG weighted scoring (GC=5, PC=3, DNC=1 points) with severity weighting (Critical=3x, High=2x, Medium/Low=1x)"
        else:
            note_text = "Compliance rate calculated using standard IAG weighted scoring (GC=5, PC=3, DNC=1 points)"
        
        ws[f'D{start_row+4}'] = note_text
        ws[f'D{start_row+4}'].font = Font(italic=True, size=9)
        ws[f'D{start_row+4}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'D{start_row+4}:H{start_row+5}')
        
        ws[f'A{start_row+5}'] = "Overall Rating:"
        ws[f'B{start_row+5}'] = rating
        ws[f'B{start_row+5}'].alignment = Alignment(horizontal='center')
        self._apply_rating_color(ws[f'B{start_row+5}'])
        
        # Override fields
        ws[f'A{start_row+6}'] = "Override Rating:"
        ws[f'B{start_row+6}'] = ""  # Blank for manual entry
        ws[f'B{start_row+6}'].alignment = Alignment(horizontal='center')
        ws[f'C{start_row+6}'] = "Rationale:"
        ws[f'D{start_row+6}'] = ""  # Blank for manual entry
        ws.merge_cells(f'D{start_row+6}:H{start_row+6}')
        
        # Summary sentence
        if weighted_score != "N/A":
            compliance_pct = f"{weighted_score * 100:.1f}%"
            scoring_type = "severity-weighted" if rule_results_by_leader else "IAG-weighted"
        else:
            compliance_pct = "N/A"
            scoring_type = ""
            
        summary_text = (f"Summary: Tested {total_rules} analytics across {num_leaders} audit leaders. "
                       f"The department achieved a {compliance_pct} {scoring_type} compliance rate, "
                       f"resulting in a \"{rating}\" rating. See Section 2 for individual audit leader "
                       f"performance and detailed test tabs for specific results.")
        
        ws[f'A{start_row+8}'] = summary_text
        ws[f'A{start_row+8}'].font = Font(size=11)
        ws[f'A{start_row+8}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{start_row+8}:H{start_row+9}')
        
        return start_row + 10
    
    def _apply_rating_color(self, cell):
        """Apply IAG-specific colors to rating cells."""
        from openpyxl.styles import PatternFill, Font
        
        if cell.value == 'GC':
            cell.fill = PatternFill("solid", start_color="90EE90")  # Light green
        elif cell.value == 'PC':
            cell.fill = PatternFill("solid", start_color="FFFF99")  # Light yellow
        elif cell.value == 'DNC':
            cell.fill = PatternFill("solid", start_color="FF6B6B")  # Light red
        elif cell.value in ['NA', 'N/A']:
            cell.fill = PatternFill("solid", start_color="D3D3D3")  # Gray
            cell.font = Font(italic=True)  # Italicize NA values
    
    def _apply_column_widths(self, ws):
        """Apply appropriate column widths for better readability."""
        from openpyxl.utils import get_column_letter
        
        # Set column widths based on content
        ws.column_dimensions['A'].width = 45  # Labels/Audit Leader names
        ws.column_dimensions['B'].width = 15  # Values
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
    
    def _generate_section2_leader_results(self, ws, results: Dict[str, Any], 
                                         responsible_party_column: str, start_row: int) -> int:
        """Generate Section 2: Audit Leader Overall Results and Ratings."""
        from core.scoring.iag_scoring_calculator import IAGScoringCalculator
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        calculator = IAGScoringCalculator()
        
        # Write section header
        ws[f'A{start_row}'] = "Audit Leader Overall Results and Ratings"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        # Column headers
        headers = ['Audit Leader', 'Total Tests', 'GC', 'PC', 'DNC', 'NA', 
                   'Compliance Rate', 'Rating', 'Override Rating', 'Rationale']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            # Add bottom border
            cell.border = Border(bottom=Side(style='thin'))
        
        current_row = start_row + 2
        
        # Get grouped summary data
        grouped_summary = results.get('grouped_summary', {})
        
        # Process each audit leader
        for leader_name in sorted(grouped_summary.keys()):
            leader_data = grouped_summary[leader_name]
            
            # Calculate IAG score for this leader
            gc_count = leader_data.get('GC', 0)
            pc_count = leader_data.get('PC', 0)
            dnc_count = leader_data.get('DNC', 0)
            na_count = leader_data.get('NA', 0)
            total_applicable = gc_count + pc_count + dnc_count
            
            if total_applicable > 0:
                score = (gc_count * 5 + pc_count * 3 + dnc_count * 1) / (total_applicable * 5)
            else:
                score = 0
            
            rating = calculator.assign_rating(score)
            
            # Write leader data
            ws[f'A{current_row}'] = leader_name
            ws[f'B{current_row}'] = leader_data.get('total_rules', 0)
            ws[f'C{current_row}'] = gc_count
            ws[f'D{current_row}'] = pc_count
            ws[f'E{current_row}'] = dnc_count
            ws[f'F{current_row}'] = na_count
            
            # Write compliance rate as a number, not text
            if total_applicable > 0:
                ws[f'G{current_row}'] = score
                ws[f'G{current_row}'].number_format = '0.0%'
            else:
                ws[f'G{current_row}'] = "N/A"
            
            ws[f'H{current_row}'] = rating
            ws[f'I{current_row}'] = ""  # Override Rating - blank for manual entry
            ws[f'J{current_row}'] = ""  # Rationale - blank for manual entry
            
            # Apply alternating row colors for readability (but not to rating column)
            if (current_row - start_row - 2) % 2 == 1:
                for col in range(1, 11):
                    if col != 8:  # Skip column H (rating column)
                        ws.cell(row=current_row, column=col).fill = PatternFill("solid", start_color="F0F0F0")
            
            # Apply rating color AFTER alternating rows so it doesn't get overwritten
            self._apply_rating_color(ws[f'H{current_row}'])
            
            current_row += 1
        
        # Add totals row
        current_row += 1
        ws[f'A{current_row}'] = "TOTALS"
        ws[f'A{current_row}'].font = Font(bold=True)
        
        # Calculate totals
        total_gc = sum(data.get('GC', 0) for data in grouped_summary.values())
        total_pc = sum(data.get('PC', 0) for data in grouped_summary.values())
        total_dnc = sum(data.get('DNC', 0) for data in grouped_summary.values())
        total_na = sum(data.get('NA', 0) for data in grouped_summary.values())
        total_tests = sum(data.get('total_rules', 0) for data in grouped_summary.values())
        
        ws[f'B{current_row}'] = total_tests
        ws[f'C{current_row}'] = total_gc
        ws[f'D{current_row}'] = total_pc
        ws[f'E{current_row}'] = total_dnc
        ws[f'F{current_row}'] = total_na
        
        # Add top border to totals row
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thin'))
        
        # Adjust column widths for Section 2
        ws.column_dimensions['I'].width = 15  # Override Rating
        ws.column_dimensions['J'].width = 30  # Rationale
        
        return current_row + 2
    
    def _generate_section3_detailed_analytics(self, ws, results: Dict[str, Any],
                                            responsible_party_column: str, start_row: int) -> None:
        """Generate Section 3: Detailed Analytics Section."""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.comments import Comment
        
        # Write section header
        ws[f'A{start_row}'] = "Detailed Analytics Section"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        # Create a matrix: Analytics (rows) x Audit Leaders (columns)
        grouped_summary = results.get('grouped_summary', {})
        rule_results = results.get('rule_results', {})
        
        # Get sorted lists
        leaders = sorted(grouped_summary.keys())
        rule_ids = sorted(rule_results.keys())
        
        # Column headers: Analytic ID, Name, Severity, then each leader
        current_row = start_row + 2
        ws[f'A{current_row}'] = "Analytic ID"
        ws[f'B{current_row}'] = "Analytic Name"
        ws[f'C{current_row}'] = "Severity"
        ws[f'D{current_row}'] = "Error Threshold"
        
        # Add leader names as column headers
        for col_idx, leader in enumerate(leaders, 5):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}{current_row}'] = leader
            ws[f'{col_letter}{current_row}'].alignment = Alignment(horizontal='center', text_rotation=45)
            ws.column_dimensions[col_letter].width = 12
        
        # Style the header row
        for col in range(1, len(leaders) + 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))
            if col >= 5:  # Leader columns
                cell.alignment = Alignment(horizontal='center', text_rotation=45)
        
        current_row += 1
        
        # Process each rule
        for rule_idx, rule_id in enumerate(rule_ids):
            rule_result = rule_results[rule_id]
            
            # Get rule metadata
            rule = self.rule_manager.get_rule(rule_id)
            using_fallback = False
            
            if rule:
                severity = rule.severity if hasattr(rule, 'severity') else 'medium'
                threshold = rule.threshold if hasattr(rule, 'threshold') else 0.0
                if not hasattr(rule, 'threshold'):
                    using_fallback = True
            else:
                severity = 'medium'
                threshold = 0.0  # Default error threshold - safer to show 0% than assume any errors are OK
                using_fallback = True
                logger.warning(f"Rule {rule_id} not found in rule manager, using fallback values")
            
            # Write rule info
            ws[f'A{current_row}'] = rule_id
            ws[f'B{current_row}'] = rule_result.get('rule_name', rule_id)
            ws[f'C{current_row}'] = severity.capitalize()
            ws[f'D{current_row}'] = threshold
            ws[f'D{current_row}'].number_format = '0%'  # Format as percentage
            
            # Highlight threshold cell if using fallback
            if using_fallback:
                ws[f'D{current_row}'].fill = PatternFill("solid", start_color="FFFF99")  # Yellow
                ws[f'D{current_row}'].comment = Comment("Using default value - rule metadata not found", "System")
            
            # Get party_results if available, otherwise calculate from grouped data
            party_results = rule_result.get('party_results', {})
            
            # Apply alternating row colors FIRST (skipping the leader status columns)
            if rule_idx % 2 == 1:
                for col in range(1, 5):  # Only columns A-D (rule info)
                    ws.cell(row=current_row, column=col).fill = PatternFill("solid", start_color="F0F0F0")
            
            # For each leader, determine the compliance status for this rule
            for col_idx, leader in enumerate(leaders, 5):
                col_letter = get_column_letter(col_idx)
                
                if party_results and leader in party_results:
                    # Use party_results if available
                    status = party_results[leader].get('status', 'NA')
                else:
                    # Fallback: use overall rule status for all leaders
                    # This is a simplification when party_results aren't available
                    status = rule_result.get('compliance_status', 'NA')
                
                ws[f'{col_letter}{current_row}'] = status
                ws[f'{col_letter}{current_row}'].alignment = Alignment(horizontal='center')
                
                # Apply color based on status (this will be the final color)
                cell = ws[f'{col_letter}{current_row}']
                self._apply_rating_color(cell)
            
            current_row += 1
        
        # Add note about color coding (moved up since we removed the totals row)
        current_row += 1
        ws[f'A{current_row}'] = "Note: GC = Generally Conforms (Green), PC = Partially Conforms (Yellow), DNC = Does Not Conform (Red), NA = Not Applicable (Gray)"
        ws[f'A{current_row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{current_row}:J{current_row}')
    
    def _extract_fields_from_formula(self, formula: str) -> List[str]:
        """Extract field names from Excel formula (e.g., [FieldName])"""
        import re
        # Find all bracketed field names in the formula
        fields = re.findall(r'\[([^\]]+)\]', formula)
        return list(set(fields))  # Remove duplicates
    
    def _create_test_tab(self, wb, rule_id: str, rule_result: Dict[str, Any], 
                        results: Dict[str, Any]) -> None:
        """Create individual test tab with summary and detailed results."""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Load rule metadata
        rule = self.rule_manager.get_rule(rule_id)
        if not rule:
            logger.warning(f"Rule {rule_id} not found, skipping test tab creation")
            return
            
        # Create worksheet with truncated title if necessary
        title = rule.name[:31] if len(rule.name) > 31 else rule.name
        ws = wb.create_sheet(title=title)
        
        # SECTION 1: Test Header (rows 1-5)
        ws['A1'] = "Test Name:"
        ws['B1'] = rule.metadata.get('title', rule.name)
        ws['A2'] = "Description:"
        ws['B2'] = rule.description
        ws['A3'] = "Risk Rating:"
        ws['B3'] = rule.severity.upper() if hasattr(rule, 'severity') else 'MEDIUM'
        ws['A4'] = "Population:"
        ws['B4'] = results.get('data_metrics', {}).get('row_count', 0)
        ws['B4'].number_format = '#,##0'
        ws['A5'] = "Error Threshold:"
        ws['B5'] = rule.threshold if hasattr(rule, 'threshold') else 0.1
        ws['B5'].number_format = '0%'
        
        # Apply header formatting
        for row in range(1, 6):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
        
        # SECTION 2: Leader Summary (starting row 7)
        ws['A7'] = "Audit Leader Summary"
        ws['A7'].font = Font(bold=True, size=12)
        
        # Leader summary column headers
        headers = ["Audit Leader", "Items Tested", "GC", "PC", "DNC", "NA", 
                   "Compliance Rate", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))
        
        # Process leader results from party_results if available
        current_row = 9
        party_results = rule_result.get('party_results', {})
        
        # Calculate counts from detailed results if available
        leader_counts = {}
        if '_result_details' in rule_result and rule_result['_result_details']:
            result_details_df = pd.DataFrame(rule_result['_result_details'])
            result_column = rule_result.get('_result_column', f"Result_{rule.name}")
            leader_column = rule.metadata.get('responsible_party_column', 'AuditLeader')
            
            if leader_column in result_details_df.columns and result_column in result_details_df.columns:
                for leader in result_details_df[leader_column].unique():
                    leader_df = result_details_df[result_details_df[leader_column] == leader]
                    
                    # Count statuses after boolean conversion
                    status_counts = {'GC': 0, 'PC': 0, 'DNC': 0, 'NA': 0}
                    for _, row in leader_df.iterrows():
                        raw_status = row.get(result_column, 'NA')
                        if isinstance(raw_status, bool) or str(raw_status).upper() in ['TRUE', 'FALSE']:
                            status = 'GC' if str(raw_status).upper() == 'TRUE' or raw_status is True else 'DNC'
                        else:
                            status = str(raw_status) if raw_status in ['GC', 'PC', 'DNC', 'NA'] else 'NA'
                        status_counts[status] = status_counts.get(status, 0) + 1
                    
                    leader_counts[leader] = {
                        'gc_count': status_counts['GC'],
                        'pc_count': status_counts['PC'],
                        'dnc_count': status_counts['DNC'],
                        'na_count': status_counts['NA'],
                        'total_count': len(leader_df)
                    }
        
        if party_results:
            for leader_name in sorted(party_results.keys()):
                leader_data = party_results[leader_name]
                
                ws[f'A{current_row}'] = leader_name
                
                # Get counts from leader data, with fallback to calculated counts
                if leader_name in leader_counts:
                    gc_count = leader_counts[leader_name]['gc_count']
                    pc_count = leader_counts[leader_name]['pc_count']
                    dnc_count = leader_counts[leader_name]['dnc_count']
                    na_count = leader_counts[leader_name]['na_count']
                    total_items = leader_counts[leader_name]['total_count']
                else:
                    gc_count = leader_data.get('gc_count', 0)
                    pc_count = leader_data.get('pc_count', 0)
                    dnc_count = leader_data.get('dnc_count', 0)
                    na_count = leader_data.get('na_count', 0)
                    total_items = leader_data.get('total_count', gc_count + pc_count + dnc_count + na_count)
                
                ws[f'B{current_row}'] = total_items
                ws[f'C{current_row}'] = gc_count
                ws[f'D{current_row}'] = pc_count
                ws[f'E{current_row}'] = dnc_count
                ws[f'F{current_row}'] = na_count
                
                # Calculate compliance rate (GC / (GC + PC + DNC))
                total_applicable = gc_count + pc_count + dnc_count
                if total_applicable > 0:
                    compliance_rate = gc_count / total_applicable
                    ws[f'G{current_row}'] = compliance_rate
                    ws[f'G{current_row}'].number_format = '0.0%'
                else:
                    ws[f'G{current_row}'] = "N/A"
                
                # Status
                status = leader_data.get('status', 'NA')
                ws[f'H{current_row}'] = status
                ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
                self._apply_rating_color(ws[f'H{current_row}'])
                
                current_row += 1
        else:
            # No party results - show overall summary
            ws[f'A{current_row}'] = "All Leaders (Combined)"
            ws[f'B{current_row}'] = rule_result.get('total_items', 0)
            ws[f'C{current_row}'] = rule_result.get('gc_count', 0)
            ws[f'D{current_row}'] = rule_result.get('pc_count', 0)
            ws[f'E{current_row}'] = rule_result.get('dnc_count', 0)
            ws[f'F{current_row}'] = 0  # NA count
            
            # Compliance rate
            if rule_result.get('total_items', 0) > 0:
                ws[f'G{current_row}'] = rule_result.get('compliance_rate', 0)
                ws[f'G{current_row}'].number_format = '0.0%'
            else:
                ws[f'G{current_row}'] = "N/A"
            
            ws[f'H{current_row}'] = rule_result.get('compliance_status', 'NA')
            self._apply_rating_color(ws[f'H{current_row}'])
            current_row += 1
        
        # SECTION 3: Detailed Results (starting ~row 15 or after summary)
        detail_start_row = current_row + 3
        ws[f'A{detail_start_row}'] = "Detailed Test Results (100% Population Coverage)"
        ws[f'A{detail_start_row}'].font = Font(bold=True, size=12)
        
        # Parse formula to determine which fields to display
        formula_fields = self._extract_fields_from_formula(rule.formula)
        
        # Check if formula contains LOOKUP and we have lookup operations
        lookup_columns = []
        if 'LOOKUP' in rule.formula and 'lookup_operations' in rule_result:
            # Extract unique lookup return columns from lookup operations
            lookup_ops = rule_result['lookup_operations']
            if lookup_ops:
                # Get unique return columns from lookup operations
                return_columns = set()
                for op in lookup_ops:
                    if 'return_column' in op:
                        return_columns.add(op['return_column'])
                lookup_columns = sorted(list(return_columns))
        
        # Detail section headers
        detail_headers = ["Item ID", "Audit Leader"] + formula_fields + lookup_columns + \
                        ["Status", "Failure Reason", "Internal Notes", "Audit Leader Response"]
        
        header_row = detail_start_row + 1
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))
        
        # Get detailed results if available
        detail_row = header_row + 1
        
        # Check if we have detailed results stored
        if '_result_details' in rule_result and rule_result['_result_details']:
            # Convert back to DataFrame for processing
            result_df = pd.DataFrame(rule_result['_result_details'])
            result_column = rule_result.get('_result_column', f"Result_{rule.name}")
            
            # Sort by status: DNC first, then PC, then GC
            status_order = {'DNC': 0, 'PC': 1, 'GC': 2, 'NA': 3}
            if result_column in result_df.columns:
                result_df['_sort_order'] = result_df[result_column].map(status_order).fillna(4)
                result_df = result_df.sort_values('_sort_order')
                
                # Write each row
                for df_idx, (idx, row) in enumerate(result_df.iterrows()):
                    col_idx = 1
                    # Item ID
                    ws.cell(row=detail_row, column=col_idx, value=row.get('AuditEntityID', idx))
                    col_idx += 1
                    
                    # Audit Leader
                    ws.cell(row=detail_row, column=col_idx, value=row.get(rule.metadata.get('responsible_party_column', 'AuditLeader'), ''))
                    col_idx += 1
                    
                    # Formula field values
                    for field in formula_fields:
                        field_value = row.get(field, '')
                        # Handle NaT (Not a Time) values
                        if pd.isna(field_value) or str(field_value) == 'NaT':
                            field_value = ''
                        ws.cell(row=detail_row, column=col_idx, value=field_value)
                        col_idx += 1
                    
                    # Lookup values
                    if lookup_columns and 'lookup_operations' in rule_result:
                        # Create a map of lookup values for this row
                        lookup_values = {}
                        # Get the original row index before sorting
                        original_idx = result_df.index.tolist().index(idx)
                        
                        for op in rule_result['lookup_operations']:
                            if op.get('row_index') == original_idx and op.get('success'):
                                lookup_values[op['return_column']] = op.get('result', '')
                        
                        # Add lookup values in the same order as headers
                        for lookup_col in lookup_columns:
                            value = lookup_values.get(lookup_col, '')
                            ws.cell(row=detail_row, column=col_idx, value=value)
                            col_idx += 1
                    
                    # Status - convert boolean to GC/DNC
                    raw_status = row.get(result_column, 'NA')
                    if isinstance(raw_status, bool) or str(raw_status).upper() in ['TRUE', 'FALSE']:
                        # Convert boolean to compliance status
                        if str(raw_status).upper() == 'TRUE' or raw_status is True:
                            status = 'GC'
                        else:
                            status = 'DNC'
                    else:
                        # Already a compliance status or NA
                        status = raw_status
                    
                    status_cell = ws.cell(row=detail_row, column=col_idx, value=status)
                    self._apply_rating_color(status_cell)
                    status_cell.alignment = Alignment(horizontal='center')
                    col_idx += 1
                    
                    # Failure reason (populate for DNC/PC items)
                    if status in ['DNC', 'PC']:
                        ws.cell(row=detail_row, column=col_idx, 
                               value=rule.metadata.get('error_message', 'Does not meet requirements'))
                    col_idx += 1
                    
                    # Leave Internal Notes and Audit Leader Response blank
                    # These are for manual entry
                    
                    detail_row += 1
                    
                # Clean up sort column
                if '_sort_order' in result_df.columns:
                    result_df.drop('_sort_order', axis=1, inplace=True)
        else:
            # No detailed results available yet
            ws[f'A{detail_row}'] = "Note: Detailed item-level results will be available when result_df is stored during validation"
            ws[f'A{detail_row}'].font = Font(italic=True, color="666666")
            ws.merge_cells(f'A{detail_row}:D{detail_row}')
        
        # Apply column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        for col in range(3, len(detail_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_guide_tab_content(self, ws):
        """Add guide content to the provided worksheet."""
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Title
        ws['A1'] = "HOW TO READ THIS REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Status Meanings Section
        row = 3
        ws[f'A{row}'] = "STATUS MEANINGS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        status_meanings = [
            ("GC (Generally Conforms)", "PASS", "The item met all validation criteria", "90EE90"),
            ("PC (Partially Conforms)", "PARTIAL", "The item had minor issues", "FFFF99"),
            ("DNC (Does Not Conform)", "FAIL", "The item failed validation criteria", "FF6B6B"),
            ("NA (Not Applicable)", "N/A", "The test didn't apply to this item", "D3D3D3")
        ]
        
        for status, result, description, color in status_meanings:
            ws[f'A{row}'] = status
            ws[f'B{row}'] = f"({result})"  # Changed from "= {result}" to avoid Excel formula
            ws[f'C{row}'] = f"- {description}"
            ws[f'A{row}'].fill = PatternFill("solid", start_color=color)
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'C{row}:F{row}')
            row += 1
        
        # Report Sections
        row += 2
        ws[f'A{row}'] = "REPORT SECTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        ws[f'A{row}'] = "1. IAG Summary Report Tab"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        sections = [
            "   • Section 1: Overall department compliance score",
            "   • Section 2: Individual audit leader scores",
            "   • Section 3: Matrix showing which leaders passed which tests"
        ]
        
        for section in sections:
            ws[f'A{row}'] = section
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        row += 1
        ws[f'A{row}'] = "2. Individual Test Tabs"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        tab_sections = [
            "   • Header: What this test checks",
            "   • Summary: How each leader performed",
            "   • Details: Every item tested (failures shown first)",
            "   • Response Columns: For documenting remediation actions"
        ]
        
        for section in tab_sections:
            ws[f'A{row}'] = section
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # What to Do
        row += 2
        ws[f'A{row}'] = "WHAT TO DO"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        actions = [
            "1. Review your compliance score in Section 2 of the IAG Summary",
            "2. Check individual test tabs for your failures (sorted to top)",
            "3. Document your remediation plan in the \"Audit Leader Response\" column",
            "4. Save and return the file by [due date]"
        ]
        
        for action in actions:
            ws[f'A{row}'] = action
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # How Compliance Scores Are Calculated
        row += 2
        ws[f'A{row}'] = "HOW COMPLIANCE SCORES ARE CALCULATED"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        ws[f'A{row}'] = "Basic Scoring:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Each test result receives points based on compliance level:"
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        scoring_rules = [
            "• GC (Generally Conforms/Pass) = 5 points",
            "• PC (Partially Conforms) = 3 points",
            "• DNC (Does Not Conform/Fail) = 1 point",
            "• NA (Not Applicable) = 0 points"
        ]
        
        for rule in scoring_rules:
            ws[f'A{row}'] = rule
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        row += 1
        ws[f'A{row}'] = "Compliance Rate = Total Points Earned ÷ Maximum Possible Points"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        # Example
        row += 2
        ws[f'A{row}'] = "Example:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        example_text = [
            "If you have 3 tests with results: GC, PC, DNC",
            "• Points earned: 5 + 3 + 1 = 9 points",
            "• Maximum possible: 3 tests × 5 points = 15 points",
            "• Compliance rate: 9 ÷ 15 = 60%"
        ]
        
        for text in example_text:
            ws[f'A{row}'] = text
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Risk-Based Weighting
        row += 2
        ws[f'A{row}'] = "Risk-Based Weighting:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Critical tests count more than routine tests:"
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        weighting_rules = [
            "• Critical Risk Tests: Count 3x (triple weight)",
            "• High Risk Tests: Count 2x (double weight)",
            "• Medium/Low Risk Tests: Count 1x (normal weight)"
        ]
        
        for rule in weighting_rules:
            ws[f'A{row}'] = rule
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Weighted Example
        row += 2
        ws[f'A{row}'] = "Example with Risk Weighting:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        weighted_example = [
            "• Critical test (DNC): 1 point × 3 = 3 weighted points",
            "• High test (GC): 5 points × 2 = 10 weighted points",
            "• Medium test (GC): 5 points × 1 = 5 weighted points",
            "Total: 18 points out of possible 40 = 45%"
        ]
        
        for text in weighted_example:
            ws[f'A{row}'] = text
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Overall Ratings
        row += 2
        ws[f'A{row}'] = "Overall Ratings:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        rating_thresholds = [
            "• GC Rating: 80% or higher",
            "• PC Rating: 50% to 79%",
            "• DNC Rating: Below 50%"
        ]
        
        for threshold in rating_thresholds:
            ws[f'A{row}'] = threshold
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Set column widths for readability
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
    
    def _create_guide_tab(self, wb):
        """Create a Guide tab with instructions on how to read the report."""
        ws = wb.create_sheet(title="Guide")
        self._create_guide_tab_content(ws)
    
    def split_report_by_leader(self, master_file_path: str, output_dir: Optional[str] = None) -> Dict[str, str]:
        """
        Split master Excel report into individual files for each audit leader.
        
        Args:
            master_file_path: Path to the master Excel report
            output_dir: Directory to save leader files (defaults to same dir as master)
            
        Returns:
            Dictionary mapping leader names to their file paths
        """
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime
        import os
        
        # Load the master workbook
        logger.info(f"Loading master report from {master_file_path}")
        master_wb = load_workbook(master_file_path, data_only=False)
        
        # Determine output directory
        if output_dir is None:
            output_dir = os.path.dirname(master_file_path)
        os.makedirs(output_dir, exist_ok=True)
        
        # Get list of audit leaders from Section 2 of IAG Summary Report
        summary_ws = master_wb['IAG Summary Report']
        leaders = []
        
        # Find Section 2 header
        section2_row = None
        for row in range(1, 50):
            if summary_ws[f'A{row}'].value == "Audit Leader Overall Results and Ratings":
                section2_row = row
                break
        
        if not section2_row:
            raise ValueError("Could not find Section 2 in IAG Summary Report")
        
        # Find Section 3 to know where Section 2 ends
        section3_row = None
        for row in range(section2_row + 1, summary_ws.max_row + 1):
            if summary_ws[f'A{row}'].value == "Detailed Analytics Section":
                section3_row = row
                break
        
        # Extract leader names (skip header row and totals row)
        leader_start_row = section2_row + 2
        end_row = section3_row - 1 if section3_row else summary_ws.max_row
        
        for row in range(leader_start_row, end_row):
            leader_name = summary_ws[f'A{row}'].value
            if leader_name and leader_name not in ["Totals", "TOTALS", None, ""]:
                # Additional check - make sure this looks like a leader row by checking if there's data
                if summary_ws[f'B{row}'].value is not None:  # Total Tests column should have data
                    leaders.append(leader_name)
        
        logger.info(f"Found {len(leaders)} audit leaders: {leaders}")
        
        # Create timestamp for filenames
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create individual files for each leader
        leader_files = {}
        
        for leader in leaders:
            logger.info(f"Creating report for {leader}")
            
            # Create new workbook for this leader
            leader_wb = Workbook()
            
            # Remove default sheet (we'll create Guide as the first)
            default_sheet = leader_wb.active
            
            # 1. Copy Guide tab as first tab
            self._copy_guide_tab(master_wb, leader_wb)
            
            # Now remove the default sheet
            leader_wb.remove(default_sheet)
            
            # 2. Copy IAG Summary Report with filtered data
            self._copy_iag_summary_for_leader(master_wb, leader_wb, leader)
            
            # 3. Copy only test tabs where this leader has results
            self._copy_test_tabs_for_leader(master_wb, leader_wb, leader)
            
            # Save the leader's file
            safe_leader_name = leader.replace(' ', '_').replace('/', '_')
            filename = f"QA_Results_{safe_leader_name}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            leader_wb.save(filepath)
            leader_files[leader] = filepath
            logger.info(f"Saved report for {leader} to {filepath}")
        
        return leader_files
    
    def _copy_iag_summary_for_leader(self, master_wb, leader_wb, leader: str):
        """Copy IAG Summary Report with only this leader's data."""
        from openpyxl.utils import get_column_letter
        from copy import copy
        
        master_ws = master_wb['IAG Summary Report']
        leader_ws = leader_wb.create_sheet('IAG Summary Report')
        
        # Find key sections
        section1_row = section2_row = section3_row = None
        for row in range(1, 100):
            cell_value = master_ws[f'A{row}'].value
            if cell_value == "IAG Overall Results and Rating":
                section1_row = row
            elif cell_value == "Audit Leader Overall Results and Ratings":
                section2_row = row
            elif cell_value == "Detailed Analytics Section":
                section3_row = row
        
        # Copy Section 1 as-is (overall department data)
        if section1_row and section2_row:
            self._copy_range(master_ws, leader_ws, 1, section2_row - 1)
        
        # Copy Section 2 header and only this leader's row
        if section2_row and section3_row:
            # Copy header rows
            self._copy_range(master_ws, leader_ws, section2_row, section2_row + 1)
            
            # Find and copy only this leader's row
            for row in range(section2_row + 2, section3_row):
                if master_ws[f'A{row}'].value == leader:
                    # Calculate destination row (header + 1 data row)
                    dest_row = section2_row + 2
                    self._copy_row(master_ws, leader_ws, row, dest_row)
                    break
        
        # Copy Section 3 with only this leader's column
        if section3_row:
            # First find which column belongs to this leader
            leader_col = None
            header_row = section3_row + 1
            
            for col in range(2, master_ws.max_column + 1):
                if master_ws.cell(row=header_row, column=col).value == leader:
                    leader_col = col
                    break
            
            if leader_col:
                # Copy section header
                self._copy_row(master_ws, leader_ws, section3_row, section3_row)
                
                # Copy header row with only first column and leader's column
                dest_row = section3_row + 1
                # Copy row headers (column A)
                self._copy_cell(master_ws, leader_ws, dest_row, 1, dest_row, 1)
                # Copy leader's header
                self._copy_cell(master_ws, leader_ws, dest_row, leader_col, dest_row, 2)
                
                # Copy data rows with only the leader's data
                for row in range(section3_row + 2, master_ws.max_row + 1):
                    if master_ws[f'A{row}'].value:  # If there's a row header
                        # Copy the row header
                        self._copy_cell(master_ws, leader_ws, row, 1, row, 1)
                        # Copy the leader's data
                        self._copy_cell(master_ws, leader_ws, row, leader_col, row, 2)
        
        # Apply column widths
        self._apply_column_widths(leader_ws)
    
    def _copy_test_tabs_for_leader(self, master_wb, leader_wb, leader: str):
        """Copy only test tabs where this leader has results."""
        from openpyxl.utils import get_column_letter
        
        # Get all sheet names except IAG Summary and Guide
        test_sheets = [name for name in master_wb.sheetnames 
                      if name not in ['IAG Summary Report', 'Guide']]
        
        for sheet_name in test_sheets:
            master_ws = master_wb[sheet_name]
            
            # Check if this leader has data in this test
            has_leader_data = False
            
            # Look for leader in the Audit Leader Summary section
            for row in range(1, min(30, master_ws.max_row + 1)):
                if master_ws[f'A{row}'].value == "Audit Leader Summary":
                    # Check the data rows below
                    for data_row in range(row + 2, min(row + 20, master_ws.max_row + 1)):
                        if master_ws[f'A{data_row}'].value == leader:
                            has_leader_data = True
                            break
                    break
            
            if has_leader_data:
                logger.debug(f"Including test tab '{sheet_name}' for {leader}")
                leader_ws = leader_wb.create_sheet(sheet_name)
                
                # Copy header section (rows 1-5)
                self._copy_range(master_ws, leader_ws, 1, 5)
                
                # Find and copy Audit Leader Summary section with only this leader
                for row in range(6, 30):
                    if master_ws[f'A{row}'].value == "Audit Leader Summary":
                        # Copy section header and column headers
                        self._copy_range(master_ws, leader_ws, row, row + 1)
                        
                        # Find and copy only this leader's row
                        for data_row in range(row + 2, row + 20):
                            if master_ws[f'A{data_row}'].value == leader:
                                dest_row = row + 2
                                self._copy_row(master_ws, leader_ws, data_row, dest_row)
                                break
                        
                        # Find detailed results section
                        detail_start = None
                        for detail_row in range(row + 3, master_ws.max_row + 1):
                            cell_val = master_ws[f'A{detail_row}'].value
                            if cell_val and "Detailed Test Results" in str(cell_val):
                                detail_start = detail_row
                                break
                        
                        if detail_start:
                            # Copy detail section header
                            self._copy_row(master_ws, leader_ws, detail_start, detail_start)
                            self._copy_row(master_ws, leader_ws, detail_start + 1, detail_start + 1)
                            
                            # Copy only this leader's detailed results
                            dest_detail_row = detail_start + 2
                            for src_row in range(detail_start + 2, master_ws.max_row + 1):
                                # Check if this row belongs to the leader (column B)
                                if master_ws[f'B{src_row}'].value == leader:
                                    self._copy_row(master_ws, leader_ws, src_row, dest_detail_row)
                                    dest_detail_row += 1
                        break
                
                # Apply column widths
                leader_ws.column_dimensions['A'].width = 20
                leader_ws.column_dimensions['B'].width = 25
                for col in range(3, 10):
                    leader_ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _copy_guide_tab(self, master_wb, leader_wb):
        """Copy the Guide tab as-is."""
        if 'Guide' in master_wb.sheetnames:
            master_ws = master_wb['Guide']
            leader_ws = leader_wb.create_sheet('Guide')
            
            # Copy all content
            for row in range(1, master_ws.max_row + 1):
                for col in range(1, master_ws.max_column + 1):
                    self._copy_cell(master_ws, leader_ws, row, col, row, col)
            
            # Copy column widths
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                if col_letter in master_ws.column_dimensions:
                    leader_ws.column_dimensions[col_letter].width = master_ws.column_dimensions[col_letter].width
    
    def _copy_range(self, source_ws, dest_ws, start_row: int, end_row: int):
        """Copy a range of rows from source to destination worksheet."""
        for row in range(start_row, end_row + 1):
            self._copy_row(source_ws, dest_ws, row, row)
    
    def _copy_row(self, source_ws, dest_ws, source_row: int, dest_row: int):
        """Copy entire row from source to destination worksheet."""
        for col in range(1, source_ws.max_column + 1):
            self._copy_cell(source_ws, dest_ws, source_row, col, dest_row, col)
    
    def _copy_cell(self, source_ws, dest_ws, source_row: int, source_col: int, 
                   dest_row: int, dest_col: int):
        """Copy a single cell with value and formatting."""
        from copy import copy
        from openpyxl.utils import get_column_letter
        
        source_cell = source_ws.cell(row=source_row, column=source_col)
        dest_cell = dest_ws.cell(row=dest_row, column=dest_col)
        
        # Check if this cell is part of a merged range in the source
        for merged_range in source_ws.merged_cells.ranges:
            if source_cell.coordinate in merged_range:
                # This cell is part of a merged range
                if source_row == merged_range.min_row and source_col == merged_range.min_col:
                    # This is the top-left cell of the merge - copy value and create merge
                    dest_cell.value = source_cell.value
                    
                    # Calculate destination merge range
                    dest_min_row = dest_row
                    dest_max_row = dest_row + (merged_range.max_row - merged_range.min_row)
                    dest_min_col = dest_col
                    dest_max_col = dest_col + (merged_range.max_col - merged_range.min_col)
                    
                    merge_range = f"{get_column_letter(dest_min_col)}{dest_min_row}:{get_column_letter(dest_max_col)}{dest_max_row}"
                    try:
                        dest_ws.merge_cells(merge_range)
                    except:
                        pass  # Already merged
                # For other cells in the merge, don't copy value (handled by merge)
                break
        else:
            # Not part of a merged range - copy value normally
            try:
                dest_cell.value = source_cell.value
            except AttributeError:
                # In case it's still a MergedCell, skip value copy
                pass
        
        # Copy formatting (always copy, even for merged cells)
        if source_cell.has_style:
            dest_cell.font = copy(source_cell.font)
            dest_cell.fill = copy(source_cell.fill)
            dest_cell.border = copy(source_cell.border)
            dest_cell.alignment = copy(source_cell.alignment)
            dest_cell.number_format = source_cell.number_format
    
    def generate_and_split_reports(self, validation_results_path: str, 
                                  output_dir: Optional[str] = None,
                                  master_filename: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate master report and split into individual leader files in one operation.
        
        Args:
            validation_results_path: Path to validation results JSON file
            output_dir: Directory for all output files (defaults to output/)
            master_filename: Name for master file (defaults to QA_Master_Report_[timestamp].xlsx)
            
        Returns:
            Dictionary with:
                - master_path: Path to master report
                - leader_files: Dict mapping leader names to their file paths
        """
        import datetime
        import os
        
        # Set output directory
        if output_dir is None:
            output_dir = str(self.output_dir)
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate master report
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if master_filename is None:
            master_filename = f"QA_Master_Report_{timestamp}.xlsx"
        
        master_path = os.path.join(output_dir, master_filename)
        self.generate_excel_report(validation_results_path, master_path)
        
        # Create subdirectory for leader files
        leader_dir = os.path.join(output_dir, f"leader_reports_{timestamp}")
        os.makedirs(leader_dir, exist_ok=True)
        
        # Split into leader files
        leader_files = self.split_report_by_leader(master_path, leader_dir)
        
        logger.info(f"Master report: {master_path}")
        logger.info(f"Leader files: {leader_dir}")
        
        return {
            'master_path': master_path,
            'leader_files': leader_files,
            'output_dir': output_dir,
            'leader_dir': leader_dir,
            'timestamp': timestamp
        }
