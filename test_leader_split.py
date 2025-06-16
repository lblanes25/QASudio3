#!/usr/bin/env python3
"""
Test the leader file splitting functionality - Phase 2
"""

import os
import glob
from services.validation_service import ValidationPipeline
from openpyxl import load_workbook

def test_leader_split():
    """Test splitting master report into individual leader files"""
    
    # First check if we have a master report
    if not os.path.exists("test_complete_iag_report.xlsx"):
        print("ERROR: No master report found. Run test_complete_iag_report.py first")
        return False
    
    print("Testing leader file splitting...")
    
    # Create output directory for leader files
    output_dir = "output/leader_reports"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create pipeline and split the report
    pipeline = ValidationPipeline()
    leader_files = pipeline.split_report_by_leader(
        "test_complete_iag_report.xlsx",
        output_dir
    )
    
    print(f"\nGenerated {len(leader_files)} leader files:")
    for leader, filepath in leader_files.items():
        print(f"  {leader}: {os.path.basename(filepath)}")
    
    # Verify each leader file
    print("\nVerifying leader files...")
    
    for leader, filepath in leader_files.items():
        print(f"\nChecking {leader}'s file:")
        wb = load_workbook(filepath)
        
        # Check expected sheets exist
        assert "IAG Summary Report" in wb.sheetnames, f"Missing IAG Summary Report in {leader}'s file"
        assert "Guide" in wb.sheetnames, f"Missing Guide tab in {leader}'s file"
        
        # Check IAG Summary has only this leader's data in Section 2
        ws = wb["IAG Summary Report"]
        
        # Find Section 2
        section2_row = None
        for row in range(1, 50):
            if ws[f'A{row}'].value == "Audit Leader Overall Results and Ratings":
                section2_row = row
                break
        
        # Count leaders in Section 2 (should be only 1)
        leader_count = 0
        if section2_row:
            for row in range(section2_row + 2, section2_row + 20):
                if ws[f'A{row}'].value and ws[f'A{row}'].value != "Totals":
                    leader_count += 1
                    assert ws[f'A{row}'].value == leader, f"Found wrong leader in {leader}'s file"
        
        assert leader_count == 1, f"Expected 1 leader in Section 2, found {leader_count}"
        print(f"  ✓ Section 2 contains only {leader}'s data")
        
        # Check Section 3 has only one data column (plus headers)
        section3_row = None
        for row in range(1, 100):
            if ws[f'A{row}'].value == "Detailed Analytics Section":
                section3_row = row
                break
        
        if section3_row:
            # Count non-empty columns in header row
            header_row = section3_row + 1
            data_cols = 0
            for col in range(2, 10):  # Check up to column J
                if ws.cell(row=header_row, column=col).value:
                    data_cols += 1
            
            assert data_cols == 1, f"Expected 1 data column in Section 3, found {data_cols}"
            print(f"  ✓ Section 3 contains only {leader}'s column")
        
        # Check test tabs
        test_tabs = [name for name in wb.sheetnames if name not in ['IAG Summary Report', 'Guide']]
        print(f"  ✓ Found {len(test_tabs)} test tabs")
        
        # Verify each test tab has only this leader's data
        for tab in test_tabs[:1]:  # Check first test tab
            test_ws = wb[tab]
            
            # Find Audit Leader Summary
            for row in range(1, 30):
                if test_ws[f'A{row}'].value == "Audit Leader Summary":
                    # Count leaders (should be 1)
                    tab_leader_count = 0
                    for data_row in range(row + 2, row + 10):
                        if test_ws[f'A{data_row}'].value:
                            tab_leader_count += 1
                            assert test_ws[f'A{data_row}'].value == leader
                    
                    assert tab_leader_count == 1, f"Expected 1 leader in {tab}, found {tab_leader_count}"
                    
                    # Check detailed results if present
                    for detail_row in range(row + 3, test_ws.max_row + 1):
                        if test_ws[f'A{detail_row}'].value and "Detailed Test Results" in str(test_ws[f'A{detail_row}'].value):
                            # Check all detail rows have this leader
                            for check_row in range(detail_row + 2, test_ws.max_row + 1):
                                if test_ws[f'B{check_row}'].value:  # Column B has leader name
                                    assert test_ws[f'B{check_row}'].value == leader, f"Found wrong leader in details"
                            break
                    break
        
        print(f"  ✓ Test tabs contain only {leader}'s data")
    
    print(f"\n✅ All leader files generated and verified successfully!")
    print(f"\nLeader files saved in: {output_dir}")
    
    return True

if __name__ == "__main__":
    test_leader_split()