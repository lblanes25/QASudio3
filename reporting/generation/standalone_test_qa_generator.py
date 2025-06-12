"""
Standalone Test for Dynamic QA Report Generator
Creates test data and generates a sample report to see the actual output
"""

import sys
import os
from pathlib import Path

# Add the current directory to Python path so we can import the generator
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

# Import the generator
try:
    from dynamic_qa_report_generator import DynamicQAReportGenerator
    print("✓ Successfully imported DynamicQAReportGenerator")
except ImportError as e:
    print(f"✗ Failed to import DynamicQAReportGenerator: {e}")
    sys.exit(1)

def create_mock_rule_results():
    """Create realistic mock data that matches the expected structure"""
    
    class MockRule:
        def __init__(self, rule_id, name, threshold=0.05):
            self.rule_id = rule_id
            self.name = name
            self.description = f"Validation analysis for {name}"
            self.formula = f"=NOT(ISBLANK([DateActivityOccurred]))"
            self.threshold = threshold
    
    class MockRuleResult:
        def __init__(self, rule_id, name, threshold=0.05):
            self.rule = MockRule(rule_id, name, threshold)
            
            # Create compliance metrics that match expected structure
            self.compliance_metrics = {
                'compliant_count': 85,
                'partially_compliant_count': 10,
                'non_compliant_count': 5,
                'not_applicable_count': 0,
                'total_count': 100
            }
            
            # Create party results with audit leader breakdown
            self.party_results = {
                'Alice Johnson': {
                    'status': 'GC',
                    'metrics': {
                        'gc_count': 45,
                        'pc_count': 3,
                        'dnc_count': 2,
                        'na_count': 0,
                        'total_count': 50,
                        'dnc_rate': 0.04
                    }
                },
                'Bob Smith': {
                    'status': 'PC', 
                    'metrics': {
                        'gc_count': 25,
                        'pc_count': 5,
                        'dnc_count': 20,
                        'na_count': 0,
                        'total_count': 50,
                        'dnc_rate': 0.50
                    }
                },
                'Carol Davis': {
                    'status': 'DNC',
                    'metrics': {
                        'gc_count': 15,
                        'pc_count': 2,
                        'dnc_count': 33,
                        'na_count': 0,
                        'total_count': 50,
                        'dnc_rate': 0.70
                    }
                }
            }
            
            # Set compliance status based on overall metrics
            total_applicable = (self.compliance_metrics['compliant_count'] + 
                              self.compliance_metrics['partially_compliant_count'] + 
                              self.compliance_metrics['non_compliant_count'])
            
            if total_applicable > 0:
                error_rate = ((self.compliance_metrics['partially_compliant_count'] + 
                             self.compliance_metrics['non_compliant_count']) / total_applicable)
                
                if error_rate <= threshold:
                    self.compliance_status = 'GC'
                elif error_rate <= threshold * 2:
                    self.compliance_status = 'PC'
                else:
                    self.compliance_status = 'DNC'
            else:
                self.compliance_status = 'N/A'
    
    # Create multiple rules with varied results
    rule_results = {
        'QA-ID-1': MockRuleResult('QA-ID-1', 'Date Activity Required', 0.05),
        'QA-ID-2': MockRuleResult('QA-ID-2', 'Submitter Reviewer Separation', 0.02),
        'QA-ID-3': MockRuleResult('QA-ID-3', 'DDAP Utilization Check', 0.10),
        'QA-ID-4': MockRuleResult('QA-ID-4', 'Impact Documentation', 0.03),
        'QA-ID-5': MockRuleResult('QA-ID-5', 'Approval Workflow', 0.05)
    }
    
    # Vary the results to show different compliance levels
    rule_results['QA-ID-2'].compliance_metrics.update({
        'compliant_count': 95,
        'partially_compliant_count': 3,
        'non_compliant_count': 2,
        'not_applicable_count': 0
    })
    rule_results['QA-ID-2'].compliance_status = 'GC'
    
    rule_results['QA-ID-3'].compliance_metrics.update({
        'compliant_count': 60,
        'partially_compliant_count': 25,
        'non_compliant_count': 15,
        'not_applicable_count': 0
    })
    rule_results['QA-ID-3'].compliance_status = 'PC'
    
    rule_results['QA-ID-4'].compliance_metrics.update({
        'compliant_count': 45,
        'partially_compliant_count': 15,
        'non_compliant_count': 40,
        'not_applicable_count': 0
    })
    rule_results['QA-ID-4'].compliance_status = 'DNC'
    
    print(f"✓ Created {len(rule_results)} mock rule results")
    return rule_results

def test_basic_functionality():
    """Test basic functionality of the generator"""
    print("\n" + "="*60)
    print("TESTING BASIC FUNCTIONALITY")
    print("="*60)
    
    try:
        # Create generator instance
        generator = DynamicQAReportGenerator()
        print("✓ Successfully created DynamicQAReportGenerator instance")
        
        # Create mock data
        rule_results = create_mock_rule_results()
        
        # Test data extraction
        rule_names, audit_leaders, leader_rule_matrix = generator._extract_dynamic_structure(
            rule_results, "Audit Leader"
        )
        
        print(f"✓ Extracted structure:")
        print(f"  - Rules: {rule_names}")
        print(f"  - Audit Leaders: {audit_leaders}")
        print(f"  - Leader-Rule Matrix keys: {list(leader_rule_matrix.keys())}")
        
        # Test if we have the expected structure
        assert len(rule_names) == 5, f"Expected 5 rules, got {len(rule_names)}"
        assert len(audit_leaders) == 3, f"Expected 3 leaders, got {len(audit_leaders)}"
        assert 'Alice Johnson' in audit_leaders, "Alice Johnson not found in audit leaders"
        
        print("✓ Basic data structure validation passed")
        
    except Exception as e:
        print(f"✗ Basic functionality test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

def test_report_generation():
    """Test actual report generation"""
    print("\n" + "="*60)
    print("TESTING REPORT GENERATION")
    print("="*60)
    
    try:
        # Create generator and mock data
        generator = DynamicQAReportGenerator()
        rule_results = create_mock_rule_results()
        
        # Generate report
        output_path = "test_dynamic_qa_report.xlsx"
        
        print(f"Generating report to: {output_path}")
        
        result_path = generator.generate_report(
            rule_results=rule_results,
            output_path=output_path,
            responsible_party_column="Audit Leader",
            review_year="2025 Test Report"
        )
        
        print(f"✓ Successfully generated report: {result_path}")
        
        # Check if file exists and has reasonable size
        if os.path.exists(result_path):
            file_size = os.path.getsize(result_path)
            print(f"✓ Report file exists, size: {file_size:,} bytes")
            
            if file_size > 10000:  # Should be at least 10KB for a real Excel file
                print("✓ File size looks reasonable")
            else:
                print(f"⚠ File size seems small ({file_size} bytes)")
        else:
            print("✗ Report file was not created")
            return False
            
    except Exception as e:
        print(f"✗ Report generation test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

def test_style_application():
    """Test that styles are being applied correctly"""
    print("\n" + "="*60)
    print("TESTING STYLE APPLICATION")
    print("="*60)
    
    try:
        generator = DynamicQAReportGenerator()
        
        # Test style setup
        print(f"✓ Header font: {generator.header_font}")
        print(f"✓ Colors defined: green={generator.green_fill.start_color}, red={generator.red_fill.start_color}")
        print(f"✓ Borders defined: {generator.thin_border}")
        print(f"✓ Number formats: {generator.number_format}, {generator.percentage_format}")
        
        print("✓ All styles properly initialized")
        
    except Exception as e:
        print(f"✗ Style application test failed: {e}")
        return False
    
    return True

def inspect_generated_file():
    """Inspect the generated Excel file"""
    print("\n" + "="*60)
    print("INSPECTING GENERATED FILE")
    print("="*60)
    
    output_path = "test_dynamic_qa_report.xlsx"
    
    if not os.path.exists(output_path):
        print(f"✗ File {output_path} not found")
        return False
    
    try:
        import openpyxl
        
        # Load the workbook
        wb = openpyxl.load_workbook(output_path)
        print(f"✓ Successfully opened workbook")
        print(f"✓ Sheet names: {wb.sheetnames}")
        
        ws = wb.active
        print(f"✓ Active sheet: {ws.title}")
        print(f"✓ Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Check some key cells
        header_cell = ws['A1']
        print(f"✓ Header (A1): '{header_cell.value}'")
        
        # Check if we have section titles
        sections_found = []
        for row in range(1, min(50, ws.max_row + 1)):
            cell_value = ws[f'A{row}'].value
            if cell_value and isinstance(cell_value, str):
                if 'Section' in cell_value:
                    sections_found.append(f"Row {row}: {cell_value}")
        
        if sections_found:
            print("✓ Found sections:")
            for section in sections_found:
                print(f"  {section}")
        else:
            print("⚠ No section headers found")
        
        # Check for data in expected areas
        data_cells = []
        for row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and cell.value != "":
                    data_cells.append(f"{cell.coordinate}: {cell.value}")
        
        print(f"✓ Found {len(data_cells)} non-empty cells in first 20×10 area")
        if len(data_cells) > 0:
            print("Sample cells:")
            for cell_info in data_cells[:10]:  # Show first 10
                print(f"  {cell_info}")
            if len(data_cells) > 10:
                print(f"  ... and {len(data_cells) - 10} more")
        
        wb.close()
        
    except Exception as e:
        print(f"✗ File inspection failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True

def main():
    """Run all tests"""
    print("DYNAMIC QA REPORT GENERATOR - STANDALONE TEST")
    print("=" * 60)
    
    # Track test results
    tests = [
        ("Basic Functionality", test_basic_functionality),
        ("Style Application", test_style_application),
        ("Report Generation", test_report_generation),
        ("File Inspection", inspect_generated_file)
    ]
    
    results = {}
    
    for test_name, test_func in tests:
        try:
            results[test_name] = test_func()
        except Exception as e:
            print(f"✗ Test '{test_name}' failed with exception: {e}")
            results[test_name] = False
    
    # Summary
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    
    passed = sum(1 for result in results.values() if result)
    total = len(results)
    
    for test_name, result in results.items():
        status = "✓ PASSED" if result else "✗ FAILED"
        print(f"{test_name:.<40} {status}")
    
    print(f"\nOverall: {passed}/{total} tests passed")
    
    if passed == total:
        print("\n🎉 All tests passed! The generator appears to be working correctly.")
        print(f"\nYou can open 'test_dynamic_qa_report.xlsx' to see the generated report.")
    else:
        print(f"\n⚠ {total - passed} test(s) failed. Check the output above for details.")
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
