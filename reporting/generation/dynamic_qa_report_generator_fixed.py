"""
Fixed Dynamic QA Report Generator - Creates Excel reports with proper summary tab naming
and correct population metrics.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple

logger = logging.getLogger(__name__)


class DynamicQAReportGeneratorFixed:
    """
    Fixed version that generates QA audit summary reports with:
    - Proper "Overall Analytics Summary" tab naming
    - Correct population counts (records not rules)
    - Complete audit leader inclusion
    """

    def __init__(self):
        """Initialize the report generator with styling configuration."""
        self._setup_styles()
        self._setup_layout_constants()

    def _setup_styles(self):
        """Define all visual styles for professional report formatting."""

        # Fonts
        self.title_font = Font(bold=True, size=12, color="FFFFFF")
        self.section_header_font = Font(bold=True, size=11, color="FFFFFF")
        self.column_header_font = Font(bold=True, size=10)
        self.data_font = Font(size=10)
        self.label_font = Font(size=10)

        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Color fills
        self.blue_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        self.blue_analytics_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.green_status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Number formats
        self.number_format = '#,##0'
        self.decimal_format = '0.00'
        self.percentage_format = '0.0%'

    def _setup_layout_constants(self):
        """Define layout spacing and positioning constants."""
        self.section1_start_row = 5
        self.section_spacing = 3
        self.section3_analytics_rows = 8

        # Column widths
        self.label_column_width = 25
        self.test_column_width = 15
        self.summary_column_width = 20
        self.leader_column_width = 20

    def generate_report(self,
                       rule_results: Dict[str, Any],
                       output_path: str,
                       responsible_party_column: str = "AuditLeader",
                       review_year: Optional[str] = None,
                       total_population: Optional[int] = None,
                       **kwargs) -> str:
        """
        Generate a complete QA summary report from scratch.

        Args:
            rule_results: Dictionary of rule_id → RuleEvaluationResult objects
            output_path: Path where Excel file should be saved
            responsible_party_column: Column name for audit leaders
            review_year: Year for the report header
            total_population: Total number of records in the dataset
            **kwargs: Additional parameters

        Returns:
            str: Path to the generated Excel file
        """

        logger.info(f"Generating QA report from scratch for {len(rule_results)} rules")

        # Calculate total population from all rule results if not provided
        if total_population is None:
            total_population = self._calculate_total_population(rule_results)
        
        logger.info(f"Total population: {total_population} records")

        # Extract dynamic structure
        rule_names, audit_leaders, leader_rule_matrix = self._extract_dynamic_structure(
            rule_results, responsible_party_column
        )

        logger.info(f"Found {len(rule_names)} rules and {len(audit_leaders)} audit leaders")

        # Create new workbook from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        # Use proper naming for the summary tab
        ws.title = "Overall Analytics Summary"

        # Add population information to header
        self._create_header_with_population(ws, review_year, total_population, len(rule_names))
        
        # PHASE 1: Build complete structure (no data, no formulas)
        section2_start = self._create_section1_structure(ws, rule_names)
        section3_start = self._create_section2_structure(ws, rule_names, audit_leaders, section2_start)
        section3_data_start = self._create_section3_structure(ws, rule_names, audit_leaders, section3_start)

        # PHASE 2: Populate Section 3 data first (this is the source of truth)
        self._populate_section3_data(ws, rule_names, audit_leaders, leader_rule_matrix, section3_data_start)

        # PHASE 3: Generate Section 1 formulas that reference Section 3
        self._generate_section1_formulas(ws, rule_names, section3_data_start, len(audit_leaders))

        # PHASE 4: Generate Section 2 formulas that reference Section 3
        self._generate_section2_formulas(ws, rule_names, audit_leaders, section2_start, section3_data_start)

        # PHASE 5: Final cleanup
        self._cleanup_worksheet(ws)
        wb.save(output_path)
        wb.close()

        logger.info(f"Generated QA report: {output_path}")
        return output_path

    def _calculate_total_population(self, rule_results: Dict[str, Any]) -> int:
        """Calculate total population from rule results."""
        max_population = 0
        
        for rule_id, rule_result in rule_results.items():
            if hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
                population = len(rule_result.result_df)
                max_population = max(max_population, population)
            elif hasattr(rule_result, 'total_records'):
                max_population = max(max_population, rule_result.total_records)
        
        return max_population

    def _create_header_with_population(self, ws, review_year: Optional[str], total_population: int, total_rules: int):
        """Create report header with population information."""
        # Title
        ws['A1'] = "IAG and AL Results and Ratings"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Year if provided
        if review_year:
            ws['A2'] = f"Review Year: {review_year}"
            ws['A2'].font = Font(bold=True, size=12)
        
        # Population information
        ws['A3'] = f"Total Population: {total_population:,} records"
        ws['A3'].font = Font(bold=True, size=11)
        
        ws['A4'] = f"Total Rules Applied: {total_rules}"
        ws['A4'].font = Font(bold=True, size=11)

    def _extract_dynamic_structure(self, rule_results: Dict[str, Any],
                                  responsible_party_column: str) -> Tuple[List[str], List[str], Dict]:
        """Extract rules, audit leaders, and performance matrix from validation results."""

        rule_names = list(rule_results.keys())
        audit_leaders = set()
        leader_rule_matrix = {}

        for rule_id, rule_result in rule_results.items():
            # Try to get data from party_results first
            if hasattr(rule_result, 'party_results') and rule_result.party_results:
                for leader, party_data in rule_result.party_results.items():
                    audit_leaders.add(leader)
                    if leader not in leader_rule_matrix:
                        leader_rule_matrix[leader] = {}

                    # Extract metrics from party_data
                    if isinstance(party_data, dict) and 'metrics' in party_data:
                        metrics = party_data['metrics']
                        leader_rule_matrix[leader][rule_id] = {
                            'gc_count': metrics.get('gc_count', 0),
                            'pc_count': metrics.get('pc_count', 0),
                            'dnc_count': metrics.get('dnc_count', 0),
                            'na_count': metrics.get('na_count', 0),
                            'total': metrics.get('total_count', 0),
                            'status': party_data.get('status', 'N/A'),
                            'error_rate': metrics.get('dnc_rate', 0),
                            'threshold': getattr(rule_result.rule, 'threshold', 0.02) if hasattr(rule_result, 'rule') else 0.02
                        }
                    else:
                        # Fallback calculation
                        leader_rule_matrix[leader][rule_id] = self._calculate_leader_metrics(
                            None, rule_result
                        )

            # Fallback to result_df if party_results not available
            elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
                df = rule_result.result_df
                if responsible_party_column in df.columns:
                    unique_leaders = df[responsible_party_column].dropna().unique()
                    for leader in unique_leaders:
                        audit_leaders.add(str(leader))
                        if str(leader) not in leader_rule_matrix:
                            leader_rule_matrix[str(leader)] = {}

                        leader_df = df[df[responsible_party_column] == leader]
                        leader_rule_matrix[str(leader)][rule_id] = self._calculate_leader_metrics(
                            leader_df, rule_result
                        )

        return sorted(rule_names), sorted(list(audit_leaders)), leader_rule_matrix

    def _calculate_leader_metrics(self, df: pd.DataFrame, rule_result: Any = None) -> Dict[str, Any]:
        """Calculate performance metrics for a leader's data in a specific rule."""

        # Try compliance_metrics first
        if rule_result and hasattr(rule_result, 'compliance_metrics'):
            metrics = rule_result.compliance_metrics
            gc_count = metrics.get('compliant_count', 0)
            pc_count = metrics.get('partially_compliant_count', 0)
            dnc_count = metrics.get('non_compliant_count', 0)
            na_count = metrics.get('not_applicable_count', 0)
        elif df is not None and not df.empty:
            # Analyze DataFrame
            total = len(df)
            result_col = self._find_result_column(df, rule_result)

            if result_col and result_col in df.columns:
                gc_count = df[result_col].isin(['GC', 'PASS', 'Pass', 'TRUE', True, 1]).sum()
                pc_count = df[result_col].isin(['PC', 'PARTIAL', 'Partial']).sum()
                dnc_count = df[result_col].isin(['DNC', 'FAIL', 'Fail', 'FALSE', False, 0]).sum()
                na_count = df[result_col].isin(['N/A', 'NA', None, '']).sum() + df[result_col].isna().sum()
            else:
                gc_count = pc_count = dnc_count = na_count = 0
        else:
            gc_count = pc_count = dnc_count = na_count = 0

        # Calculate derived metrics
        total = gc_count + pc_count + dnc_count + na_count
        applicable_total = gc_count + pc_count + dnc_count

        if applicable_total > 0:
            error_rate = (pc_count + dnc_count) / applicable_total
        else:
            error_rate = 0

        threshold = 0.02
        if rule_result and hasattr(rule_result, 'threshold'):
            threshold = rule_result.threshold

        # Determine status
        if applicable_total == 0:
            status = 'N/A'
        elif error_rate > threshold:
            status = 'DNC'
        else:
            status = 'GC'

        return {
            'total': total,
            'gc_count': gc_count,
            'pc_count': pc_count,
            'dnc_count': dnc_count,
            'na_count': na_count,
            'applicable_total': applicable_total,
            'error_rate': error_rate,
            'threshold': threshold,
            'status': status
        }

    def _find_result_column(self, df: pd.DataFrame, rule_result: Any = None) -> Optional[str]:
        """Find the result column in a DataFrame."""
        if rule_result and hasattr(rule_result, 'result_column'):
            return rule_result.result_column

        for col in ['Result', 'Status', 'Compliance', 'Overall Test Result']:
            if col in df.columns:
                return col
        
        return None

    # Include all other necessary methods from the original class...
    # (The rest would continue with the actual implementation methods)