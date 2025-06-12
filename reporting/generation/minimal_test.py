"""
Minimal QA Report Generator for Testing
Simplified version focusing on core functionality
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import datetime
from typing import Dict, Any, List, Tuple

class MinimalQAGenerator:
    """Simplified version for testing core report generation"""
    
    def __init__(self):
        self._setup_styles()
    
    def _setup_styles(self):
        """Set up basic styles"""
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.section_font = Font(bold=True, size=11)
        self.data_font = Font(size=10)
        
        # Colors
        self.blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Borders and alignment
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.center = Alignment(horizontal='center', vertical='center')
    
    def generate_simple_report(self, rule_results: Dict, output_path: str) -> str:
        """Generate a simple 3-section report"""
        
        # Extract basic data
        rules = list(rule_results.keys())
        leaders = self._get_leaders(rule_results)
        
        print(f"Generating report with {len(rules)} rules and {len(leaders)} leaders")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QA Report"
        
        current_row = 1
        
        # Header
        current_row = self._create_header(ws, current_row)
        
        # Section 1: Overall Summary
        current_row = self._create_section1(ws, rule_results, current_row)
        
        # Section 2: Leader Summary  
        current_row = self._create_section2(ws, rules, leaders, rule_results, current_row)
        
        # Section 3: Detailed Grid
        current_row = self._create_section3(ws, rules, leaders, rule_results, current_row)
        
        # Set column widths
        self._set_column_widths(ws, len(rules))
        
        # Save
        wb.save(output_path)
        wb.close()
        
        return output_path
    
    def _get_leaders(self, rule_results: Dict) -> List[str]:
        """Extract unique leaders from rule results"""
        leaders = set()
        for rule_result in rule_results.values():
            if hasattr(rule_result, 'party_results'):
                leaders.update(rule_result.party_results.keys())
        return sorted(list(leaders))
    
    def _create_header(self, ws, start_row: int) -> int:
        """Create report header"""
        # Main title
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws[f'A{start_row}'] = f"QA Summary Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
        ws[f'A{start_row}'].font = self.header_font
        ws[f'A{start_row}'].fill = self.blue_fill
        ws[f'A{start_row}'].alignment = self.center
        
        return start_row + 3
    
    def _create_section1(self, ws, rule_results: Dict, start_row: int) -> int:
        """Create Section 1: Overall Results"""
        
        # Section title
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws[f'A{start_row}'] = "Section 1: Overall Results"
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        current_row = start_row + 2
        
        # Calculate totals
        total_gc = total_pc = total_dnc = 0
        for rule_result in rule_results.values():
            if hasattr(rule_result, 'compliance_metrics'):
                metrics = rule_result.compliance_metrics
                total_gc += metrics.get('compliant_count', 0)
                total_pc += metrics.get('partially_compliant_count', 0)
                total_dnc += metrics.get('non_compliant_count', 0)
        
        # Headers
        headers = ['Metric', 'Count', 'Percentage']
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = self.border
        
        current_row += 1
        
        # Data rows
        total = total_gc + total_pc + total_dnc
        rows = [
            ('GC Count', total_gc, total_gc/total if total > 0 else 0),
            ('PC Count', total_pc, total_pc/total if total > 0 else 0),
            ('DNC Count', total_dnc, total_dnc/total if total > 0 else 0),
            ('Total', total, 1.0)
        ]
        
        for label, count, pct in rows:
            ws.cell(row=current_row, column=1).value = label
            ws.cell(row=current_row, column=2).value = count
            ws.cell(row=current_row, column=3).value = f"{pct:.1%}"
            
            # Apply borders
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        return current_row + 2
    
    def _create_section2(self, ws, rules: List[str], leaders: List[str], 
                        rule_results: Dict, start_row: int) -> int:
        """Create Section 2: Leader Summary"""
        
        # Section title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws[f'A{start_row}'] = "Section 2: Leader Summary"
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        current_row = start_row + 2
        
        # Headers
        ws.cell(row=current_row, column=1).value = "Leader"
        ws.cell(row=current_row, column=2).value = "GC"
        ws.cell(row=current_row, column=3).value = "PC" 
        ws.cell(row=current_row, column=4).value = "DNC"
        ws.cell(row=current_row, column=5).value = "Total"
        ws.cell(row=current_row, column=6).value = "Status"
        
        # Format headers
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = self.center
        
        current_row += 1
        
        # Leader rows
        for leader in leaders:
            gc_count = pc_count = dnc_count = 0
            
            # Calculate leader totals
            for rule_result in rule_results.values():
                if hasattr(rule_result, 'party_results') and leader in rule_result.party_results:
                    party_data = rule_result.party_results[leader]
                    if 'metrics' in party_data:
                        metrics = party_data['metrics']
                        gc_count += metrics.get('gc_count', 0)
                        pc_count += metrics.get('pc_count', 0)
                        dnc_count += metrics.get('dnc_count', 0)
            
            total = gc_count + pc_count + dnc_count
            
            # Determine status
            if total == 0:
                status = 'N/A'
                status_fill = None
            else:
                error_rate = (pc_count + dnc_count) / total
                if error_rate <= 0.05:
                    status = 'GC'
                    status_fill = self.green_fill
                elif error_rate <= 0.15:
                    status = 'PC'
                    status_fill = self.yellow_fill
                else:
                    status = 'DNC'
                    status_fill = self.red_fill
            
            # Write data
            ws.cell(row=current_row, column=1).value = leader
            ws.cell(row=current_row, column=2).value = gc_count
            ws.cell(row=current_row, column=3).value = pc_count
            ws.cell(row=current_row, column=4).value = dnc_count
            ws.cell(row=current_row, column=5).value = total
            
            status_cell = ws.cell(row=current_row, column=6)
            status_cell.value = status
            if status_fill:
                status_cell.fill = status_fill
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = self.border
            
            current_row += 1
        
        return current_row + 2
    
    def _create_section3(self, ws, rules: List[str], leaders: List[str], 
                        rule_results: Dict, start_row: int) -> int:
        """Create Section 3: Detailed Grid"""
        
        # Section title
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws[f'A{start_row}'] = "Section 3: Detailed Results Grid"
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        current_row = start_row + 2
        
        # Headers
        ws.cell(row=current_row, column=1).value = "Leader"
        
        for i, rule in enumerate(rules):
            ws.cell(row=current_row, column=i+2).value = f"Rule {i+1}"
        
        # Format headers
        for col in range(1, len(rules) + 2):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = self.center
        
        current_row += 1
        
        # Data grid
        for leader in leaders:
            ws.cell(row=current_row, column=1).value = leader
            
            for i, rule_id in enumerate(rules):
                rule_result = rule_results[rule_id]
                status = 'N/A'
                fill = None
                
                if hasattr(rule_result, 'party_results') and leader in rule_result.party_results:
                    party_data = rule_result.party_results[leader]
                    status = party_data.get('status', 'N/A')
                    
                    if status == 'GC':
                        fill = self.green_fill
                    elif status == 'PC':
                        fill = self.yellow_fill
                    elif status == 'DNC':
                        fill = self.red_fill
                
                cell = ws.cell(row=current_row, column=i+2)
                cell.value = status
                cell.border = self.border
                cell.alignment = self.center
                if fill:
                    cell.fill = fill
            
            # Border for leader name
            ws.cell(row=current_row, column=1).border = self.border
            
            current_row += 1
        
        return current_row
    
    def _set_column_widths(self, ws, num_rules: int):
        """Set appropriate column widths"""
        ws.column_dimensions['A'].width = 20  # Leader names
        
        for i in range(2, num_rules + 10):  # Extra columns for summaries
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 12

# Simple test function
def test_minimal_generator():
    """Test the minimal generator"""
    
    class MockRule:
        def __init__(self, name):
            self.name = name
    
    class MockResult:
        def __init__(self):
            self.compliance_metrics = {
                'compliant_count': 80,
                'partially_compliant_count': 15,
                'non_compliant_count': 5
            }
            self.party_results = {
                'Alice': {'status': 'GC', 'metrics': {'gc_count': 40, 'pc_count': 5, 'dnc_count': 0}},
                'Bob': {'status': 'PC', 'metrics': {'gc_count': 25, 'pc_count': 10, 'dnc_count': 15}},
                'Carol': {'status': 'DNC', 'metrics': {'gc_count': 15, 'pc_count': 0, 'dnc_count': 35}}
            }
    
    # Create test data
    rule_results = {
        'Rule-1': MockResult(),
        'Rule-2': MockResult(),
        'Rule-3': MockResult()
    }
    
    # Generate report
    generator = MinimalQAGenerator()
    output_path = generator.generate_simple_report(rule_results, "minimal_test_report.xlsx")
    
    print(f"Generated minimal test report: {output_path}")
    return output_path

if __name__ == "__main__":
    test_minimal_generator()
