"""
Template-Based Report Generator for QA Analytics Framework
Uses Excel template structure (QA-ID-70 format) for generating reports
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import shutil
import pandas as pd
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging
import datetime

logger = logging.getLogger(__name__)


class TemplateBasedReportGenerator:
    """
    Generates Excel reports using your QA-ID-70 template structure.
    
    Template Structure Identified:
    - Rows 1-10: Analytic metadata (Title, ID, Description, etc.)
    - Rows 13-21: IAG Summary Results with formulas
    - Rows 24-52: Summary Results by Audit Leader
    - Rows 58-66: Detailed Results table
    """
    
    def __init__(self, template_path: str = None):
        self.template_path = Path(template_path) if template_path else None
        
        # Define the template structure based on your QA-ID-70 sheet
        self.template_structure = {
            'metadata': {
                'analytic_title': 'C1',
                'analytic_id': 'C2', 
                'analytic_type': 'C3',
                'description': 'C4',
                'observation_criteria': 'C5',
                'population_criteria': 'C6',
                'population_completeness': 'C7',
                'sample_selection': 'C8',
                'threshold_rationale': 'C9',
                'threshold': 'C10'
            },
            'summary_results': {
                'start_row': 13,
                'gc_count_cell': 'E15',
                'pc_count_cell': 'E16', 
                'dnc_count_cell': 'E17',
                'na_count_cell': 'E18',
                'failure_rate_cell': 'E19',
                'threshold_cell': 'E20',
                'overall_result_cell': 'E21'
            },
            'leader_results': {
                'start_row': 24,
                'template_rows': 7  # Each leader gets 7 rows (matching QA-ID-70)
            },
            'detailed_results': {
                'header_row': 60,
                'data_start_row': 61,
                'columns': {
                    'audit_leader': 'B',
                    'entity_id': 'C', 
                    'entity_name': 'D',
                    'entity_status': 'E',
                    'it_risk': 'F',
                    'is_risk': 'G',
                    'applications': 'H',
                    'test1_result': 'I',
                    'test2_result': 'J',
                    'override': 'K',
                    'overall_result': 'L',
                    'qa_comments': 'M',
                    'finding_id': 'N'
                }
            }
        }
    
    def generate_excel_from_template(self, 
                                   results: Dict[str, Any],
                                   rule_results: Dict[str, Any], 
                                   output_path: str,
                                   analytic_id: str = None,
                                   analytic_title: str = None,
                                   group_by: str = None) -> str:
        """
        Generate Excel report using your template structure.
        
        Args:
            results: Validation results from ValidationPipeline
            rule_results: Rule evaluation results
            output_path: Path for output file
            analytic_id: QA Analytic ID (e.g., "70")
            analytic_title: Title for the analytic
            group_by: Column for responsible party grouping (e.g., "Audit Leader")
        """
        
        try:
            # Create workbook (copy template if available, otherwise create new)
            if self.template_path and self.template_path.exists():
                shutil.copy2(self.template_path, output_path)
                # Load with data_only=True to avoid formula corruption
                workbook = openpyxl.load_workbook(output_path, data_only=True)
                
                # Remove external links to prevent corruption
                if hasattr(workbook, 'external_links') and workbook.external_links:
                    workbook.external_links.clear()
                
                logger.info(f"Using template: {self.template_path}")
            else:
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)  # Remove default sheet
                logger.info("Creating new workbook (no template found)")
            
            # Determine sheet name
            sheet_name = f"QA-ID-{analytic_id}" if analytic_id else f"QA-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}"
            
            # Create or get the target sheet
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                self._setup_sheet_template(sheet)
            
            # Populate the sheet with validation results
            self._populate_metadata_section(sheet, results, rule_results, analytic_id, analytic_title)
            self._populate_summary_section(sheet, results, rule_results)
            self._populate_leader_section(sheet, results, rule_results, group_by)
            self._populate_detailed_results(sheet, results, rule_results, group_by)
            
            # Save and close
            workbook.save(output_path)
            workbook.close()
            
            logger.info(f"Template-based report generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating template-based report: {e}")
            # Fallback to creating a basic report
            return self._generate_fallback_report(results, rule_results, output_path)
    
    def _setup_sheet_template(self, sheet):
        """Set up the basic template structure when creating a new sheet."""
        
        # Set up metadata section headers
        metadata_headers = [
            ("B1", "Analytic Title:", "C1", ""),
            ("B2", "QA Analytic ID:", "C2", ""),
            ("B3", "Analytic Type", "C3", "Direct ACT"),
            ("B4", "High Level Analytic Description:", "C4", ""),
            ("B5", "Observation Criteria:", "C5", ""),
            ("B6", "Population Criteria:", "C6", ""),
            ("B7", "Population Completeness:", "C7", ""),
            ("B8", "Sample Size & Selection Rationale:", "C8", ""),
            ("B9", "Threshold Rationale:", "C9", ""),
            ("B10", "Threshold:", "C10", "")
        ]
        
        for header_cell, header_text, value_cell, default_value in metadata_headers:
            sheet[header_cell] = header_text
            sheet[header_cell].font = Font(bold=True)
            sheet[value_cell] = default_value
            
        # Merge cells for long descriptions
        for row in [1, 4, 5, 6, 7, 8, 9]:
            sheet.merge_cells(f'C{row}:H{row}')
        
        # Set up IAG Summary Results section
        sheet["B13"] = "IAG Summary Results"
        sheet["B13"].font = Font(bold=True, size=14)
        
        # Headers row 14
        sheet["B14"] = "Name for Aggregating Results"
        sheet["C14"] = "Name"
        sheet["D14"] = "Score and Rating"
        sheet["E14"] = "Results"
        
        for col in ['B', 'C', 'D', 'E']:
            sheet[f'{col}14'].font = Font(bold=True)
        
        # Summary data rows
        summary_labels = [
            (15, '-', 'IAG Overall', 'Count of "GC"'),
            (16, '-', 'IAG Overall', 'Count of "PC"'),
            (17, '-', 'IAG Overall', 'Count of "DNC"'),
            (18, '-', 'IAG Overall', 'Count of "N/A"'),
            (19, '-', 'IAG Overall', 'Error Rate'),
            (20, '-', 'IAG Overall', 'Analytic Threshold'),
            (21, 'IAG Overall', 'IAG Overall', 'Test Result')
        ]
        
        for row, col_b, col_c, col_d in summary_labels:
            sheet[f'B{row}'] = col_b
            sheet[f'C{row}'] = col_c
            sheet[f'D{row}'] = col_d
        
        # Set up detailed results headers
        sheet["B58"] = "Detailed Results"
        sheet["B58"].font = Font(bold=True, size=14)
        
        sheet["B59"] = "Archer Record Data"
        sheet["I59"] = "QA Analytic Test(s) and Results"
        sheet.merge_cells('B59:H59')
        sheet.merge_cells('I59:N59')
        
        detailed_headers = [
            ("B60", "Audit Leader"),
            ("C60", "Audit Entity ID"),
            ("D60", "Audit Entity Name"), 
            ("E60", "Audit Entity Status"),
            ("F60", "IT Risk"),
            ("G60", "IS Risk"),
            ("H60", "Key Primary and Secondary IT Applications"),
            ("I60", "Test1: Validation Result"),
            ("J60", "Test2: Additional Test"),
            ("K60", "Overall Test Result Override\n(If applicable)"),
            ("L60", "Overall Test Result (after considering any applicable test result overrides)"),
            ("M60", "QA comments, including any Analytic Override Rationale(s)\nAll observations are considered reportable findings unless otherwise noted."),
            ("N60", "Finding ID, If applicable")
        ]
        
        for cell, text in detailed_headers:
            sheet[cell] = text
            sheet[cell].font = Font(bold=True)
            sheet[cell].alignment = Alignment(wrap_text=True, vertical='center')
            
        # Apply basic formatting
        self._apply_basic_formatting(sheet)
    
    def _populate_metadata_section(self, sheet, results, rule_results, analytic_id, analytic_title):
        """Populate the metadata section (rows 1-10)."""
        
        # Basic metadata from results
        sheet["C1"] = analytic_title or results.get('analytic_id', 'Validation Analysis')
        sheet["C2"] = analytic_id or results.get('analytic_id', 'AUTO')
        sheet["C4"] = f"Automated validation analysis with {len(rule_results)} rules"
        
        # Extract threshold from rules (use the most common threshold)
        thresholds = []
        for rule_result in rule_results.values():
            if hasattr(rule_result, 'rule') and hasattr(rule_result.rule, 'threshold'):
                thresholds.append(rule_result.rule.threshold)
        
        if thresholds:
            # Convert from 0-1 scale to percentage if needed
            common_threshold = max(set(thresholds), key=thresholds.count)
            if common_threshold <= 1.0:
                sheet["C10"] = common_threshold
            else:
                sheet["C10"] = common_threshold / 100.0
        else:
            sheet["C10"] = 0.02  # Default 2% threshold
        
        # Format threshold as percentage
        sheet["C10"].number_format = '0%'
        
        # Population information
        data_metrics = results.get('data_metrics', {})
        total_items = data_metrics.get('row_count', 0)
        
        sheet["C5"] = "Compliance validation based on defined business rules"
        sheet["C6"] = f"All records in the dataset. Total population: {total_items:,} items."
        sheet["C7"] = "Data was loaded directly from the provided data source."
        sheet["C8"] = f"QA tested 100% of the population ({total_items:,} items)."
        sheet["C9"] = "Threshold based on rule configuration and industry best practices."
    
    def _populate_summary_section(self, sheet, results, rule_results):
        """Populate the IAG Summary Results section with formulas."""
        
        # Calculate overall counts across all rules
        overall_gc = 0
        overall_pc = 0
        overall_dnc = 0
        overall_na = 0
        
        # Count by looking at rule compliance status
        for rule_result in rule_results.values():
            if hasattr(rule_result, 'compliance_status'):
                status = str(rule_result.compliance_status)
                if 'COMPLIANT' in status and 'NOT' not in status:
                    overall_gc += 1
                elif 'PARTIALLY' in status:
                    overall_pc += 1
                elif 'NOT_COMPLIANT' in status or 'NON_COMPLIANT' in status:
                    overall_dnc += 1
                else:
                    overall_na += 1
        
        # Set the counts
        sheet["E15"] = overall_gc  # GC count
        sheet["E16"] = overall_pc  # PC count  
        sheet["E17"] = overall_dnc # DNC count
        sheet["E18"] = overall_na  # N/A count
        
        # Calculate error rate
        total_rules = overall_gc + overall_pc + overall_dnc
        if total_rules > 0:
            error_rate = (overall_pc + overall_dnc) / total_rules
            sheet["E19"] = error_rate
        else:
            sheet["E19"] = 0
        
        # Format as percentage
        sheet["E19"].number_format = '0.0%'
        
        # Threshold reference - use the actual value instead of formula
        sheet["E20"] = sheet["C10"].value if sheet["C10"].value else 0.02
        
        # Overall result - calculate instead of using formula
        error_rate = sheet["E19"].value if sheet["E19"].value else 0
        threshold = sheet["E20"].value if sheet["E20"].value else 0.02
        sheet["E21"] = "GC" if error_rate <= threshold else "DNC"
    
    def _populate_leader_section(self, sheet, results, rule_results, group_by):
        """Populate the Summary Results by Audit Leader section."""
        
        # Start of leader section
        sheet["B24"] = "Summary Results by Audit Leader"
        sheet["B24"].font = Font(bold=True, size=14)
        
        # Headers
        sheet["B25"] = "Name for Aggregating Results"
        sheet["C25"] = "Name"
        sheet["D25"] = "Score and Rating"
        sheet["E25"] = "Results"
        
        for col in ['B', 'C', 'D', 'E']:
            sheet[f'{col}25'].font = Font(bold=True)
        
        if not group_by or 'grouped_summary' not in results:
            # No grouping data available
            sheet["B26"] = "-"
            sheet["C26"] = "No grouping data available"
            return
        
        grouped_data = results.get('grouped_summary', {})
        current_row = 26
        
        # Process up to 4 leaders (matching template structure)
        leader_count = 0
        for leader, stats in sorted(grouped_data.items())[:4]:
            # Create section for this leader
            base_row = current_row
            
            # Reference row for this leader (where the name appears)
            ref_row = base_row + 6
            
            # Add data rows
            rows_data = [
                ('-', leader, 'Count of "GC"', stats.get('GC', 0)),
                ('-', leader, 'Count of "PC"', stats.get('PC', 0)),
                ('-', leader, 'Count of "DNC"', stats.get('DNC', 0)),
                ('-', leader, 'Count of "N/A"', 0),
                ('-', leader, 'Error Rate', None),  # Will calculate
                ('-', leader, 'Analytic Threshold', None),  # Will reference C10
                (leader, leader, 'Test Result', None)  # Will calculate
            ]
            
            for i, (col_b, col_c, col_d, col_e) in enumerate(rows_data):
                row = base_row + i
                sheet[f'B{row}'] = col_b
                sheet[f'C{row}'] = col_c
                sheet[f'D{row}'] = col_d
                
                if i == 4:  # Error rate calculation
                    total = stats.get('total_rules', 0)
                    if total > 0:
                        error_rate = (stats.get('PC', 0) + stats.get('DNC', 0)) / total
                        sheet[f'E{row}'] = error_rate
                        sheet[f'E{row}'].number_format = '0.0%'
                    else:
                        sheet[f'E{row}'] = 0
                elif i == 5:  # Threshold reference
                    sheet[f'E{row}'] = sheet["C10"].value if sheet["C10"].value else 0.02
                elif i == 6:  # Test result
                    # Calculate test result based on error rate vs threshold
                    error_rate_row = base_row + 4
                    threshold_row = base_row + 5
                    if sheet[f'E{error_rate_row}'].value <= sheet[f'E{threshold_row}'].value:
                        sheet[f'E{row}'] = "GC"
                    else:
                        sheet[f'E{row}'] = "DNC"
                elif col_e is not None:
                    sheet[f'E{row}'] = col_e
            
            current_row = base_row + 7  # Move to next leader section (7 rows per leader)
            leader_count += 1
            
            if leader_count >= 4:  # Template supports 4 leaders
                break
    
    def _populate_detailed_results(self, sheet, results, rule_results, group_by):
        """Populate the detailed results table."""
        
        current_row = 61
        
        # If we have detailed failure data from rules, show it
        has_detailed_data = False
        
        for rule_id, rule_result in rule_results.items():
            if hasattr(rule_result, 'result_df') and hasattr(rule_result, 'result_column'):
                # We have detailed results
                result_df = rule_result.result_df
                result_col = rule_result.result_column
                
                # Get failing items
                if result_col in result_df.columns:
                    failing_items = result_df[result_df[result_col] == False]
                    
                    if not failing_items.empty:
                        has_detailed_data = True
                        # Add up to 10 failing items as examples
                        for idx, (_, row_data) in enumerate(failing_items.head(10).iterrows()):
                            self._populate_result_row(sheet, current_row, row_data, 
                                                    rule_result.rule.name if hasattr(rule_result, 'rule') else rule_id,
                                                    'DNC', group_by)
                            current_row += 1
        
        # If no detailed data, add summary rows
        if not has_detailed_data:
            # Add a summary row for each rule
            for rule_id, rule_result in rule_results.items():
                rule_name = rule_result.rule.name if hasattr(rule_result, 'rule') else rule_id
                status = str(rule_result.compliance_status) if hasattr(rule_result, 'compliance_status') else 'UNKNOWN'
                
                # Determine simple status
                if 'COMPLIANT' in status and 'NOT' not in status:
                    simple_status = 'GC'
                elif 'PARTIALLY' in status:
                    simple_status = 'PC'
                elif 'NOT_COMPLIANT' in status:
                    simple_status = 'DNC'
                else:
                    simple_status = 'N/A'
                
                sheet[f'B{current_row}'] = 'Overall'
                sheet[f'C{current_row}'] = rule_id
                sheet[f'D{current_row}'] = rule_name
                sheet[f'L{current_row}'] = simple_status
                sheet[f'M{current_row}'] = f"Rule validation result: {status}"
                
                current_row += 1
    
    def _populate_result_row(self, sheet, row_num, row_data, rule_name, compliance_status, group_by):
        """Populate a single result row with data."""
        
        # Map data to template columns
        col_mapping = self.template_structure['detailed_results']['columns']
        
        # Audit Leader (from group_by column or default)
        if group_by and group_by in row_data:
            sheet[f"{col_mapping['audit_leader']}{row_num}"] = str(row_data[group_by])
        else:
            sheet[f"{col_mapping['audit_leader']}{row_num}"] = "Default"
        
        # Try to map common column names
        column_mappings = {
            'entity_id': ['entity_id', 'id', 'audit_entity_id', 'ae_id', 'Audit Entity ID'],
            'entity_name': ['entity_name', 'name', 'audit_entity_name', 'ae_name', 'Audit Entity Name'],
            'entity_status': ['status', 'entity_status', 'ae_status', 'Audit Entity Status'],
            'it_risk': ['it_risk', 'it_risk_rating', 'IT Risk'],
            'is_risk': ['is_risk', 'is_risk_rating', 'IS Risk'],
            'applications': ['applications', 'apps', 'mapped_apps', 'Key Primary and Secondary IT Applications']
        }
        
        for template_col, possible_names in column_mappings.items():
            value = None
            for name in possible_names:
                if name in row_data:
                    value = str(row_data[name]) if pd.notna(row_data[name]) else ""
                    break
            
            if value:
                sheet[f"{col_mapping[template_col]}{row_num}"] = value
        
        # Test results
        sheet[f"{col_mapping['test1_result']}{row_num}"] = compliance_status
        sheet[f"{col_mapping['overall_result']}{row_num}"] = compliance_status
        sheet[f"{col_mapping['qa_comments']}{row_num}"] = f"Failed validation: {rule_name}"
    
    def _apply_basic_formatting(self, sheet):
        """Apply basic formatting to match template style."""
        
        # Set column widths
        column_widths = {
            'A': 5,
            'B': 30,
            'C': 20,
            'D': 30,
            'E': 20,
            'F': 15,
            'G': 15,
            'H': 40,
            'I': 30,
            'J': 30,
            'K': 25,
            'L': 40,
            'M': 50,
            'N': 20
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Apply borders to header rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header row formatting
        for col in range(2, 15):  # B to N
            cell = sheet.cell(row=60, column=col)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Row heights
        sheet.row_dimensions[60].height = 40
        
        # Apply section formatting
        for row in [13, 24, 58]:  # Section headers
            if sheet[f'B{row}'].value:
                sheet[f'B{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                sheet[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    
    def _generate_fallback_report(self, results, rule_results, output_path):
        """Generate a basic fallback report if template processing fails."""
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Validation Results"
        
        # Basic summary
        sheet["A1"] = "QA Analytics Validation Results"
        sheet["A1"].font = Font(bold=True, size=16)
        
        sheet["A3"] = "Status:"
        sheet["B3"] = results.get('status', 'UNKNOWN')
        
        sheet["A4"] = "Total Rules:"
        sheet["B4"] = len(rule_results)
        
        # Rule results
        sheet["A6"] = "Rule Results:"
        sheet["A6"].font = Font(bold=True)
        
        row = 7
        for rule_id, rule_result in rule_results.items():
            sheet[f"A{row}"] = rule_id
            if hasattr(rule_result, 'compliance_status'):
                sheet[f"B{row}"] = str(rule_result.compliance_status)
            row += 1
        
        # Save fallback
        workbook.save(output_path)
        workbook.close()
        
        logger.info(f"Fallback report generated: {output_path}")
        return output_path