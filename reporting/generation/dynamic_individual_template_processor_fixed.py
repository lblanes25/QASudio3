"""
Fixed Dynamic Individual Analytic Template Processor
Addresses:
- Correct population metrics (records not rules)
- Dynamic column mapping based on actual data
- Complete audit leader expansion beyond 4
- Proper detailed results columns
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple
import shutil

logger = logging.getLogger(__name__)


class DynamicIndividualTemplateProcessorFixed:
    """
    Fixed processor that addresses all identified reporting issues.
    """
    
    def __init__(self, template_path: str):
        """Initialize the template processor."""
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Define styles for dynamic elements
        self.header_font = Font(bold=True, size=10)
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        
        # Define color fills for results
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Yellow highlight for referenced columns
        self.highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def generate_individual_report(self, 
                                  rule_id: str,
                                  rule_result: Any,
                                  output_path: str,
                                  responsible_party_column: str = "Audit Leader",
                                  test_columns: Optional[List[str]] = None,
                                  generation_date: Optional[str] = None,
                                  total_population: Optional[int] = None) -> str:
        """Generate an individual analytic report with all fixes applied."""
        
        logger.info(f"Generating comprehensive individual report for rule {rule_id}")
        
        # Copy template to output location
        shutil.copy2(self.template_path, output_path)
        
        # Open the copied template with data_only=True to avoid formula issues
        wb = openpyxl.load_workbook(output_path, data_only=True)
        
        # Remove external links to prevent corruption
        if hasattr(wb, 'external_links') and wb.external_links:
            wb.external_links.clear()
        ws = wb.active
        
        # Calculate total population for this rule
        if total_population is None:
            total_population = self._get_total_population(rule_result)
        
        # Update metadata section with correct population
        threshold = self._update_metadata_section_fixed(ws, rule_id, rule_result, total_population)
        
        # Get all unique audit leaders
        responsible_parties = self._get_responsible_parties(rule_result, responsible_party_column)
        
        # Calculate detailed results first (this determines the actual compliance)
        detailed_results = self._calculate_detailed_results(rule_result, responsible_party_column)
        
        # Calculate compliance statistics from the detailed results
        compliance_stats = self._calculate_compliance_from_detailed_results(detailed_results, responsible_parties)
        
        # Update IAG summary section based on actual detailed results
        self._update_iag_summary_fixed(ws, compliance_stats, threshold, total_population)
        
        # Update audit leader sections with proper counts (handle more than 4)
        detail_start_adjustment = self._update_all_audit_leaders_expanded(
            ws, responsible_parties, compliance_stats, threshold
        )
        
        # Sort detailed results by compliance status (DNC, GC, NA, PC order)
        detailed_results_sorted = sorted(detailed_results, key=lambda x: x['compliance'])
        
        # Create dynamic detailed results section with proper columns
        self._create_dynamic_detailed_results(
            ws, rule_result, detailed_results_sorted, responsible_party_column, 
            detail_start_adjustment
        )
        
        # Save the workbook
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Generated comprehensive individual report: {output_path}")
        return output_path
    
    def _get_total_population(self, rule_result: Any) -> int:
        """Get the total population (number of records) for this rule."""
        if hasattr(rule_result, 'total_records'):
            return rule_result.total_records
        elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
            return len(rule_result.result_df)
        else:
            # Try to get from compliance metrics
            if hasattr(rule_result, 'compliance_metrics'):
                metrics = rule_result.compliance_metrics
                total = (metrics.get('compliant_count', 0) + 
                        metrics.get('partially_compliant_count', 0) + 
                        metrics.get('non_compliant_count', 0) + 
                        metrics.get('not_applicable_count', 0))
                if total > 0:
                    return total
            return 0
    
    def _update_metadata_section_fixed(self, ws, rule_id: str, rule_result: Any, total_population: int) -> float:
        """Update the metadata section with correct population information."""
        
        # Extract rule information from RuleEvaluationResult or rule object
        if hasattr(rule_result, 'rule'):
            # This is a RuleEvaluationResult, get info from the rule
            rule = rule_result.rule
            rule_name = rule.name
            rule_id = rule.rule_id  
            description = rule.description
            formula = rule.formula
            threshold = rule.threshold
        else:
            # Direct rule object or other structure
            rule_name = getattr(rule_result, 'rule_name', getattr(rule_result, 'name', rule_id))
            description = getattr(rule_result, 'description', f"Validation analysis for {rule_name}")
            formula = getattr(rule_result, 'formula', getattr(rule_result, 'condition', ''))
            threshold = getattr(rule_result, 'threshold', 0.05)
        
        # Format rule name if it's the same as ID
        if rule_name == rule_id:
            rule_name = rule_id.replace('_', ' ').title()
        
        # C1: Analytic Title
        ws['C1'] = rule_name
        
        # C2: QA Analytic ID
        ws['C2'] = rule_id
        
        # C3: Analytic Type
        if ws['C3'].value is None or '{{' in str(ws['C3'].value):
            ws['C3'] = "Rule Validation"
        
        # C4: Description
        ws['C4'] = description
        
        # C5: Observation Criteria
        if formula:
            ws['C5'] = f"Rule Formula: {formula}"
        
        # C6: Population Criteria - FIXED to show total records
        ws['C6'] = f"Total records analyzed: {total_population:,}"
        
        # C7: Population Completeness
        ws['C7'] = "QA validated 100% of the available data"
        
        # C8: Sample Size & Selection
        ws['C8'] = f"Full population tested ({total_population:,} records)"
        
        # C9: Threshold Rationale
        if threshold:
            ws['C9'] = f"Threshold set based on rule requirements: {threshold:.1%}"
        else:
            threshold = 0.05  # Default 5%
            ws['C9'] = f"Default threshold applied: {threshold:.1%}"
        
        # C10: Threshold
        ws['C10'] = threshold
        
        logger.info(f"Updated metadata for rule {rule_id} ({rule_name}) with {total_population:,} records and threshold {threshold:.1%}")
        return threshold
    
    def _update_iag_summary_fixed(self, ws, compliance_stats: Dict, threshold: float, total_population: int):
        """Update the IAG summary section with proper compliance counts and population context."""
        
        overall = compliance_stats['overall']
        
        # Calculate actual counts (rows 15-18)
        gc_count = overall['GC']
        pc_count = overall['PC']
        dnc_count = overall['DNC']
        na_count = overall['N/A']
        
        # Add a header to show this is based on total population
        ws['B14'] = f"Results for {total_population:,} records:"
        ws['B14'].font = Font(italic=True, size=9)
        
        # Update counts
        ws['E15'] = gc_count
        ws['E16'] = pc_count
        ws['E17'] = dnc_count
        ws['E18'] = na_count
        
        # Calculate error rate: (PC + DNC) / (Total - N/A) (row 19)
        non_na_total = gc_count + pc_count + dnc_count
        if non_na_total > 0:
            error_rate = (pc_count + dnc_count) / non_na_total
            ws['E19'] = error_rate
            ws['E19'].number_format = '0.0%'
            
            # Add compliance rate next to error rate
            compliance_rate = gc_count / non_na_total
            ws['F19'] = f"({compliance_rate:.1%} compliance)"
            ws['F19'].font = Font(italic=True, size=9)
        else:
            ws['E19'] = "N/A"
        
        # Analytic threshold (row 20)
        ws['E20'] = threshold
        ws['E20'].number_format = '0.0%'
        
        # Test result based on error rate vs threshold (row 21)
        result_cell = ws['E21']
        if ws['E19'].value == "N/A":
            result_cell.value = "N/A"
            result_cell.fill = self.gray_fill
        else:
            if error_rate <= threshold:
                result_cell.value = "GC"
                result_cell.fill = self.green_fill
            elif error_rate <= threshold * 2:  # PC if within 2x threshold
                result_cell.value = "PC"
                result_cell.fill = self.yellow_fill
            else:
                result_cell.value = "DNC"
                result_cell.fill = self.red_fill
        
        # Apply borders to all cells
        for row in range(15, 22):
            ws[f'E{row}'].border = self.border
    
    def _update_all_audit_leaders_expanded(self, ws, responsible_parties: List[str], 
                                          compliance_stats: Dict, threshold: float) -> int:
        """Update audit leader sections, properly expanding beyond 4 leaders."""
        
        # Original template positions for first 4 leaders
        original_positions = [32, 39, 46, 53]
        rows_added = 0
        
        # Update the first 4 leaders in their original positions
        for i, row in enumerate(original_positions):
            if i < len(responsible_parties):
                leader = responsible_parties[i]
                self._update_leader_section(ws, row, leader, compliance_stats, threshold)
            else:
                # Clear unused sections properly
                self._clear_leader_section(ws, row)
        
        # If we have more than 4 leaders, dynamically add sections
        if len(responsible_parties) > 4:
            logger.info(f"Expanding report to include {len(responsible_parties)} audit leaders")
            
            # Calculate where to insert new leader sections
            insert_after_row = 53  # After the 4th leader
            
            # Each leader section is 7 rows
            for i in range(4, len(responsible_parties)):
                leader = responsible_parties[i]
                
                # Insert 7 rows for the new leader section
                ws.insert_rows(insert_after_row + 1, 7)
                rows_added += 7
                
                # Calculate positions for the new section
                section_start = insert_after_row + 1
                name_row = section_start + 6
                
                # Create the new leader section
                self._create_new_leader_section(ws, section_start, name_row, leader, compliance_stats, threshold)
                
                # Update insert position for next leader
                insert_after_row = name_row
                
            logger.info(f"Added {len(responsible_parties) - 4} additional leader sections")
        
        return rows_added
    
    def _clear_leader_section(self, ws, name_row: int):
        """Clear an unused leader section."""
        section_start = name_row - 6
        for row in range(section_start, name_row + 1):
            for col in ['B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                if not hasattr(cell, '__class__') or 'MergedCell' not in str(cell.__class__):
                    cell.value = ""
    
    def _create_new_leader_section(self, ws, section_start: int, name_row: int, 
                                   leader: str, compliance_stats: Dict, threshold: float):
        """Create a new leader section with proper formatting."""
        
        # Copy formatting from the first leader section
        template_start = 26  # First leader section starts at row 26
        
        # Copy the structure
        for offset in range(7):
            source_row = template_start + offset
            target_row = section_start + offset
            
            # Copy column headers/labels
            for col in ['B', 'C', 'D']:
                source_cell = ws[f'{col}{source_row}']
                target_cell = ws[f'{col}{target_row}']
                target_cell.value = source_cell.value
                if hasattr(source_cell, 'font'):
                    target_cell.font = source_cell.font.copy()
                if hasattr(source_cell, 'alignment'):
                    target_cell.alignment = source_cell.alignment.copy()
                if hasattr(source_cell, 'border'):
                    target_cell.border = source_cell.border.copy()
        
        # Now update with actual leader data
        self._update_leader_section(ws, name_row, leader, compliance_stats, threshold)
    
    def _create_dynamic_detailed_results(self, ws, rule_result: Any, detailed_results: List[Dict],
                                       responsible_party_column: str, detail_start_adjustment: int):
        """Create detailed results section with dynamic columns based on actual data."""
        
        # Find the start row for detailed results (after all leader sections)
        detail_start_row = 67 + detail_start_adjustment
        
        # Clear any existing headers
        header_row = detail_start_row
        for col in range(1, 20):  # Clear columns A through S
            cell = ws.cell(row=header_row, column=col)
            cell.value = ""
        
        # Determine relevant columns from the actual data
        if detailed_results:
            sample_data = detailed_results[0]['row_data']
            
            # Define columns to show based on what's actually in the data
            relevant_columns = [responsible_party_column]
            
            # Add other meaningful columns from the data
            for col in sample_data.keys():
                if col != responsible_party_column and not col.startswith('_'):
                    # Skip internal/system columns
                    if any(keyword in col.lower() for keyword in ['id', 'name', 'date', 'status', 'result', 'test']):
                        relevant_columns.append(col)
            
            # Ensure we have the compliance result column
            relevant_columns.append('Overall Test Result')
            
            # Write headers
            for i, col_name in enumerate(relevant_columns):
                cell = ws.cell(row=header_row, column=i+1)
                cell.value = col_name
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = self.center_align
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Write data rows
            for idx, result in enumerate(detailed_results):
                data_row = header_row + 1 + idx
                row_data = result['row_data']
                
                for col_idx, col_name in enumerate(relevant_columns[:-1]):  # All except last
                    cell = ws.cell(row=data_row, column=col_idx+1)
                    cell.value = row_data.get(col_name, '')
                    cell.font = self.data_font
                    cell.border = self.border
                
                # Add compliance result in last column
                compliance_cell = ws.cell(row=data_row, column=len(relevant_columns))
                compliance_cell.value = result['compliance']
                compliance_cell.font = self.data_font
                compliance_cell.border = self.border
                compliance_cell.alignment = self.center_align
                
                # Apply color based on result
                if result['compliance'] == 'GC':
                    compliance_cell.fill = self.green_fill
                elif result['compliance'] == 'PC':
                    compliance_cell.fill = self.yellow_fill
                elif result['compliance'] == 'DNC':
                    compliance_cell.fill = self.red_fill
                else:
                    compliance_cell.fill = self.gray_fill
            
            # Auto-fit column widths
            for col_idx, col_name in enumerate(relevant_columns):
                column_letter = get_column_letter(col_idx + 1)
                max_length = len(col_name)
                
                # Check data for max length
                for result in detailed_results[:50]:  # Sample first 50 rows
                    if col_name in result['row_data']:
                        max_length = max(max_length, len(str(result['row_data'][col_name])))
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        logger.info(f"Created detailed results section with {len(detailed_results)} rows")
    
    # Include other necessary methods from the original class...
    def _get_responsible_parties(self, rule_result: Any, responsible_party_column: str) -> List[str]:
        """Get sorted list of unique responsible parties."""
        responsible_parties = []
        
        if hasattr(rule_result, 'party_results') and rule_result.party_results:
            responsible_parties = list(rule_result.party_results.keys())
        elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
            df = rule_result.result_df
            if responsible_party_column in df.columns:
                responsible_parties = df[responsible_party_column].dropna().unique().tolist()
        
        return sorted(responsible_parties)
    
    def _calculate_detailed_results(self, rule_result: Any, responsible_party_column: str) -> List[Dict]:
        """Use the actual test results that already exist in the system."""
        detailed_results = []
        
        # Check if this is a RuleEvaluationResult object
        if hasattr(rule_result, 'result_df') and hasattr(rule_result, 'result_column'):
            logger.info(f"Using RuleEvaluationResult with result_column: {rule_result.result_column}")
            df = rule_result.result_df
            result_column = rule_result.result_column
            
            # Use the actual results from the system evaluation
            for _, row in df.iterrows():
                # Get the compliance result from the result column
                if result_column in row.index and not pd.isna(row[result_column]):
                    raw_result = row[result_column]
                    compliance = self._map_to_compliance_status(raw_result)
                else:
                    compliance = 'N/A'
                
                result_dict = {
                    'audit_leader': row.get(responsible_party_column, 'Unknown'),
                    'compliance': compliance,
                    'row_data': row.to_dict()
                }
                detailed_results.append(result_dict)
        
        logger.info(f"Generated {len(detailed_results)} detailed results using system data")
        return detailed_results
    
    def _map_to_compliance_status(self, value: Any) -> str:
        """Map various result values to standard compliance statuses."""
        if pd.isna(value):
            return 'N/A'
        
        value_str = str(value).upper().strip()
        
        if value_str in ['GC', 'PASS', 'TRUE', '1', 'YES', 'COMPLIANT']:
            return 'GC'
        elif value_str in ['PC', 'PARTIAL', 'PARTIALLY']:
            return 'PC'
        elif value_str in ['DNC', 'FAIL', 'FALSE', '0', 'NO', 'NON-COMPLIANT']:
            return 'DNC'
        else:
            return 'N/A'
    
    def _calculate_compliance_from_detailed_results(self, detailed_results: List[Dict], 
                                                  responsible_parties: List[str]) -> Dict:
        """Calculate compliance statistics from detailed results."""
        stats = {
            'overall': {'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0},
            'by_party': {}
        }
        
        # Initialize party stats
        for party in responsible_parties:
            stats['by_party'][party] = {'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0}
        
        # Count results from detailed calculations
        for result in detailed_results:
            compliance = result['compliance']
            party = result['audit_leader']
            
            # Update overall stats
            if compliance in stats['overall']:
                stats['overall'][compliance] += 1
                stats['overall']['total'] += 1
            
            # Update party stats
            if party in stats['by_party'] and compliance in stats['by_party'][party]:
                stats['by_party'][party][compliance] += 1
                stats['by_party'][party]['total'] += 1
        
        logger.info(f"Calculated compliance stats - Overall: {stats['overall']}")
        return stats
    
    def _update_leader_section(self, ws, name_row: int, leader: str, 
                               compliance_stats: Dict, threshold: float):
        """Update an existing leader section with calculated values."""
        
        # Set leader name
        ws[f'B{name_row}'] = leader
        
        # Get stats for this leader
        leader_stats = compliance_stats['by_party'].get(leader, {
            'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0
        })
        
        # Update counts (section starts 6 rows before the name row)
        section_start = name_row - 6
        ws[f'E{section_start}'] = leader_stats['GC']      # Count of GC
        ws[f'E{section_start + 1}'] = leader_stats['PC']  # Count of PC
        ws[f'E{section_start + 2}'] = leader_stats['DNC'] # Count of DNC
        ws[f'E{section_start + 3}'] = leader_stats['N/A'] # Count of N/A
        
        # Calculate error rate
        non_na_total = leader_stats['GC'] + leader_stats['PC'] + leader_stats['DNC']
        if non_na_total > 0:
            error_rate = (leader_stats['PC'] + leader_stats['DNC']) / non_na_total
            ws[f'E{section_start + 4}'] = error_rate
            ws[f'E{section_start + 4}'].number_format = '0.0%'
        else:
            ws[f'E{section_start + 4}'] = "N/A"
        
        # Test result
        result_cell = ws[f'E{section_start + 5}']
        if non_na_total == 0:
            result_cell.value = "N/A"
            result_cell.fill = self.gray_fill
        elif error_rate <= threshold:
            result_cell.value = "GC"
            result_cell.fill = self.green_fill
        elif error_rate <= threshold * 2:
            result_cell.value = "PC"
            result_cell.fill = self.yellow_fill
        else:
            result_cell.value = "DNC"
            result_cell.fill = self.red_fill
        
        # Apply borders
        for offset in range(6):
            ws[f'E{section_start + offset}'].border = self.border