"""
Dynamic Individual Analytic Template Processor V5

Comprehensive fix for all identified issues:
- Proper IAG summary calculation based on actual detailed results
- Fixed Overall Test Result column logic
- Added formula description column
- Yellow highlighting for referenced columns
- Clean formatting without None values
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple
import shutil

logger = logging.getLogger(__name__)


class DynamicIndividualTemplateProcessorV5:
    """
    Comprehensive processor that fixes all identified reporting issues.
    """
    
    def __init__(self, template_path: str):
        """Initialize the template processor."""
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Define styles for dynamic elements
        self.header_font = Font(bold=True, size=10)
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        
        # Define color fills for results
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Yellow highlight for referenced columns
        self.highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def generate_individual_report(self, 
                                  rule_id: str,
                                  rule_result: Any,
                                  output_path: str,
                                  responsible_party_column: str = "Audit Leader",
                                  test_columns: Optional[List[str]] = None,
                                  generation_date: Optional[str] = None) -> str:
        """Generate an individual analytic report with all fixes applied."""
        
        logger.info(f"Generating comprehensive individual report for rule {rule_id}")
        
        # Copy template to output location
        shutil.copy2(self.template_path, output_path)
        
        # Open the copied template
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Update metadata section
        threshold = self._update_metadata_section(ws, rule_id, rule_result)
        
        # Get all unique audit leaders
        responsible_parties = self._get_responsible_parties(rule_result, responsible_party_column)
        
        # Calculate detailed results first (this determines the actual compliance)
        detailed_results = self._calculate_detailed_results(rule_result, responsible_party_column)
        
        # Calculate compliance statistics from the detailed results
        compliance_stats = self._calculate_compliance_from_detailed_results(detailed_results, responsible_parties)
        
        # Update IAG summary section based on actual detailed results
        self._update_iag_summary(ws, compliance_stats, threshold)
        
        # Update audit leader sections with proper counts
        detail_start_adjustment = self._update_all_audit_leaders(
            ws, responsible_parties, compliance_stats, threshold
        )
        
        # Create dynamic detailed results section with all enhancements
        self._create_enhanced_detailed_results(
            ws, rule_result, detailed_results, responsible_party_column, 
            test_columns, detail_start_adjustment
        )
        
        # Clean up extra formatting after all data is populated
        # Find where QA Comments column ends in the detail section
        detail_start_row = 60 + detail_start_adjustment
        # QA Comments is at comment_col_idx (result_col_idx + 1), so 3 columns after that
        qa_comments_col = len(self._determine_columns_for_detailed_results(detailed_results[0]['row_data'] if detailed_results else {}, responsible_party_column)) + 3
        self._cleanup_extra_formatting(ws, qa_comments_col + 1)  # Start cleanup after QA Comments
        
        # Save the workbook
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Generated comprehensive individual report: {output_path}")
        return output_path
    
    def _get_responsible_parties(self, rule_result: Any, responsible_party_column: str) -> List[str]:
        """Get sorted list of unique responsible parties."""
        responsible_parties = []
        
        if hasattr(rule_result, 'party_results') and rule_result.party_results:
            responsible_parties = list(rule_result.party_results.keys())
        elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
            df = rule_result.result_df
            if responsible_party_column in df.columns:
                responsible_parties = df[responsible_party_column].dropna().unique().tolist()
        
        return sorted(responsible_parties)
    
    def _calculate_detailed_results(self, rule_result: Any, responsible_party_column: str) -> List[Dict]:
        """Use the actual test results that already exist in the system."""
        detailed_results = []
        
        # Check if this is a RuleEvaluationResult object
        if hasattr(rule_result, 'result_df') and hasattr(rule_result, 'result_column'):
            logger.info(f"Using RuleEvaluationResult with result_column: {rule_result.result_column}")
            df = rule_result.result_df
            result_column = rule_result.result_column
            
            # Use the actual results from the system evaluation
            for _, row in df.iterrows():
                # Get the compliance result from the result column
                if result_column in row.index and not pd.isna(row[result_column]):
                    raw_result = row[result_column]
                    compliance = self._map_to_compliance_status(raw_result)
                    logger.debug(f"Using system result {result_column}: {raw_result} -> {compliance}")
                else:
                    # Fallback if result column is missing/null
                    compliance = 'N/A'
                    logger.warning(f"Result column {result_column} missing or null for record")
                
                result_dict = {
                    'audit_leader': row.get(responsible_party_column, 'Unknown'),
                    'compliance': compliance,
                    'row_data': row.to_dict()
                }
                detailed_results.append(result_dict)
                
        else:
            # Fallback: check for other data structures
            if hasattr(rule_result, 'evaluated_data') and rule_result.evaluated_data is not None:
                logger.info(f"Using rule_result.evaluated_data")
                df = rule_result.evaluated_data
            elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
                logger.info(f"Using rule_result.result_df (no result_column)")
                df = rule_result.result_df
            else:
                logger.warning("No result data found in rule_result")
                return detailed_results
            
            logger.info(f"Processing {len(df)} records - searching for existing result columns")
            
            for _, row in df.iterrows():
                compliance = None
                
                # Look for existing compliance/result columns
                for result_col in ['Compliance', 'Result', 'Overall_Result', 'Status', 'TestResult', 'Passes', 'Fails']:
                    if result_col in row.index and not pd.isna(row[result_col]):
                        compliance = self._map_to_compliance_status(row[result_col])
                        logger.debug(f"Found existing result {result_col}: {row[result_col]} -> {compliance}")
                        break
                
                # If no existing result found, use rule logic as fallback
                if compliance is None:
                    logger.warning(f"No existing result found for record, using rule logic")
                    compliance = self._evaluate_row_compliance(row, rule_result)
                
                result_dict = {
                    'audit_leader': row.get(responsible_party_column, 'Unknown'),
                    'compliance': compliance,
                    'row_data': row.to_dict()
                }
                detailed_results.append(result_dict)
        
        logger.info(f"Generated {len(detailed_results)} detailed results using system data")
        return detailed_results
    
    def _evaluate_row_compliance(self, row_data: pd.Series, rule_result: Any) -> str:
        """
        Evaluate compliance for a single row based on the actual rule logic.
        This is the core logic that determines GC/PC/DNC/N/A for each record.
        """
        
        # Check if there's an existing result column first
        for col in ['Result', 'Overall Result', 'Compliance', 'Status']:
            if col in row_data.index and not pd.isna(row_data[col]):
                return self._map_to_compliance_status(row_data[col])
        
        # Apply rule-specific logic based on the rule formula/condition
        if hasattr(rule_result, 'formula') and rule_result.formula:
            # If we have a rule formula, evaluate it for this specific row
            return self._apply_rule_logic(row_data, rule_result)
        
        # Fall back to test column analysis
        return self._evaluate_test_columns(row_data)
    
    def _apply_rule_logic(self, row_data: pd.Series, rule_result: Any) -> str:
        """Apply specific rule logic to determine compliance for this specific row."""
        
        formula = getattr(rule_result, 'formula', '')
        rule_name = getattr(rule_result, 'rule_name', '')
        rule_id = getattr(rule_result, 'rule_id', '')
        
        logger.debug(f"Applying rule logic for rule {rule_id}: {formula}")
        
        # If the rule result already has the evaluation built in, use it
        if hasattr(rule_result, 'compliance_status'):
            return rule_result.compliance_status
        
        # Apply rule-specific logic based on formula content
        formula_lower = formula.lower()
        
        # Date/Activity rules
        if 'dateactivityoccurred' in formula_lower or 'not(isblank(' in formula_lower:
            return self._evaluate_date_activity_rule(row_data, formula)
        
        # Submitter/Reviewer rules  
        elif 'submitter' in formula_lower and 'reviewer' in formula_lower:
            return self._evaluate_submitter_reviewer_rule(row_data, formula)
        
        # Submitter/Approver rules
        elif 'submitter' in formula_lower and 'approver' in formula_lower:
            return self._evaluate_submitter_approver_rule(row_data, formula)
        
        # Data-Driven Auditing Procedure (DDAP) rules
        elif 'ddap' in formula_lower or 'data' in formula_lower or 'procedure' in formula_lower:
            return self._evaluate_ddap_compliance(row_data)
        
        # Impact assessment rules
        elif 'impact' in formula_lower:
            return self._evaluate_impact_compliance(row_data)
        
        # Default evaluation using test columns
        else:
            return self._evaluate_test_columns(row_data)
    
    def _evaluate_date_activity_rule(self, row_data: pd.Series, formula: str) -> str:
        """Evaluate date activity rules like NOT(ISBLANK([DateActivityOccurred]))"""
        
        date_field = 'DateActivityOccurred'
        if date_field not in row_data.index:
            return 'N/A'
        
        date_value = row_data[date_field]
        
        # Rule checks if date is NOT blank
        if 'not(isblank(' in formula.lower():
            if pd.isna(date_value) or str(date_value).strip() == '':
                return 'DNC'  # Date is blank, rule fails
            else:
                return 'GC'   # Date is populated, rule passes
        
        return 'N/A'
    
    def _evaluate_submitter_reviewer_rule(self, row_data: pd.Series, formula: str) -> str:
        """Evaluate submitter != reviewer rules"""
        
        submitter_field = 'Submitter'
        reviewer_field = 'Reviewer'
        
        if submitter_field not in row_data.index or reviewer_field not in row_data.index:
            return 'N/A'
        
        submitter = row_data[submitter_field]
        reviewer = row_data[reviewer_field]
        
        # Check if either is missing
        if pd.isna(submitter) or pd.isna(reviewer):
            return 'N/A'
        
        # Rule checks if submitter != reviewer
        if str(submitter).strip().lower() == str(reviewer).strip().lower():
            return 'DNC'  # Same person, rule fails
        else:
            return 'GC'   # Different people, rule passes
    
    def _evaluate_submitter_approver_rule(self, row_data: pd.Series, formula: str) -> str:
        """Evaluate submitter != approver rules"""
        
        submitter_field = 'Submitter'
        approver_field = 'Approver'
        
        if submitter_field not in row_data.index or approver_field not in row_data.index:
            return 'N/A'
        
        submitter = row_data[submitter_field]
        approver = row_data[approver_field]
        
        # Check if either is missing
        if pd.isna(submitter) or pd.isna(approver):
            return 'N/A'
        
        # Rule checks if submitter != approver
        if str(submitter).strip().lower() == str(approver).strip().lower():
            return 'DNC'  # Same person, rule fails
        else:
            return 'GC'   # Different people, rule passes
    
    def _evaluate_ddap_compliance(self, row_data: pd.Series) -> str:
        """Evaluate DDAP (Data-Driven Auditing Procedure) compliance for this row."""
        
        ddap_field = 'WasADataDrivenAuditingProcedureUtilized'
        
        # Check if DDAP field exists and has a value
        if ddap_field not in row_data.index or pd.isna(row_data[ddap_field]):
            return 'N/A'
        
        ddap_value = str(row_data[ddap_field]).strip().upper()
        
        # Evaluate based on DDAP utilization
        if ddap_value in ['YES', 'TRUE', '1', 'Y']:
            # DDAP was used - check if it was properly documented
            ddap_type_field = 'DDAPTypeUtilized'
            if ddap_type_field in row_data.index and not pd.isna(row_data[ddap_type_field]):
                return 'GC'  # Properly utilized and documented
            else:
                return 'PC'  # Used but not fully documented
                
        elif ddap_value in ['NO', 'FALSE', '0', 'N']:
            # DDAP was not used - this is generally a compliance issue
            return 'DNC'
            
        elif ddap_value in ['PARTIAL', 'SOMETIMES', 'PARTIALLY', 'MAYBE']:
            # Partial usage
            return 'PC'
            
        else:
            # Unknown/unclear value
            return 'N/A'
    
    def _evaluate_impact_compliance(self, row_data: pd.Series) -> str:
        """Evaluate impact-related compliance for this row."""
        
        impact_field = 'ImpactOccurred'
        
        # Check if impact field exists
        if impact_field not in row_data.index or pd.isna(row_data[impact_field]):
            return 'N/A'
        
        impact_value = str(row_data[impact_field]).strip().upper()
        
        if impact_value in ['NO', 'FALSE', '0', 'N', 'NONE']:
            # No impact occurred - this is good
            return 'GC'
            
        elif impact_value in ['YES', 'TRUE', '1', 'Y']:
            # Impact occurred - check if it's properly documented
            desc_field = 'ImpactDescription'
            if desc_field in row_data.index and not pd.isna(row_data[desc_field]):
                impact_desc = str(row_data[desc_field]).strip()
                if len(impact_desc) > 5:  # Meaningful description
                    return 'PC'  # Impact documented
                else:
                    return 'DNC'  # Poor documentation
            else:
                return 'DNC'  # Impact not documented
                
        else:
            # Unknown impact status
            return 'N/A'
    
    def _evaluate_test_columns(self, row_data: pd.Series) -> str:
        """Evaluate based on test columns."""
        
        # Look for test-related columns
        test_results = []
        for col in row_data.index:
            if any(keyword in col.lower() for keyword in ['test', 'check', 'validation', 'utilized', 'occurred']):
                if not pd.isna(row_data[col]):
                    test_results.append(str(row_data[col]).upper())
        
        if test_results:
            # Count different types of results
            pass_indicators = ['YES', 'TRUE', '1', 'PASS', 'GC', 'COMPLIANT']
            fail_indicators = ['NO', 'FALSE', '0', 'FAIL', 'DNC', 'NON-COMPLIANT']
            partial_indicators = ['PARTIAL', 'SOMETIMES', 'PC', 'PARTIALLY']
            
            passes = sum(1 for r in test_results if r in pass_indicators)
            fails = sum(1 for r in test_results if r in fail_indicators)
            partials = sum(1 for r in test_results if r in partial_indicators)
            
            # Determine overall compliance
            if fails > 0:
                return 'DNC'
            elif partials > 0:
                return 'PC'
            elif passes > 0:
                return 'GC'
            else:
                return 'N/A'
        
        return 'N/A'
    
    def _calculate_compliance_from_detailed_results(self, detailed_results: List[Dict], 
                                                   responsible_parties: List[str]) -> Dict:
        """Calculate compliance statistics from the actual detailed results."""
        
        stats = {
            'overall': {'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0},
            'by_party': {}
        }
        
        # Initialize party stats
        for party in responsible_parties:
            stats['by_party'][party] = {'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0}
        
        # Count results from detailed calculations
        for result in detailed_results:
            compliance = result['compliance']
            party = result['audit_leader']
            
            # Update overall stats
            if compliance in stats['overall']:
                stats['overall'][compliance] += 1
                stats['overall']['total'] += 1
            
            # Update party stats
            if party in stats['by_party'] and compliance in stats['by_party'][party]:
                stats['by_party'][party][compliance] += 1
                stats['by_party'][party]['total'] += 1
        
        logger.info(f"Calculated compliance stats - Overall: {stats['overall']}")
        return stats
    
    def _update_metadata_section(self, ws, rule_id: str, rule_result: Any) -> float:
        """Update the metadata section and return threshold."""
        
        # Extract rule information from RuleEvaluationResult or rule object
        if hasattr(rule_result, 'rule'):
            # This is a RuleEvaluationResult, get info from the rule
            rule = rule_result.rule
            rule_name = rule.name
            rule_id = rule.rule_id  
            description = rule.description
            formula = rule.formula
            threshold = rule.threshold
        else:
            # Direct rule object or other structure
            rule_name = getattr(rule_result, 'rule_name', getattr(rule_result, 'name', rule_id))
            description = getattr(rule_result, 'description', f"Validation analysis for {rule_name}")
            formula = getattr(rule_result, 'formula', getattr(rule_result, 'condition', ''))
            threshold = getattr(rule_result, 'threshold', 0.05)
        
        # Format rule name if it's the same as ID
        if rule_name == rule_id:
            rule_name = rule_id.replace('_', ' ').title()
        
        # C1: Analytic Title
        ws['C1'] = rule_name
        
        # C2: QA Analytic ID
        ws['C2'] = rule_id
        
        # C3: Analytic Type
        if ws['C3'].value is None or '{{' in str(ws['C3'].value):
            ws['C3'] = "Rule Validation"
        
        # C4: Description
        ws['C4'] = description
        
        # C5: Observation Criteria
        if formula:
            ws['C5'] = f"Rule Formula: {formula}"
        
        # C6: Population Criteria
        if hasattr(rule_result, 'total_records'):
            ws['C6'] = f"Total records analyzed: {rule_result.total_records}"
        elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
            ws['C6'] = f"Total records analyzed: {len(rule_result.result_df)}"
        
        # C7: Population Completeness
        ws['C7'] = "QA validated 100% of the available data"
        
        # C8: Sample Size & Selection
        ws['C8'] = "Full population tested (100%)"
        
        # C9: Threshold Rationale
        if threshold:
            ws['C9'] = f"Threshold set based on rule requirements: {threshold:.1%}"
        else:
            threshold = 0.05  # Default 5%
            ws['C9'] = f"Default threshold applied: {threshold:.1%}"
        
        # C10: Threshold
        ws['C10'] = threshold
        
        logger.info(f"Updated metadata for rule {rule_id} ({rule_name}) with threshold {threshold:.1%}")
        return threshold
    
    def _update_iag_summary(self, ws, compliance_stats: Dict, threshold: float):
        """Update the IAG summary section with proper compliance counts."""
        
        overall = compliance_stats['overall']
        
        # Calculate actual counts (rows 15-18)
        gc_count = overall['GC']
        pc_count = overall['PC']
        dnc_count = overall['DNC']
        na_count = overall['N/A']
        
        # Update counts
        ws['E15'] = gc_count
        ws['E16'] = pc_count
        ws['E17'] = dnc_count
        ws['E18'] = na_count
        
        # Calculate error rate: (PC + DNC) / (Total - N/A) (row 19)
        non_na_total = gc_count + pc_count + dnc_count
        if non_na_total > 0:
            error_rate = (pc_count + dnc_count) / non_na_total
            ws['E19'] = error_rate
            ws['E19'].number_format = '0.0%'
        else:
            ws['E19'] = "N/A"
        
        # Analytic threshold (row 20)
        ws['E20'] = threshold
        ws['E20'].number_format = '0.0%'
        
        # Test result based on error rate vs threshold (row 21)
        result_cell = ws['E21']
        if ws['E19'].value == "N/A":
            result_cell.value = "N/A"
            result_cell.fill = self.gray_fill
        else:
            if error_rate <= threshold:
                result_cell.value = "GC"
                result_cell.fill = self.green_fill
            elif error_rate <= threshold * 2:  # PC if within 2x threshold
                result_cell.value = "PC"
                result_cell.fill = self.yellow_fill
            else:
                result_cell.value = "DNC"
                result_cell.fill = self.red_fill
        
        # Apply borders to all cells
        for row in range(15, 22):
            ws[f'E{row}'].border = self.border
    
    def _update_all_audit_leaders(self, ws, responsible_parties: List[str], 
                                   compliance_stats: Dict, threshold: float) -> int:
        """Update audit leader sections with proper formatting and no None values."""
        
        # Original template positions
        original_positions = [32, 39, 46, 53]
        rows_added = 0
        
        # Update the first 4 leaders in their original positions
        for i, row in enumerate(original_positions):
            if i < len(responsible_parties):
                leader = responsible_parties[i]
                self._update_leader_section(ws, row, leader, compliance_stats, threshold)
            else:
                # Clear unused sections properly
                ws[f'B{row}'] = ""
                for offset in range(6):
                    section_row = row - 6 + offset
                    ws[f'E{section_row}'] = ""
                    ws[f'D{section_row}'] = ""
                    ws[f'C{section_row}'] = ""
                    ws[f'B{section_row}'] = ""
        
        # If we have more than 4 leaders, insert new sections
        if len(responsible_parties) > 4:
            # Insert rows after the last leader section (row 53)
            insert_after_row = 53
            
            # Each additional leader needs 7 rows
            for i in range(4, len(responsible_parties)):
                leader = responsible_parties[i]
                
                # Insert 7 rows for the new leader section
                ws.insert_rows(insert_after_row + 1, 7)
                rows_added += 7
                
                # Calculate positions for the new section
                section_start = insert_after_row + 1
                name_row = section_start + 6
                
                # Populate the new section
                self._create_leader_section(ws, section_start, name_row, leader, compliance_stats, threshold)
                
                # Update insert position for next leader
                insert_after_row = name_row
        
        # Clean up any None values in the remaining area
        self._clean_none_values(ws, 54, 66)  # Clean area between leaders and detailed results
        
        return rows_added
    
    def _clean_none_values(self, ws, start_row: int, end_row: int):
        """Clean up None values in the specified row range."""
        for row in range(start_row, end_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E']:
                try:
                    cell = ws[f'{col}{row}']
                    # Skip merged cells
                    if hasattr(cell, '__class__') and 'MergedCell' in str(cell.__class__):
                        continue
                    if cell.value == 'None' or cell.value is None:
                        cell.value = ""
                except Exception:
                    # Skip any problematic cells
                    continue
    
    def _update_leader_section(self, ws, name_row: int, leader: str, 
                               compliance_stats: Dict, threshold: float):
        """Update an existing leader section with calculated values."""
        
        # Set leader name
        ws[f'B{name_row}'] = leader
        
        # Get stats for this leader
        leader_stats = compliance_stats['by_party'].get(leader, {
            'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0
        })
        
        # Update counts (section starts 6 rows before the name row)
        section_start = name_row - 6
        ws[f'E{section_start}'] = leader_stats['GC']      # Count of GC
        ws[f'E{section_start + 1}'] = leader_stats['PC']  # Count of PC
        ws[f'E{section_start + 2}'] = leader_stats['DNC'] # Count of DNC
        ws[f'E{section_start + 3}'] = leader_stats['N/A'] # Count of N/A
        
        # Calculate error rate
        non_na_total = leader_stats['GC'] + leader_stats['PC'] + leader_stats['DNC']
        if non_na_total > 0:
            error_rate = (leader_stats['PC'] + leader_stats['DNC']) / non_na_total
            ws[f'E{section_start + 4}'] = error_rate
            ws[f'E{section_start + 4}'].number_format = '0.0%'
        else:
            ws[f'E{section_start + 4}'] = "N/A"
        
        # Threshold
        ws[f'E{section_start + 5}'] = threshold
        ws[f'E{section_start + 5}'].number_format = '0.0%'
        
        # Test result with color coding
        result_cell = ws[f'E{name_row}']
        if ws[f'E{section_start + 4}'].value == "N/A":
            result_cell.value = "N/A"
            result_cell.fill = self.gray_fill
        else:
            if error_rate <= threshold:
                result_cell.value = "GC"
                result_cell.fill = self.green_fill
            elif error_rate <= threshold * 2:
                result_cell.value = "PC"
                result_cell.fill = self.yellow_fill
            else:
                result_cell.value = "DNC"
                result_cell.fill = self.red_fill
    
    def _create_leader_section(self, ws, section_start: int, name_row: int, 
                               leader: str, compliance_stats: Dict, threshold: float):
        """Create a new leader section with calculated values."""
        
        # Get stats for this leader
        leader_stats = compliance_stats['by_party'].get(leader, {
            'GC': 0, 'PC': 0, 'DNC': 0, 'N/A': 0, 'total': 0
        })
        
        # Create the section structure
        labels = ['Count of "GC"', 'Count of "PC"', 'Count of "DNC"', 'Count of "N/A"', 'Error Rate', 'Analytic Threshold']
        values = [leader_stats['GC'], leader_stats['PC'], leader_stats['DNC'], leader_stats['N/A']]
        
        # Calculate error rate and add to values
        non_na_total = leader_stats['GC'] + leader_stats['PC'] + leader_stats['DNC']
        if non_na_total > 0:
            error_rate = (leader_stats['PC'] + leader_stats['DNC']) / non_na_total
            values.extend([error_rate, threshold])
        else:
            values.extend(["N/A", threshold])
        
        # Populate the section
        for i, (label, value) in enumerate(zip(labels, values)):
            row = section_start + i
            ws[f'B{row}'] = '-'
            ws[f'C{row}'] = leader
            ws[f'D{row}'] = label
            ws[f'E{row}'] = value
            
            # Format percentages
            if i in [4, 5] and isinstance(value, (int, float)) and value != "N/A":
                ws[f'E{row}'].number_format = '0.0%'
        
        # Leader name and test result
        ws[f'B{name_row}'] = leader
        ws[f'C{name_row}'] = leader
        ws[f'D{name_row}'] = 'Test Result'
        
        # Calculate and set result with color
        result_cell = ws[f'E{name_row}']
        if values[4] == "N/A":
            result_cell.value = "N/A"
            result_cell.fill = self.gray_fill
        else:
            if error_rate <= threshold:
                result_cell.value = "GC"
                result_cell.fill = self.green_fill
            elif error_rate <= threshold * 2:
                result_cell.value = "PC"
                result_cell.fill = self.yellow_fill
            else:
                result_cell.value = "DNC"
                result_cell.fill = self.red_fill
        
        # Apply formatting
        for row in range(section_start, name_row + 1):
            for col in ['B', 'C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                cell.font = self.data_font
                cell.border = self.border
    
    def _create_enhanced_detailed_results(self, ws, rule_result: Any, detailed_results: List[Dict],
                                        responsible_party_column: str, test_columns: Optional[List[str]],
                                        row_adjustment: int):
        """Create enhanced detailed results section with all requested features."""
        
        if not detailed_results:
            logger.warning("No detailed results data available")
            return
        
        # Calculate where detailed results should start
        detail_start_row = 60 + row_adjustment
        
        # Clear any existing data from detail start row onwards
        max_row = ws.max_row
        for row in range(detail_start_row, max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Add section headers
        ws[f'B{detail_start_row - 2}'] = "Detailed Results"
        ws[f'B{detail_start_row - 2}'].font = Font(bold=True, size=12)
        
        ws[f'B{detail_start_row - 1}'] = "Archer Record Data"
        
        # Get the data structure from the first result
        if detailed_results and 'row_data' in detailed_results[0]:
            sample_data = detailed_results[0]['row_data']
            columns_to_include = self._determine_columns_for_detailed_results(sample_data, responsible_party_column, test_columns)
        else:
            columns_to_include = [responsible_party_column]
        
        # Create headers with yellow highlighting for referenced columns
        header_row = detail_start_row
        col_idx = 2  # Start at column B
        
        # Track which columns are referenced in the rule
        referenced_columns = self._get_referenced_columns(rule_result)
        
        for col_name in columns_to_include:
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col_name
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply standard blue fill to all columns (no yellow highlighting)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            col_idx += 1
        
        # Add "Overall Test Result" column
        result_col_idx = col_idx
        result_cell = ws.cell(row=header_row, column=result_col_idx)
        result_cell.value = "Overall Test Result"
        result_cell.font = self.header_font
        result_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        result_cell.border = self.border
        
        # Add "QA Comments" column (removed Test Logic/Formula column)
        comment_col_idx = result_col_idx + 1
        comment_cell = ws.cell(row=header_row, column=comment_col_idx)
        comment_cell.value = "QA Comments"
        comment_cell.font = self.header_font
        comment_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        comment_cell.border = self.border
        
        # Add data rows
        data_start_row = header_row + 1
        
        for idx, result_dict in enumerate(detailed_results):
            current_row = data_start_row + idx
            row_data = result_dict['row_data']
            compliance = result_dict['compliance']
            
            # Add data for each column
            col_idx = 2
            for col_name in columns_to_include:
                cell = ws.cell(row=current_row, column=col_idx)
                
                if col_name in row_data:
                    value = row_data[col_name]
                    if pd.isna(value):
                        cell.value = ""
                    else:
                        cell.value = value
                
                cell.border = self.border
                cell.font = self.data_font
                col_idx += 1
            
            # Add overall test result with color coding
            result_cell = ws.cell(row=current_row, column=result_col_idx)
            result_cell.value = compliance
            result_cell.border = self.border
            result_cell.font = self.data_font
            
            # Apply color based on result
            if compliance == "GC":
                result_cell.fill = self.green_fill
            elif compliance == "PC":
                result_cell.fill = self.yellow_fill
            elif compliance == "DNC":
                result_cell.fill = self.red_fill
            else:  # N/A
                result_cell.fill = self.gray_fill
            
            # Add comments (removed formula explanation)
            comment_cell = ws.cell(row=current_row, column=comment_col_idx)
            if 'Comments' in row_data:
                comment_cell.value = str(row_data['Comments']) if not pd.isna(row_data['Comments']) else ""
            comment_cell.border = self.border
            comment_cell.font = self.data_font
    
    def _determine_columns_for_detailed_results(self, sample_data: Dict, responsible_party_column: str,
                                              test_columns: Optional[List[str]] = None) -> List[str]:
        """Determine which columns to include in the detailed results."""
        
        essential_columns = [responsible_party_column]
        
        # Add ID columns
        for col in ['AuditEntityID', 'Entity ID', 'Entity_ID', 'ID']:
            if col in sample_data and col not in essential_columns:
                essential_columns.append(col)
                break
        
        # Add date columns
        for col in ['DateActivityOccurred', 'Date', 'Activity Date']:
            if col in sample_data and col not in essential_columns:
                essential_columns.append(col)
                break
        
        # Add test columns if specified
        if test_columns:
            for col in test_columns:
                if col in sample_data and col not in essential_columns:
                    essential_columns.append(col)
        else:
            # Add columns that are likely to be test-related
            for col in sample_data.keys():
                if col not in essential_columns and col not in ['Comments', 'Result', 'Overall Result']:
                    if any(keyword in col for keyword in ['Was', 'Impact', 'Utilized', 'Type', 'Description']):
                        essential_columns.append(col)
        
        return essential_columns[:8]  # Limit for readability
    
    def _get_referenced_columns(self, rule_result: Any) -> List[str]:
        """Get columns that are referenced in the rule formula."""
        
        referenced = []
        
        # Common columns that are typically referenced in validation rules
        common_references = [
            'WasADataDrivenAuditingProcedureUtilized',
            'ImpactOccurred',
            'ImpactDescription',
            'AuditEntityID',
            'DateActivityOccurred'
        ]
        
        # If we have a formula, try to extract column references
        if hasattr(rule_result, 'formula') and rule_result.formula:
            formula = rule_result.formula.lower()
            for col in common_references:
                if col.lower() in formula:
                    referenced.append(col)
        else:
            # Default to common validation columns
            referenced = common_references
        
        return referenced
    
    def _get_formula_explanation(self, row_data: Dict, rule_result: Any, compliance: str) -> str:
        """Generate explanation of how the test result was determined."""
        
        if hasattr(rule_result, 'formula') and rule_result.formula:
            base_explanation = f"Rule: {rule_result.formula}"
        else:
            base_explanation = "Standard validation logic applied"
        
        # Add specific explanation based on the data
        if 'WasADataDrivenAuditingProcedureUtilized' in row_data:
            ddap_value = row_data['WasADataDrivenAuditingProcedureUtilized']
            base_explanation += f" | DDAP: {ddap_value}"
        
        if 'ImpactOccurred' in row_data:
            impact_value = row_data['ImpactOccurred']
            base_explanation += f" | Impact: {impact_value}"
        
        return base_explanation
    
    def _map_to_compliance_status(self, result_value) -> str:
        """Map result values to compliance status."""
        
        if pd.isna(result_value):
            return "N/A"
        
        result_str = str(result_value).upper()
        
        if result_str in ['PASS', 'TRUE', '1', 'YES', 'COMPLIANT', 'GC', 'SUCCESS']:
            return "GC"
        elif result_str in ['FAIL', 'FALSE', '0', 'NO', 'NON-COMPLIANT', 'DNC', 'FAILURE']:
            return "DNC"
        elif result_str in ['PARTIAL', 'WARN', 'WARNING', 'PC', 'PARTIALLY', 'SOMETIMES']:
            return "PC"
        elif result_str in ['N/A', 'NA', 'NOT APPLICABLE', 'SKIP']:
            return "N/A"
        else:
            return "DNC"  # Default to DNC for unknown values