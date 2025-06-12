"""
Fixed Dynamic Summary Template Processor for IAG and AL Results and Ratings

This version adds proper template clearing functionality to prevent duplication
of template content when generating dynamic reports.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.cell import MergedCell
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple
import shutil

logger = logging.getLogger(__name__)


class DynamicSummaryTemplateProcessor:
    """
    Processes the IAG and AL Results and Ratings summary template dynamically.
    Handles variable numbers of rules and audit leaders.
    FIXED: Properly clears template placeholder content before inserting dynamic data.
    """

    def __init__(self, template_path: str):
        """Initialize the template processor."""
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")

        # Define common styles
        self._setup_styles()

    def _setup_styles(self):
        """Define common Excel styles used throughout the template."""
        self.header_font = Font(bold=True, size=11)
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.number_format = '#,##0'
        self.percentage_format = '0.0%'

        # Define color fills for compliance statuses
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ==========================================
    # TEMPLATE CLEARING METHODS (NEW)
    # ==========================================
    
    def _clear_template_placeholders(self, ws):
        """
        Clear template placeholder content before inserting dynamic data.
        This prevents duplication of template sections.
        """
        logger.info("Clearing template placeholder content")
        
        # Find and clear Section 3 template content
        # Look for "Audit Leader Average Test Results" in the template
        template_section3_row = None
        for row in range(20, 50):  # Search in reasonable range
            cell_value = self._safe_read_cell(ws, f'A{row}')
            if cell_value and "Audit Leader Average Test Results" in str(cell_value):
                template_section3_row = row
                logger.info(f"Found template Section 3 at row {row}")
                break
        
        if template_section3_row:
            # Clear the template content from this row onwards
            # We need to clear until we find the next major section or end of content
            rows_to_clear = []
            for row in range(template_section3_row, template_section3_row + 20):
                # Check if this row has template content (Area, Analytics, etc.)
                has_template_content = False
                for col in range(1, 15):  # Check first 14 columns
                    cell_value = self._safe_read_cell(ws, f'{get_column_letter(col)}{row}')
                    if cell_value and any(keyword in str(cell_value) for keyword in 
                                         ["Area", "IAG-Wide Analytic", "Analytic Error Threshold",
                                          "Risk Level", "Budget", "Analytic ID", "Manual Samples"]):
                        has_template_content = True
                        break
                
                if has_template_content:
                    rows_to_clear.append(row)
            
            # Clear the identified rows
            for row in rows_to_clear:
                logger.debug(f"Clearing template row {row}")
                for col in range(1, 20):  # Clear first 19 columns
                    cell_ref = f'{get_column_letter(col)}{row}'
                    try:
                        # Preserve formatting but clear content
                        cell = ws[cell_ref]
                        if not isinstance(cell, MergedCell):
                            cell.value = None
                    except:
                        pass
        
        # Also clear any other template placeholders that might cause issues
        self._clear_other_template_sections(ws)
    
    def _clear_other_template_sections(self, ws):
        """Clear other template sections that might interfere with dynamic content."""
        # Clear empty audit leader rows in Section 2
        for row in range(15, 20):
            cell_b = self._safe_read_cell(ws, f'B{row}')
            if cell_b is None or str(cell_b).strip() == "":
                # This is likely a template placeholder row for audit leaders
                cell_c = self._safe_read_cell(ws, f'C{row}')
                if cell_c and "Total Weighted Score" in str(cell_c):
                    logger.debug(f"Clearing template audit leader row {row}")
                    # Clear the row but preserve column C's label
                    for col in [1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                        cell_ref = f'{get_column_letter(col)}{row}'
                        try:
                            cell = ws[cell_ref]
                            if not isinstance(cell, MergedCell):
                                cell.value = None
                        except:
                            pass

    # ==========================================
    # COMMON UTILITIES AND DATA PROCESSING
    # ==========================================

    def _safe_write_cell(self, ws, cell_ref: str, value: Any, **kwargs):
        """Safely write to a cell, handling merged cells properly."""
        cell = ws[cell_ref]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    ws[top_left.coordinate].value = value
                    for attr, attr_value in kwargs.items():
                        setattr(ws[top_left.coordinate], attr, attr_value)
                    return

        cell.value = value
        for attr, attr_value in kwargs.items():
            setattr(cell, attr, attr_value)

    def _safe_read_cell(self, ws, cell_ref: str):
        """Safely read from a cell, handling merged cells properly."""
        cell = ws[cell_ref]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    return ws[top_left.coordinate].value

        return cell.value

    # ==========================================
    # MAIN GENERATION METHOD
    # ==========================================

    def generate_summary_report(self,
                               results: Dict[str, Any],
                               rule_results: Dict[str, Any],
                               output_path: str,
                               responsible_party_column: str = "Audit Leader",
                               review_year: Optional[str] = None,
                               **kwargs) -> str:
        """Generate a dynamic summary report using the template."""

        logger.info(f"Generating dynamic summary report for {len(rule_results)} rules")

        # Copy template and load workbook
        shutil.copy2(self.template_path, output_path)
        # Load with data_only=True to avoid formula corruption
        wb = openpyxl.load_workbook(output_path, data_only=True)
        
        # Remove external links to prevent corruption
        if hasattr(wb, 'external_links') and wb.external_links:
            wb.external_links.clear()
        ws = wb.active

        # IMPORTANT: Clear template placeholders BEFORE adding dynamic content
        self._clear_template_placeholders(ws)

        # Extract dynamic structure from validation data
        rule_names, audit_leaders, leader_rule_matrix = self._extract_dynamic_structure(
            rule_results, results, responsible_party_column
        )

        # Update header with review year
        self._update_header(ws, review_year)

        # Process each section
        self._process_section1_iag_overall(ws, rule_results, rule_names, leader_rule_matrix)
        self._process_section2_audit_leader_overall(ws, rule_names, audit_leaders, leader_rule_matrix)
        self._process_section3_detailed_test_results(ws, rule_names, audit_leaders, leader_rule_matrix)

        # Final cleanup and save
        self._cleanup_none_values(ws)
        wb.save(output_path)
        wb.close()

        logger.info(f"Generated dynamic summary report: {output_path}")
        return output_path

    # ==========================================
    # SECTION 3: AUDIT LEADER AVERAGE TEST RESULTS (FIXED)
    # ==========================================

    def _process_section3_detailed_test_results(self, ws, rule_names: List[str],
                                               audit_leaders: List[str], leader_rule_matrix: Dict[str, Dict]):
        """Process Section 3: Detailed Test Results."""
        logger.info("Processing Section 3: Audit Leader Average Test Results")

        # Setup Section 3 structure
        section3_start_row = self._setup_section3_structure(ws, rule_names, audit_leaders)

        # Add blue analytics section
        self._add_section3_analytics_rows(ws, rule_names, section3_start_row)

        # Populate detailed grid with results
        self._populate_section3_detailed_grid(ws, rule_names, audit_leaders, leader_rule_matrix, section3_start_row)

    def _setup_section3_structure(self, ws, rule_names: List[str], audit_leaders: List[str]) -> int:
        """Setup the structure for Section 3 and return starting row."""
        # Find where Section 3 should start (after Section 2)
        # This is dynamic based on the number of audit leaders
        section3_start_row = 35 + len(audit_leaders)
        
        # Check if we've already cleared template content here
        existing_content = self._safe_read_cell(ws, f'A{section3_start_row}')
        if existing_content and "Audit Leader Average Test Results" in str(existing_content):
            # Template content wasn't properly cleared, clear it now
            logger.warning(f"Found residual template content at row {section3_start_row}, clearing it")
            for col in range(1, 20):
                cell_ref = f'{get_column_letter(col)}{section3_start_row}'
                try:
                    cell = ws[cell_ref]
                    if not isinstance(cell, MergedCell):
                        cell.value = None
                except:
                    pass

        # Add section title
        ws[f'A{section3_start_row}'] = "Audit Leader Average Test Results"
        ws[f'A{section3_start_row}'].font = self.header_font

        # Add headers
        header_row = section3_start_row + 2
        ws[f'B{header_row}'] = "Audit Leader"
        ws[f'C{header_row}'] = "Samples Tested for Audit Leader"

        # Add rule headers
        start_col = 4
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(start_col + i)
            ws[f'{col_letter}{header_row}'] = rule_name
            ws[f'{col_letter}{header_row}'].font = self.header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.border
            ws.column_dimensions[col_letter].width = 20

        # Add aggregate headers
        agg_start_col = start_col + len(rule_names)
        agg_headers = ["GC Count", "PC Count", "DNC Count", "Total Applicable Count", "Average Score", "Average Rating: 4 Point Scale"]

        for i, header in enumerate(agg_headers):
            col_letter = get_column_letter(agg_start_col + i)
            ws[f'{col_letter}{header_row}'] = header
            ws[f'{col_letter}{header_row}'].font = self.header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.border
            ws.column_dimensions[col_letter].width = 15

        return header_row + 1  # Return data starting row

    # Note: Include all other methods from the original file that aren't shown here
    # This is just showing the key changes needed to fix the duplication issue