#!/usr/bin/env python3
"""
Template-free summary report generator.
Creates IAG and AL Results and Ratings report without using any template.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from typing import Dict, List, Tuple, Any
import datetime

logger = logging.getLogger(__name__)


class TemplateFreeSummaryGenerator:
    """Generate summary reports from scratch without templates."""
    
    def __init__(self):
        """Initialize styles."""
        self.header_font = Font(bold=True, size=11)
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    def generate_summary_report(self, rule_results: Dict, output_path: str, 
                               responsible_party_column: str = "Responsible Party") -> str:
        """Generate a complete summary report from scratch."""
        logger.info(f"Generating template-free summary report with {len(rule_results)} rules")
        
        # Extract data structure
        rule_names, audit_leaders, leader_rule_matrix = self._extract_structure(
            rule_results, responsible_party_column
        )
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IAG and AL Results and Ratings"
        
        # Build the report sections
        current_row = 1
        current_row = self._create_header(ws, current_row)
        current_row = self._create_section1_iag_overall(ws, current_row, rule_results, rule_names, leader_rule_matrix)
        current_row = self._create_section2_audit_leader_overall(ws, current_row, rule_names, audit_leaders, leader_rule_matrix)
        current_row = self._create_section3_detailed_results(ws, current_row, rule_names, audit_leaders, leader_rule_matrix)
        
        # Save workbook
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Generated template-free summary report: {output_path}")
        return output_path
    
    def _extract_structure(self, rule_results: Dict, responsible_party_column: str) -> Tuple[List[str], List[str], Dict]:
        """Extract rules, leaders, and matrix from results."""
        rule_names = []
        audit_leaders = set()
        leader_rule_matrix = {}
        
        for rule_id, eval_result in rule_results.items():
            rule_name = eval_result.rule.name if hasattr(eval_result, 'rule') else rule_id
            rule_names.append(rule_name)
            
            if hasattr(eval_result, 'party_results') and eval_result.party_results:
                for leader, party_data in eval_result.party_results.items():
                    audit_leaders.add(leader)
                    if leader not in leader_rule_matrix:
                        leader_rule_matrix[leader] = {}
                    
                    metrics = party_data.get('metrics', {}) if isinstance(party_data, dict) else {}
                    leader_rule_matrix[leader][rule_name] = {
                        'gc_count': metrics.get('gc_count', 0),
                        'pc_count': metrics.get('pc_count', 0),
                        'dnc_count': metrics.get('dnc_count', 0),
                        'na_count': metrics.get('na_count', 0),
                        'total_count': metrics.get('total_count', 0),
                        'status': party_data.get('status', 'N/A') if isinstance(party_data, dict) else 'N/A',
                        'error_rate': metrics.get('dnc_rate', 0),
                        'threshold': getattr(eval_result.rule, 'threshold', 0.02) if hasattr(eval_result, 'rule') else 0.02
                    }
        
        return sorted(rule_names), sorted(list(audit_leaders)), leader_rule_matrix
    
    def _create_header(self, ws, start_row: int) -> int:
        """Create report header."""
        ws.merge_cells(f'B{start_row}:J{start_row}')
        ws[f'B{start_row}'] = f"QA {datetime.datetime.now().year} QA Review and Summary Report"
        ws[f'B{start_row}'].font = Font(bold=True, size=14)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='center')
        
        return start_row + 3
    
    def _create_section1_iag_overall(self, ws, start_row: int, rule_results: Dict, 
                                    rule_names: List[str], leader_rule_matrix: Dict) -> int:
        """Create Section 1: IAG Overall Results."""
        # Section header
        ws[f'B{start_row}'] = "IAG Overall Results and Rating"
        ws[f'B{start_row}'].font = Font(bold=True, size=12)
        
        # Row labels
        labels = ["Total Score", "GC Score", "PC Score", "DNC Score", 
                 "Total Count of Applicable Tests Across Audit Leaders",
                 "Weighted Score Across Audit Leaders",
                 "Weighted Rating Across Audit Leaders"]
        
        for i, label in enumerate(labels):
            ws[f'B{start_row + 1 + i}'] = label
            ws[f'B{start_row + 1 + i}'].font = self.data_font
        
        # Calculate totals
        total_gc = total_pc = total_dnc = total_na = 0
        for rule_id, eval_result in rule_results.items():
            if hasattr(eval_result, 'compliance_metrics'):
                metrics = eval_result.compliance_metrics
                total_gc += metrics.get('gc_count', 0)
                total_pc += metrics.get('pc_count', 0)
                total_dnc += metrics.get('dnc_count', 0)
                total_na += metrics.get('na_count', 0)
        
        # Add rule columns
        col_start = 4  # Column D
        for i, rule_name in enumerate(rule_names):
            col = get_column_letter(col_start + i)
            ws[f'{col}{start_row + 1}'] = rule_name
            ws[f'{col}{start_row + 1}'].font = self.header_font
            ws[f'{col}{start_row + 1}'].alignment = self.center_align
            ws.column_dimensions[col].width = 15
        
        # Add summary columns
        summary_col = col_start + len(rule_names)
        summary_headers = ["Overall IAG Score and Rating", "", "Volume of Sampled Audit Entities for IAG"]
        
        for i, header in enumerate(summary_headers):
            col = get_column_letter(summary_col + i)
            ws[f'{col}{start_row}'] = header
            ws[f'{col}{start_row}'].font = self.header_font
            ws.column_dimensions[col].width = 20
        
        # Fill in IAG totals
        ws[f'D{start_row + 2}'] = total_gc
        ws[f'D{start_row + 3}'] = total_pc
        ws[f'D{start_row + 4}'] = total_dnc
        ws[f'D{start_row + 5}'] = total_gc + total_pc + total_dnc
        
        return start_row + 10
    
    def _create_section2_audit_leader_overall(self, ws, start_row: int, rule_names: List[str],
                                            audit_leaders: List[str], leader_rule_matrix: Dict) -> int:
        """Create Section 2: Audit Leader Overall Results."""
        # Section header
        ws[f'B{start_row}'] = "Audit Leader Overall Results and Ratings"
        ws[f'B{start_row}'].font = Font(bold=True, size=12)
        
        # Headers
        ws[f'B{start_row + 1}'] = "Audit Leader"
        ws[f'C{start_row + 1}'] = "Measurement Description"
        
        # Rule headers
        col_start = 4
        for i, rule_name in enumerate(rule_names):
            col = get_column_letter(col_start + i)
            ws[f'{col}{start_row + 1}'] = rule_name
            ws[f'{col}{start_row + 1}'].font = self.header_font
            ws[f'{col}{start_row + 1}'].alignment = self.center_align
        
        # Summary headers
        summary_col = col_start + len(rule_names)
        summary_headers = ["Weighted Score", "Weighted Average Rating: 4 Point Scale", 
                          "Volume of Sampled Audit Entities by AL"]
        
        for i, header in enumerate(summary_headers):
            col = get_column_letter(summary_col + i)
            ws[f'{col}{start_row + 1}'] = header
            ws[f'{col}{start_row + 1}'].font = self.header_font
            ws.column_dimensions[col].width = 20
        
        # Add audit leader data
        for i, leader in enumerate(audit_leaders):
            row = start_row + 2 + i
            ws[f'B{row}'] = leader
            ws[f'C{row}'] = "Total Weighted Score"
            
            # Calculate scores for each rule
            for j, rule_name in enumerate(rule_names):
                col = get_column_letter(col_start + j)
                if leader in leader_rule_matrix and rule_name in leader_rule_matrix[leader]:
                    rule_data = leader_rule_matrix[leader][rule_name]
                    weighted_score = (rule_data.get('gc_count', 0) * 5 +
                                    rule_data.get('pc_count', 0) * 3 +
                                    rule_data.get('dnc_count', 0) * 1)
                    ws[f'{col}{row}'] = weighted_score
                else:
                    ws[f'{col}{row}'] = 0
        
        return start_row + len(audit_leaders) + 5
    
    def _create_section3_detailed_results(self, ws, start_row: int, rule_names: List[str],
                                         audit_leaders: List[str], leader_rule_matrix: Dict) -> int:
        """Create Section 3: Detailed Test Results."""
        # Section title
        ws[f'A{start_row}'] = "Audit Leader Average Test Results"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        # Blue analytics section
        analytics_row = start_row + 2
        col_start = 4
        
        for i, rule_name in enumerate(rule_names):
            col = get_column_letter(col_start + i)
            # Analytics header
            ws[f'{col}{analytics_row}'] = "Analytics"
            ws[f'{col}{analytics_row}'].fill = self.blue_fill
            ws[f'{col}{analytics_row}'].font = Font(color="FFFFFF", bold=True)
            
            # Placeholder values
            ws[f'{col}{analytics_row + 2}'] = "2%"
            ws[f'{col}{analytics_row + 3}'] = "3"
            for j in range(3):
                ws[f'{col}{analytics_row + 4 + j}'] = "Not Applicable"
            ws[f'{col}{analytics_row + 7}'] = rule_name
            
            # Apply blue formatting
            for offset in range(8):
                ws[f'{col}{analytics_row + offset}'].fill = self.blue_fill
                ws[f'{col}{analytics_row + offset}'].font = Font(color="FFFFFF")
        
        # Data headers
        data_header_row = analytics_row + 10
        ws[f'B{data_header_row}'] = "Audit Leader"
        ws[f'C{data_header_row}'] = "Samples Tested for Audit Leader"
        
        # Rule headers
        for i, rule_name in enumerate(rule_names):
            col = get_column_letter(col_start + i)
            ws[f'{col}{data_header_row}'] = rule_name
            ws[f'{col}{data_header_row}'].font = self.header_font
        
        # Aggregate headers
        agg_col = col_start + len(rule_names)
        agg_headers = ["GC Count", "PC Count", "DNC Count", "Total Applicable Count", 
                      "Average Score", "Average Rating: 4 Point Scale"]
        
        for i, header in enumerate(agg_headers):
            col = get_column_letter(agg_col + i)
            ws[f'{col}{data_header_row}'] = header
            ws[f'{col}{data_header_row}'].font = self.header_font
            ws.column_dimensions[col].width = 15
        
        # Add audit leader data
        for i, leader in enumerate(audit_leaders):
            row = data_header_row + 1 + i
            ws[f'B{row}'] = leader
            ws[f'C{row}'] = ""  # Manual entry
            
            total_gc = total_pc = total_dnc = 0
            
            # Status for each rule
            for j, rule_name in enumerate(rule_names):
                col = get_column_letter(col_start + j)
                if leader in leader_rule_matrix and rule_name in leader_rule_matrix[leader]:
                    status = leader_rule_matrix[leader][rule_name].get('status', 'N/A')
                    ws[f'{col}{row}'] = status
                    
                    # Apply color
                    if status == 'GC':
                        ws[f'{col}{row}'].fill = self.green_fill
                    elif status == 'PC':
                        ws[f'{col}{row}'].fill = self.yellow_fill
                    elif status == 'DNC':
                        ws[f'{col}{row}'].fill = self.red_fill
                    
                    # Add to totals
                    rule_data = leader_rule_matrix[leader][rule_name]
                    total_gc += rule_data.get('gc_count', 0)
                    total_pc += rule_data.get('pc_count', 0)
                    total_dnc += rule_data.get('dnc_count', 0)
                else:
                    ws[f'{col}{row}'] = 'N/A'
                
                ws[f'{col}{row}'].alignment = self.center_align
            
            # Aggregate columns
            ws[f'{get_column_letter(agg_col)}{row}'] = total_gc
            ws[f'{get_column_letter(agg_col + 1)}{row}'] = total_pc
            ws[f'{get_column_letter(agg_col + 2)}{row}'] = total_dnc
            ws[f'{get_column_letter(agg_col + 3)}{row}'] = total_gc + total_pc + total_dnc
            
            # Average score and rating
            total = total_gc + total_pc + total_dnc
            if total > 0:
                avg_score = (total_gc * 5 + total_pc * 3 + total_dnc * 1) / total
                ws[f'{get_column_letter(agg_col + 4)}{row}'] = round(avg_score, 2)
                
                if avg_score >= 4:
                    rating = "GC"
                elif avg_score >= 2.5:
                    rating = "PC"
                else:
                    rating = "DNC"
                ws[f'{get_column_letter(agg_col + 5)}{row}'] = rating
            else:
                ws[f'{get_column_letter(agg_col + 4)}{row}'] = 0
                ws[f'{get_column_letter(agg_col + 5)}{row}'] = "N/A"
        
        return row + 3


# Integration function
def generate_template_free_summary(rule_results: Dict, output_path: str, 
                                  responsible_party_column: str = "Responsible Party", **kwargs) -> str:
    """Generate summary report without using any template."""
    generator = TemplateFreeSummaryGenerator()
    return generator.generate_summary_report(rule_results, output_path, responsible_party_column)