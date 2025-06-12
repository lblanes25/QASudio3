"""
Template Integration Mixin for ReportGenerator - Complete Rewrite for DynamicQAReportGenerator
Provides template-based report generation capabilities with new dynamic generator integration
"""

import os
import logging
import datetime
import tempfile
import shutil
import time
import re
from typing import Dict, List, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# Import template processors
try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from .template_report_generator import TemplateBasedReportGenerator
    from .dynamic_qa_report_generator import DynamicQAReportGenerator
    TEMPLATE_PROCESSORS_AVAILABLE = True
except ImportError as e:
    logger.warning(f"Template processors not available: {e}")
    TEMPLATE_PROCESSORS_AVAILABLE = False


class TemplateIntegrationMixin:
    """
    Mixin class that adds template-based generation capabilities to ReportGenerator.
    Now prioritizes DynamicQAReportGenerator for summary reports.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._init_template_integration()

    def _init_template_integration(self):
        """Initialize template integration with new dynamic generator priority"""

        # Initialize all generators
        self.template_generator = None
        self.dynamic_qa_generator = None
        self.individual_template_processor = None

        # Get template paths from config
        template_config = getattr(self, 'config', {}).get('template_config', {})

        # Set default paths
        self.summary_template_path = template_config.get(
            'summary_template_path',
            'qa_summary_template.xlsx'
        )

        self.individual_template_path = template_config.get(
            'individual_template_path',
            'qa_individual_analytic_template.xlsx'
        )

        # Enable template usage by default if processors available
        self.use_templates = template_config.get('use_templates', TEMPLATE_PROCESSORS_AVAILABLE)

        # Initialize processors if enabled
        if self.use_templates and TEMPLATE_PROCESSORS_AVAILABLE:
            try:
                # Initialize the new dynamic QA report generator (highest priority)
                self.dynamic_qa_generator = DynamicQAReportGenerator()
                logger.info("DynamicQAReportGenerator initialized successfully")

                # Initialize the legacy template generator as fallback
                self.template_generator = TemplateBasedReportGenerator(
                    template_path=self.summary_template_path if os.path.exists(self.summary_template_path) else None
                )
                logger.info("Legacy template-based report generator initialized")

                # Initialize individual template processor if template exists
                self._init_individual_template_processor()

            except Exception as e:
                logger.error(f"Error initializing template processors: {e}")
                self.use_templates = False
                self.dynamic_qa_generator = None

    def _init_individual_template_processor(self):
        """Initialize the individual template processor"""
        if not os.path.exists(self.individual_template_path):
            logger.info(f"Individual template not found: {self.individual_template_path}")
            return

        try:
            # Try to import the dynamic template processor
            try:
                from .dynamic_individual_template_processor_v5 import DynamicIndividualTemplateProcessorV5
                self.individual_template_processor = DynamicIndividualTemplateProcessorV5(self.individual_template_path)
                logger.info("Individual template processor initialized successfully")
            except ImportError:
                # Try alternative import path
                import sys
                project_root = Path(__file__).parent.parent.parent
                if str(project_root) not in sys.path:
                    sys.path.insert(0, str(project_root))
                from dynamic_individual_template_processor_v5 import DynamicIndividualTemplateProcessorV5
                self.individual_template_processor = DynamicIndividualTemplateProcessorV5(self.individual_template_path)
                logger.info("Individual template processor initialized via alternative path")
        except ImportError:
            logger.warning("DynamicIndividualTemplateProcessorV5 not available")
        except Exception as e:
            logger.warning(f"Could not initialize individual template processor: {e}")

    def generate_excel_with_template(self, results: Dict[str, Any], rule_results: Dict[str, Any],
                                   output_path: str, group_by: Optional[str] = None) -> str:
        """
        Generate Excel report using template processor with individual tabs.
        Now prioritizes DynamicQAReportGenerator for the main report.

        Args:
            results: Validation results
            rule_results: Rule evaluation results
            output_path: Output file path
            group_by: Grouping column (responsible party)

        Returns:
            Path to generated report
        """
        logger.info(f"Generating template-based Excel report with individual tabs at {output_path}")

        responsible_party_column = group_by or "Audit Leader"

        # Method 1: Try DynamicQAReportGenerator (preferred)
        generated_path = self._try_dynamic_qa_generation(
            results, rule_results, output_path, responsible_party_column
        )

        if generated_path:
            # Add individual tabs to the generated report
            self._add_individual_tabs_to_report(
                output_path=generated_path,
                rule_results=rule_results,
                responsible_party_column=responsible_party_column
            )
            return generated_path

        # Method 2: Try legacy template-based generation
        generated_path = self._try_legacy_template_generation(
            results, rule_results, output_path, responsible_party_column
        )

        if generated_path:
            self._add_individual_tabs_to_report(
                output_path=generated_path,
                rule_results=rule_results,
                responsible_party_column=responsible_party_column
            )
            return generated_path

        # Method 3: Create basic fallback report
        generated_path = self._create_basic_fallback_report(output_path, results)

        logger.info(f"Template-based report generated: {generated_path}")
        return generated_path

    def _try_dynamic_qa_generation(self, results: Dict[str, Any], rule_results: Dict[str, Any],
                                 output_path: str, responsible_party_column: str) -> Optional[str]:
        """Try generating report using DynamicQAReportGenerator"""
        if not self.dynamic_qa_generator:
            return None

        try:
            logger.info("Using DynamicQAReportGenerator for main report")
            generated_path = self.dynamic_qa_generator.generate_report(
                rule_results=rule_results,
                output_path=output_path,
                responsible_party_column=responsible_party_column,
                review_year=results.get('review_year', None)
            )
            logger.info("Successfully generated report using DynamicQAReportGenerator")
            return generated_path
        except Exception as e:
            logger.error(f"DynamicQAReportGenerator failed: {e}")
            return None

    def _try_legacy_template_generation(self, results: Dict[str, Any], rule_results: Dict[str, Any],
                                      output_path: str, responsible_party_column: str) -> Optional[str]:
        """Try generating report using legacy template methods"""
        try:
            # Try dynamic summary template processor
            dynamic_summary_available = self._try_import_dynamic_summary_processor()

            if dynamic_summary_available:
                from .template_free_summary_generator import generate_template_free_summary
                generated_path = generate_template_free_summary(
                    rule_results=rule_results,
                    output_path=output_path,
                    responsible_party_column=responsible_party_column
                )
                logger.info("Generated report using dynamic summary template processor")
                return generated_path

            # Fall back to original template generator
            if self.template_generator:
                generated_path = self.template_generator.generate_excel_from_template(
                    results=results,
                    rule_results=rule_results,
                    output_path=output_path,
                    analytic_id="Analytics Summary",
                    analytic_title="QA Analytics Validation Report",
                    group_by=responsible_party_column
                )
                logger.info("Generated report using legacy template generator")
                return generated_path

        except Exception as e:
            logger.error(f"Legacy template generation failed: {e}")

        return None

    def _try_import_dynamic_summary_processor(self) -> bool:
        """Try to import dynamic summary processor"""
        try:
            from .dynamic_summary_template_processor import DynamicSummaryTemplateProcessor
            return True
        except ImportError:
            try:
                import sys
                project_root = Path(__file__).parent.parent.parent
                if str(project_root) not in sys.path:
                    sys.path.insert(0, str(project_root))
                from dynamic_summary_template_processor import DynamicSummaryTemplateProcessor
                return True
            except ImportError:
                logger.warning("Dynamic summary processor not available")
                return False

    def _create_basic_fallback_report(self, output_path: str, results: Dict[str, Any]) -> str:
        """Create a basic fallback Excel report"""
        logger.warning("Creating basic fallback report")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws['A1'] = "QA Analytics Report"
            ws['A2'] = f"Generated: {datetime.datetime.now()}"
            ws['A3'] = "Error: Advanced report generation failed"
            ws['A4'] = f"Status: {results.get('status', 'Unknown')}"
            ws['A5'] = f"Total Rules: {len(results.get('rule_results', {}))}"
            wb.save(output_path)
            wb.close()
            return output_path
        except Exception as e:
            logger.error(f"Even basic fallback failed: {e}")
            # Create empty file as last resort
            with open(output_path, 'w') as f:
                f.write('Report generation failed')
            return output_path

    def generate_standalone_summary_report(self, results: Dict[str, Any], rule_results: Dict[str, Any],
                                         output_path: str, group_by: Optional[str] = None, **kwargs) -> str:
        """
        Generate standalone summary report using DynamicQAReportGenerator.
        This is the new recommended method for summary-only reports.

        Args:
            results: Validation results
            rule_results: Rule evaluation results
            output_path: Output file path
            group_by: Grouping column (responsible party)
            **kwargs: Additional parameters (review_year, etc.)

        Returns:
            Path to generated report
        """
        if not self.dynamic_qa_generator:
            raise RuntimeError("DynamicQAReportGenerator not available")

        logger.info(f"Generating standalone summary report at {output_path}")

        responsible_party_column = group_by or "Audit Leader"

        generated_path = self.dynamic_qa_generator.generate_report(
            rule_results=rule_results,
            output_path=output_path,
            responsible_party_column=responsible_party_column,
            review_year=kwargs.get('review_year', results.get('review_year', None)),
            **kwargs
        )

        logger.info(f"Standalone summary report generated: {generated_path}")
        return generated_path

    def _add_individual_tabs_to_report(self, output_path: str, rule_results: Dict[str, Any],
                                     responsible_party_column: str):
        """
        Add individual analytic tabs to the existing Excel report.

        Args:
            output_path: Path to the existing Excel file
            rule_results: Rule evaluation results
            responsible_party_column: Grouping column
        """
        if not self.individual_template_processor:
            logger.warning("Cannot add individual tabs - individual template processor not available")
            return

        try:
            # Load the existing workbook
            wb = openpyxl.load_workbook(output_path)
            existing_sheets = [ws.title for ws in wb.worksheets]

            # Track successful tab creation
            successful_tabs = 0
            failed_tabs = []

            # Process each rule
            for rule_id, rule_result in rule_results.items():
                try:
                    # Create safe worksheet name
                    safe_name = self._create_safe_sheet_name(rule_id, rule_result, existing_sheets)

                    logger.debug(f"Creating individual tab for {rule_id} -> {safe_name}")

                    # Check if rule has required data
                    if not self._validate_rule_data(rule_result):
                        logger.warning(f"Rule {rule_id} missing required data for template generation")
                        failed_tabs.append(rule_id)
                        continue

                    # Generate individual report using template
                    temp_path = self._generate_individual_report_temp(
                        rule_id, rule_result, responsible_party_column
                    )

                    if temp_path:
                        # Copy the generated worksheet to main workbook
                        self._copy_worksheet_to_workbook(wb, temp_path, safe_name)
                        successful_tabs += 1
                        existing_sheets.append(safe_name)

                        # Clean up temp file
                        self._cleanup_temp_file(temp_path)
                    else:
                        failed_tabs.append(rule_id)

                except Exception as e:
                    logger.error(f"Error creating individual tab for {rule_id}: {e}")
                    failed_tabs.append(rule_id)
                    continue

            # Save the workbook
            wb.save(output_path)
            wb.close()

            logger.info(f"Successfully added {successful_tabs} individual tabs to report")
            if failed_tabs:
                logger.warning(f"Failed to create tabs for rules: {failed_tabs}")

        except Exception as e:
            logger.error(f"Error adding individual tabs to report: {e}")
            import traceback
            logger.debug(f"Full traceback: {traceback.format_exc()}")

    def _create_safe_sheet_name(self, rule_id: str, rule_result: Any, existing_sheets: List[str]) -> str:
        """Create a safe Excel sheet name"""
        # Get rule name/title
        if hasattr(rule_result, 'rule') and hasattr(rule_result.rule, 'name'):
            rule_name = rule_result.rule.name
        else:
            rule_name = rule_id

        # Create the tab name
        tab_name = f"{rule_id} - {rule_name}"

        # Make it safe for Excel (31 char limit, no invalid chars)
        safe_name = re.sub(r'[^\w\-_ ]', '', tab_name)[:31]

        # Ensure uniqueness
        if safe_name in existing_sheets:
            base_name = safe_name[:28]
            suffix = 1
            while f"{base_name}_{suffix}" in existing_sheets:
                suffix += 1
            safe_name = f"{base_name}_{suffix}"

        return safe_name

    def _validate_rule_data(self, rule_result: Any) -> bool:
        """Validate that rule has required data for template generation"""
        return (hasattr(rule_result, 'result_df') and
                rule_result.result_df is not None and
                not rule_result.result_df.empty)

    def _generate_individual_report_temp(self, rule_id: str, rule_result: Any,
                                       responsible_party_column: str) -> Optional[str]:
        """Generate individual report in temporary file"""
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)  # Close the file descriptor immediately

        try:
            logger.debug(f"Generating individual report for {rule_id} at {temp_path}")
            self.individual_template_processor.generate_individual_report(
                rule_id=rule_id,
                rule_result=rule_result,
                output_path=temp_path,
                responsible_party_column=responsible_party_column
            )
            logger.debug(f"Successfully generated individual report for {rule_id}")
            return temp_path
        except Exception as e:
            logger.error(f"Failed to generate individual report for {rule_id}: {e}")
            self._cleanup_temp_file(temp_path)
            return None

    def _copy_worksheet_to_workbook(self, wb: openpyxl.Workbook, temp_path: str, safe_name: str):
        """Copy worksheet from temporary file to main workbook"""
        logger.debug(f"Loading generated report from {temp_path}")
        temp_wb = openpyxl.load_workbook(temp_path, data_only=True)
        temp_ws = temp_wb.active

        logger.debug(f"Creating new sheet with title: {safe_name}")
        ws = wb.create_sheet(title=safe_name)

        # Copy merged cells first
        logger.debug(f"Copying {len(temp_ws.merged_cells.ranges)} merged cell ranges")
        for merged_range in temp_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_range))

        # Copy all cells from template worksheet
        logger.debug("Copying cells from template worksheet")
        for row in temp_ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                try:
                    new_cell = ws.cell(row=cell.row, column=cell.column)

                    # Copy value
                    new_cell.value = cell.value

                    # Copy styles
                    if cell.has_style:
                        if cell.font:
                            new_cell.font = cell.font.copy()
                        if cell.fill and hasattr(cell.fill, 'copy'):
                            new_cell.fill = cell.fill.copy()
                        if cell.border:
                            new_cell.border = cell.border.copy()
                        if cell.alignment:
                            new_cell.alignment = cell.alignment.copy()
                        if cell.number_format:
                            new_cell.number_format = cell.number_format
                except Exception as cell_error:
                    logger.error(f"Error processing cell {cell.coordinate}: {cell_error}")
                    raise

        # Copy column and row dimensions
        for col_letter, col_dim in temp_ws.column_dimensions.items():
            ws.column_dimensions[col_letter].width = col_dim.width

        for row_num, row_dim in temp_ws.row_dimensions.items():
            ws.row_dimensions[row_num].height = row_dim.height

        # Close temp workbook
        temp_wb.close()
        del temp_wb

    def _cleanup_temp_file(self, temp_path: str):
        """Clean up temporary file with retry logic for Windows"""
        for attempt in range(3):
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.1)  # Wait 100ms and retry
                else:
                    logger.warning(f"Could not delete temp file {temp_path}, will be cleaned up later")

    def _create_basic_individual_tab(self, ws, rule_id: str, rule_result):
        """Create a basic individual tab when full data is not available"""
        ws['C1'] = f"Individual Analytic Report - {rule_id}"
        ws['C2'] = rule_id
        ws['C4'] = getattr(rule_result, 'description', f"Analysis for rule {rule_id}")

        # Summary results
        ws['A13'] = "Summary Results"
        ws['A15'] = "Generally Conforms (GC)"
        ws['A16'] = "Partially Conforms (PC)"
        ws['A17'] = "Does Not Conform (DNC)"

        # Add basic counts if available
        if hasattr(rule_result, 'compliance_counts'):
            counts = rule_result.compliance_counts
            ws['E15'] = counts.get('GC', 0)
            ws['E16'] = counts.get('PC', 0)
            ws['E17'] = counts.get('DNC', 0)

    def generate_individual_reports_with_template(self, rule_results: Dict[str, Any],
                                                output_dir: str, responsible_party_column: str = "Audit Leader") -> List[str]:
        """
        Generate individual analytic reports for all rules using templates.

        Args:
            rule_results: Dictionary of rule_id -> RuleEvaluationResult
            output_dir: Directory to save individual reports
            responsible_party_column: Grouping column

        Returns:
            List of generated file paths
        """
        if not self.individual_template_processor:
            logger.warning("Individual template processor not available")
            return []

        logger.info(f"Generating individual reports for {len(rule_results)} rules")

        # Create output directory
        os.makedirs(output_dir, exist_ok=True)

        generated_paths = []
        failed_rules = []

        for rule_id, rule_result in rule_results.items():
            try:
                # Generate filename
                safe_rule_name = re.sub(r'[^\w\-_]', '_', rule_id)
                output_path = os.path.join(
                    output_dir,
                    f"individual_analytic_{safe_rule_name}.xlsx"
                )

                # Generate report
                path = self.individual_template_processor.generate_individual_report(
                    rule_id=rule_id,
                    rule_result=rule_result,
                    output_path=output_path,
                    responsible_party_column=responsible_party_column,
                    generation_date=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )

                generated_paths.append(path)
                logger.info(f"Generated individual report for {rule_id}: {os.path.basename(path)}")

            except Exception as e:
                logger.error(f"Error generating individual report for {rule_id}: {e}")
                failed_rules.append(rule_id)
                continue

        logger.info(f"Generated {len(generated_paths)} individual reports")
        if failed_rules:
            logger.warning(f"Failed to generate reports for rules: {failed_rules}")

        return generated_paths

    def validate_templates(self) -> Dict[str, bool]:
        """
        Validate that required templates exist and are accessible.

        Returns:
            Dictionary with validation results
        """
        validation_results = {
            'summary_template_exists': False,
            'summary_template_valid': False,
            'individual_template_exists': False,
            'individual_template_valid': False,
            'processors_available': TEMPLATE_PROCESSORS_AVAILABLE,
            'dynamic_qa_generator_available': self.dynamic_qa_generator is not None,
            'individual_processor_available': self.individual_template_processor is not None
        }

        # Check summary template
        if self.summary_template_path:
            validation_results['summary_template_exists'] = os.path.exists(self.summary_template_path)
            if validation_results['summary_template_exists']:
                try:
                    wb = openpyxl.load_workbook(self.summary_template_path, read_only=True)
                    wb.close()
                    validation_results['summary_template_valid'] = True
                except Exception as e:
                    logger.error(f"Summary template validation failed: {e}")

        # Check individual template
        if self.individual_template_path:
            validation_results['individual_template_exists'] = os.path.exists(self.individual_template_path)
            if validation_results['individual_template_exists']:
                try:
                    wb = openpyxl.load_workbook(self.individual_template_path, read_only=True)
                    wb.close()
                    validation_results['individual_template_valid'] = True
                except Exception as e:
                    logger.error(f"Individual template validation failed: {e}")

        return validation_results

    def get_available_generators(self) -> Dict[str, bool]:
        """
        Get status of available report generators.

        Returns:
            Dictionary indicating which generators are available
        """
        return {
            'dynamic_qa_generator': self.dynamic_qa_generator is not None,
            'legacy_template_generator': self.template_generator is not None,
            'individual_template_processor': self.individual_template_processor is not None,
            'dynamic_summary_processor': self._try_import_dynamic_summary_processor()
        }

    def set_generator_preference(self, prefer_dynamic: bool = True):
        """
        Set preference for which generator to use.

        Args:
            prefer_dynamic: Whether to prefer DynamicQAReportGenerator
        """
        if not hasattr(self, 'config'):
            self.config = {}
        if 'display_options' not in self.config:
            self.config['display_options'] = {}

        self.config['display_options']['prefer_dynamic_qa_generator'] = prefer_dynamic
        logger.info(f"Generator preference set to: {'Dynamic' if prefer_dynamic else 'Legacy'}")