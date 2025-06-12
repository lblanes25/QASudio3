"""
Dynamic QA Report Generator - Creates Excel reports from scratch without templates
Generates professional QA audit summary reports with dynamic structure.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple

logger = logging.getLogger(__name__)


class DynamicQAReportGenerator:
    """
    Generates QA audit summary reports entirely from scratch in Excel format.
    No template files required - builds everything programmatically.
    """

    def __init__(self):
        """Initialize the report generator with styling configuration."""
        self._setup_styles()
        self._setup_layout_constants()

    def _setup_styles(self):
        """Define all visual styles for professional report formatting."""

        # Fonts
        self.title_font = Font(bold=True, size=12, color="FFFFFF")
        self.section_header_font = Font(bold=True, size=11, color="FFFFFF")
        self.column_header_font = Font(bold=True, size=10)
        self.data_font = Font(size=10)
        self.label_font = Font(size=10)

        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Color fills
        self.blue_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        self.blue_analytics_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.green_status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Number formats
        self.number_format = '#,##0'
        self.decimal_format = '0.00'
        self.percentage_format = '0.0%'

    def _setup_layout_constants(self):
        """Define layout spacing and positioning constants."""
        self.section1_start_row = 5
        self.section_spacing = 3
        self.section3_analytics_rows = 8

        # Column widths
        self.label_column_width = 25
        self.test_column_width = 15
        self.summary_column_width = 20
        self.leader_column_width = 20

    def generate_report(self,
                       rule_results: Dict[str, Any],
                       output_path: str,
                       responsible_party_column: str = "AuditLeader",
                       review_year: Optional[str] = None,
                       **kwargs) -> str:
        """
        Generate a complete QA summary report from scratch.

        Args:
            rule_results: Dictionary of rule_id → RuleEvaluationResult objects
            output_path: Path where Excel file should be saved
            responsible_party_column: Column name for audit leaders
            review_year: Year for the report header
            **kwargs: Additional parameters

        Returns:
            str: Path to the generated Excel file
        """

        logger.info(f"Generating QA report from scratch for {len(rule_results)} rules")

        # Extract dynamic structure
        rule_names, audit_leaders, leader_rule_matrix = self._extract_dynamic_structure(
            rule_results, responsible_party_column
        )

        logger.info(f"Found {len(rule_names)} rules and {len(audit_leaders)} audit leaders")

        # Create new workbook from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QA Results Summary"

        # PHASE 1: Build complete structure (no data, no formulas)
        self._create_header(ws, review_year)
        section2_start = self._create_section1_structure(ws, rule_names)
        section3_start = self._create_section2_structure(ws, rule_names, audit_leaders, section2_start)
        section3_data_start = self._create_section3_structure(ws, rule_names, audit_leaders, section3_start)

        # PHASE 2: Populate Section 3 data first (this is the source of truth)
        self._populate_section3_data(ws, rule_names, audit_leaders, leader_rule_matrix, section3_data_start)

        # PHASE 3: Generate Section 1 formulas that reference Section 3
        self._generate_section1_formulas(ws, rule_names, section3_data_start, len(audit_leaders))

        # PHASE 4: Generate Section 2 formulas that reference Section 3
        self._generate_section2_formulas(ws, rule_names, audit_leaders, section2_start, section3_data_start)

        # PHASE 5: Final cleanup
        self._cleanup_worksheet(ws)
        wb.save(output_path)
        wb.close()

        logger.info(f"Generated QA report: {output_path}")
        return output_path

    def _extract_dynamic_structure(self, rule_results: Dict[str, Any],
                                  responsible_party_column: str) -> Tuple[List[str], List[str], Dict]:
        """Extract rules, audit leaders, and performance matrix from validation results."""

        rule_names = list(rule_results.keys())
        audit_leaders = set()
        leader_rule_matrix = {}

        for rule_id, rule_result in rule_results.items():
            # Try to get data from party_results first
            if hasattr(rule_result, 'party_results') and rule_result.party_results:
                for leader, party_data in rule_result.party_results.items():
                    audit_leaders.add(leader)
                    if leader not in leader_rule_matrix:
                        leader_rule_matrix[leader] = {}

                    # Extract metrics from party_data
                    if isinstance(party_data, dict) and 'metrics' in party_data:
                        metrics = party_data['metrics']
                        leader_rule_matrix[leader][rule_id] = {
                            'gc_count': metrics.get('gc_count', 0),
                            'pc_count': metrics.get('pc_count', 0),
                            'dnc_count': metrics.get('dnc_count', 0),
                            'na_count': metrics.get('na_count', 0),
                            'total': metrics.get('total_count', 0),
                            'status': party_data.get('status', 'N/A'),
                            'error_rate': metrics.get('dnc_rate', 0),
                            'threshold': getattr(rule_result.rule, 'threshold', 0.02) if hasattr(rule_result, 'rule') else 0.02
                        }
                    else:
                        # Fallback calculation
                        leader_rule_matrix[leader][rule_id] = self._calculate_leader_metrics(
                            None, rule_result
                        )

            # Fallback to result_df if party_results not available
            elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
                df = rule_result.result_df
                if responsible_party_column in df.columns:
                    unique_leaders = df[responsible_party_column].dropna().unique()
                    for leader in unique_leaders:
                        audit_leaders.add(str(leader))
                        if str(leader) not in leader_rule_matrix:
                            leader_rule_matrix[str(leader)] = {}

                        leader_df = df[df[responsible_party_column] == leader]
                        leader_rule_matrix[str(leader)][rule_id] = self._calculate_leader_metrics(
                            leader_df, rule_result
                        )

        return sorted(rule_names), sorted(list(audit_leaders)), leader_rule_matrix

    def _calculate_leader_metrics(self, df: pd.DataFrame, rule_result: Any = None) -> Dict[str, Any]:
        """Calculate performance metrics for a leader's data in a specific rule."""

        # Try compliance_metrics first
        if rule_result and hasattr(rule_result, 'compliance_metrics'):
            metrics = rule_result.compliance_metrics
            gc_count = metrics.get('compliant_count', 0)
            pc_count = metrics.get('partially_compliant_count', 0)
            dnc_count = metrics.get('non_compliant_count', 0)
            na_count = metrics.get('not_applicable_count', 0)
        elif df is not None and not df.empty:
            # Analyze DataFrame
            total = len(df)
            result_col = self._find_result_column(df, rule_result)

            if result_col and result_col in df.columns:
                gc_count = df[result_col].isin(['GC', 'PASS', 'Pass', 'TRUE', True, 1]).sum()
                pc_count = df[result_col].isin(['PC', 'PARTIAL', 'Partial']).sum()
                dnc_count = df[result_col].isin(['DNC', 'FAIL', 'Fail', 'FALSE', False, 0]).sum()
                na_count = df[result_col].isin(['N/A', 'NA', None, '']).sum() + df[result_col].isna().sum()
            else:
                gc_count = pc_count = dnc_count = na_count = 0
        else:
            gc_count = pc_count = dnc_count = na_count = 0

        # Calculate derived metrics
        total = gc_count + pc_count + dnc_count + na_count
        applicable_total = gc_count + pc_count + dnc_count

        if applicable_total > 0:
            error_rate = (pc_count + dnc_count) / applicable_total
        else:
            error_rate = 0

        threshold = 0.02
        if rule_result and hasattr(rule_result, 'threshold'):
            threshold = rule_result.threshold

        # Determine status
        if applicable_total == 0:
            status = 'N/A'
        elif error_rate > threshold:
            status = 'DNC'
        else:
            status = 'GC'

        return {
            'total': total,
            'gc_count': gc_count,
            'pc_count': pc_count,
            'dnc_count': dnc_count,
            'na_count': na_count,
            'applicable_total': applicable_total,
            'error_rate': error_rate,
            'threshold': threshold,
            'status': status
        }

    def _find_result_column(self, df: pd.DataFrame, rule_result: Any = None) -> Optional[str]:
        """Find the result column in a DataFrame."""
        if rule_result and hasattr(rule_result, 'result_column'):
            return rule_result.result_column

        for col in ['Result', 'Status', 'Compliance', 'Overall Test Result']:
            if col in df.columns:
                return col
        return None

    def _create_header(self, ws, review_year: Optional[str]):
        """Create the report header section."""
        if not review_year:
            review_year = datetime.datetime.now().strftime('%Y')

        # Main title
        title = f"QA {review_year} Summary Ratings and Results Report"
        ws.merge_cells('B2:H2')
        ws['B2'] = title
        ws['B2'].font = self.title_font
        ws['B2'].fill = self.blue_header_fill
        ws['B2'].alignment = self.center_align

        # Merge title across multiple columns
        ws.merge_cells('B2:H2')

        # Set row height
        ws.row_dimensions[2].height = 25

    def _create_section1_structure(self, ws, rule_names: List[str]) -> int:
        """Create Section 1 structure (labels and headers only, no formulas yet)."""

        start_row = self.section1_start_row

        # Section title
        ws[f'B{start_row}'] = "IAG Overall Results and Rating"
        ws[f'B{start_row}'].font = self.section_header_font
        ws[f'B{start_row}'].fill = self.blue_header_fill
        ws[f'B{start_row}'].alignment = self.center_align
        ws.merge_cells(f'B{start_row}:C{start_row}')

        # Labels column
        labels = [
            "GC Score",
            "PC Score",
            "DNC Score",
            "Total Count of Applications",
            "Weighted Score Across All Tests",
            "Weighted Rating Across All Tests"
        ]

        for i, label in enumerate(labels):
            row = start_row + 1 + i
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = self.label_font
            ws[f'B{row}'].alignment = self.left_align
            ws[f'B{row}'].border = self.thin_border

        # Test column headers (dynamic)
        test_start_col = 4  # Column D
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(test_start_col + i)

            # Header only (formulas added later)
            ws[f'{col_letter}{start_row + 1}'] = rule_name
            ws[f'{col_letter}{start_row + 1}'].font = self.column_header_font
            ws[f'{col_letter}{start_row + 1}'].alignment = self.center_align
            ws[f'{col_letter}{start_row + 1}'].border = self.thin_border
            ws.column_dimensions[col_letter].width = self.test_column_width

            # Create empty cells with borders for formulas later
            for j in range(6):  # 6 data rows
                cell_row = start_row + 2 + j
                ws[f'{col_letter}{cell_row}'].border = self.thin_border
                ws[f'{col_letter}{cell_row}'].alignment = self.center_align

        # Summary column headers
        summary_start_col = test_start_col + len(rule_names)
        summary_headers = [
            "Weighted Score across Audit Leaders",
            "Weighted Average Rating: 4 Point Scale",
            "Volume of Sampled Audit Entities for IAG",
            "Overridden IAG Rating, Where Applicable",
            "Rating Override Rationale, Where Applicable"
        ]

        for i, header in enumerate(summary_headers):
            col_letter = get_column_letter(summary_start_col + i)
            ws[f'{col_letter}{start_row + 1}'] = header
            ws[f'{col_letter}{start_row + 1}'].font = self.column_header_font
            ws[f'{col_letter}{start_row + 1}'].alignment = self.center_align
            ws[f'{col_letter}{start_row + 1}'].border = self.thin_border
            ws.column_dimensions[col_letter].width = self.summary_column_width

            # Create empty cells with borders
            for j in range(6):
                cell_row = start_row + 2 + j
                ws[f'{col_letter}{cell_row}'].border = self.thin_border
                ws[f'{col_letter}{cell_row}'].alignment = self.center_align

        return start_row + len(labels) + 1 + self.section_spacing

    def _create_section2_structure(self, ws, rule_names: List[str], audit_leaders: List[str], start_row: int) -> int:
        """Create Section 2 structure (no formulas yet)."""

        # Section title
        ws[f'B{start_row}'] = "Audit Leader Overall Results and Ratings"
        ws[f'B{start_row}'].font = self.section_header_font
        ws[f'B{start_row}'].fill = self.blue_header_fill
        ws[f'B{start_row}'].alignment = self.center_align
        ws.merge_cells(f'B{start_row}:C{start_row}')

        # Headers
        header_row = start_row + 1
        ws[f'B{header_row}'] = "Audit Leader"
        ws[f'B{header_row}'].font = self.column_header_font
        ws[f'B{header_row}'].border = self.thin_border
        ws[f'B{header_row}'].alignment = self.center_align
        ws.column_dimensions['B'].width = self.leader_column_width

        ws[f'C{header_row}'] = "Measurement Description"
        ws[f'C{header_row}'].font = self.column_header_font
        ws[f'C{header_row}'].border = self.thin_border
        ws[f'C{header_row}'].alignment = self.center_align
        ws.column_dimensions['C'].width = self.label_column_width

        # Test column headers (dynamic)
        test_start_col = 4
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(test_start_col + i)
            ws[f'{col_letter}{header_row}'] = rule_name
            ws[f'{col_letter}{header_row}'].font = self.column_header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.thin_border
            ws.column_dimensions[col_letter].width = self.test_column_width

        # Summary headers (dynamic)
        summary_start_col = test_start_col + len(rule_names)
        summary_headers = [
            "Weighted Score",
            "Weighted Average Rating: 4 Point Scale",
            "Volume of Sampled Audit Entities by AL",
            "Overridden AL Rating, Where Applicable",
            "Rating Override Rationale, Where Applicable"
        ]

        for i, header in enumerate(summary_headers):
            col_letter = get_column_letter(summary_start_col + i)
            ws[f'{col_letter}{header_row}'] = header
            ws[f'{col_letter}{header_row}'].font = self.column_header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.thin_border
            ws.column_dimensions[col_letter].width = self.summary_column_width

        # Leader name rows (structure only, no formulas)
        data_start_row = header_row + 1
        for i, leader in enumerate(audit_leaders):
            row = data_start_row + i

            ws[f'B{row}'] = leader
            ws[f'B{row}'].font = self.data_font
            ws[f'B{row}'].border = self.thin_border

            ws[f'C{row}'] = "Total Weighted Score"
            ws[f'C{row}'].font = self.data_font
            ws[f'C{row}'].border = self.thin_border

            # Create empty cells with borders for all test and summary columns
            for j in range(len(rule_names) + len(summary_headers)):
                col_letter = get_column_letter(test_start_col + j)
                ws[f'{col_letter}{row}'].border = self.thin_border
                ws[f'{col_letter}{row}'].alignment = self.center_align

        return data_start_row + len(audit_leaders) + self.section_spacing

    def _create_section3_structure(self, ws, rule_names: List[str], audit_leaders: List[str], start_row: int) -> int:
        """Create Section 3 structure and return data start row."""

        # Section title
        ws[f'A{start_row}'] = "Audit Leader Average Test Results"
        ws[f'A{start_row}'].font = self.section_header_font
        ws[f'A{start_row}'].fill = self.blue_header_fill
        ws[f'A{start_row}'].alignment = self.center_align

        # Full-width blue analytics section
        analytics_start_row = start_row + 2
        self._create_analytics_section(ws, rule_names, analytics_start_row)

        # Column headers
        header_row = analytics_start_row + self.section3_analytics_rows + 1

        ws[f'B{header_row}'] = "Audit Leader"
        ws[f'B{header_row}'].font = self.column_header_font
        ws[f'B{header_row}'].border = self.thin_border

        ws[f'C{header_row}'] = "Samples Tested for Audit Leader"
        ws[f'C{header_row}'].font = self.column_header_font
        ws[f'C{header_row}'].border = self.thin_border

        # Test column headers (dynamic)
        test_start_col = 4
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(test_start_col + i)
            ws[f'{col_letter}{header_row}'] = rule_name
            ws[f'{col_letter}{header_row}'].font = self.column_header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.thin_border

        # Aggregate column headers (dynamic)
        agg_start_col = test_start_col + len(rule_names)
        agg_headers = ["GC Count", "PC Count", "DNC Count", "Total Applicable Count", "Average Score", "Average Rating: 4 Point Scale"]

        for i, header in enumerate(agg_headers):
            col_letter = get_column_letter(agg_start_col + i)
            ws[f'{col_letter}{header_row}'] = header
            ws[f'{col_letter}{header_row}'].font = self.column_header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.thin_border
            ws.column_dimensions[col_letter].width = 15

        # Leader name rows (structure only, data added later)
        data_start_row = header_row + 1
        for i, leader in enumerate(audit_leaders):
            row = data_start_row + i

            ws[f'B{row}'] = leader
            ws[f'B{row}'].border = self.thin_border

            ws[f'C{row}'] = ""  # Manual entry field
            ws[f'C{row}'].border = self.thin_border

            # Create empty cells with borders for all test and aggregate columns
            for j in range(len(rule_names) + len(agg_headers)):
                col_letter = get_column_letter(test_start_col + j)
                ws[f'{col_letter}{row}'].border = self.thin_border
                ws[f'{col_letter}{row}'].alignment = self.center_align

        return data_start_row

    def _create_analytics_section(self, ws, rule_names: List[str], start_row: int):
        """Create the blue analytics section with FULL width coverage."""

        # Calculate total columns needed: A + B + C + test columns + 6 aggregate columns
        total_columns = 3 + len(rule_names) + 6

        analytics_labels = [
            "Analytics",
            "",  # Blank row
            "Error Threshold",
            "Risk Level",
            "Not Applicable",
            "Not Applicable",
            "Not Applicable",
            "Rule ID"
        ]

        # Apply analytics section across ALL columns
        for row_offset in range(self.section3_analytics_rows):
            row = start_row + row_offset

            # Cover ALL columns from A to the last aggregate column
            for col_index in range(1, total_columns + 1):
                col_letter = get_column_letter(col_index)

                # Set blue background for all cells
                ws[f'{col_letter}{row}'].fill = self.blue_analytics_fill
                ws[f'{col_letter}{row}'].font = Font(color="FFFFFF", bold=(row_offset == 0))
                ws[f'{col_letter}{row}'].alignment = self.center_align

                # Set content based on column and row
                if col_index == 1:  # Column A
                    if row_offset < len(analytics_labels):
                        ws[f'{col_letter}{row}'] = analytics_labels[row_offset]
                elif col_index <= 3:  # Columns B, C
                    if row_offset == 0:
                        ws[f'{col_letter}{row}'] = "Analytics"
                elif col_index <= 3 + len(rule_names):  # Test columns
                    rule_index = col_index - 4
                    if row_offset == 0:
                        ws[f'{col_letter}{row}'] = "Analytics"
                    elif row_offset == 1:
                        ws[f'{col_letter}{row}'] = ""
                    elif row_offset == 2:
                        ws[f'{col_letter}{row}'] = "2%"  # Should come from rule threshold
                    elif row_offset == 3:
                        ws[f'{col_letter}{row}'] = "3"   # Should come from rule risk level
                    elif row_offset in [4, 5, 6]:
                        ws[f'{col_letter}{row}'] = "Not Applicable"
                    elif row_offset == 7:
                        ws[f'{col_letter}{row}'] = rule_names[rule_index] if rule_index < len(rule_names) else ""
                else:  # Aggregate columns
                    if row_offset == 0:
                        ws[f'{col_letter}{row}'] = "Analytics"

    def _populate_section3_data(self, ws, rule_names: List[str], audit_leaders: List[str],
                               leader_rule_matrix: Dict[str, Dict], data_start_row: int):
        """Populate Section 3 with compliance status and aggregate calculations (SOURCE OF TRUTH)."""

        test_start_col = 4
        agg_start_col = test_start_col + len(rule_names)

        for i, leader in enumerate(audit_leaders):
            row = data_start_row + i

            # Track totals for aggregate columns
            total_gc = total_pc = total_dnc = 0

            # Test result columns (dynamic number of test columns)
            for j, rule_name in enumerate(rule_names):
                col_letter = get_column_letter(test_start_col + j)

                if leader in leader_rule_matrix and rule_name in leader_rule_matrix[leader]:
                    rule_data = leader_rule_matrix[leader][rule_name]
                    status = rule_data.get('status', 'N/A')

                    # Apply status-based coloring
                    if status == 'GC':
                        fill = self.green_status_fill
                    elif status == 'PC':
                        fill = self.yellow_status_fill
                    elif status == 'DNC':
                        fill = self.red_status_fill
                    else:
                        fill = None

                    ws[f'{col_letter}{row}'] = status
                    if fill:
                        ws[f'{col_letter}{row}'].fill = fill

                    # Add to totals
                    total_gc += rule_data.get('gc_count', 0)
                    total_pc += rule_data.get('pc_count', 0)
                    total_dnc += rule_data.get('dnc_count', 0)
                else:
                    ws[f'{col_letter}{row}'] = 'N/A'

                ws[f'{col_letter}{row}'].alignment = self.center_align
                ws[f'{col_letter}{row}'].border = self.thin_border

            # Aggregate columns (dynamic positioning)
            gc_col = get_column_letter(agg_start_col)
            pc_col = get_column_letter(agg_start_col + 1)
            dnc_col = get_column_letter(agg_start_col + 2)
            total_col = get_column_letter(agg_start_col + 3)
            avg_col = get_column_letter(agg_start_col + 4)
            rating_col = get_column_letter(agg_start_col + 5)

            # Populate aggregate data
            ws[f'{gc_col}{row}'] = total_gc
            ws[f'{pc_col}{row}'] = total_pc
            ws[f'{dnc_col}{row}'] = total_dnc
            ws[f'{total_col}{row}'] = total_gc + total_pc + total_dnc

            # Calculate average score and rating
            applicable_total = total_gc + total_pc + total_dnc
            if applicable_total > 0:
                avg_score = (total_gc * 5 + total_pc * 3 + total_dnc * 1) / applicable_total
                ws[f'{avg_col}{row}'] = avg_score
                ws[f'{avg_col}{row}'].number_format = self.decimal_format

                if avg_score >= 4:
                    rating = "GC"
                    rating_fill = self.green_status_fill
                elif avg_score >= 2.5:
                    rating = "PC"
                    rating_fill = self.yellow_status_fill
                else:
                    rating = "DNC"
                    rating_fill = self.red_status_fill

                ws[f'{rating_col}{row}'] = rating
                ws[f'{rating_col}{row}'].fill = rating_fill
            else:
                ws[f'{avg_col}{row}'] = 0
                ws[f'{rating_col}{row}'] = "N/A"

            # Apply formatting to all aggregate cells
            for col in [gc_col, pc_col, dnc_col, total_col, avg_col, rating_col]:
                ws[f'{col}{row}'].alignment = self.center_align
                ws[f'{col}{row}'].border = self.thin_border

    def _generate_section1_formulas(self, ws, rule_names: List[str], section3_data_start: int, num_leaders: int):
        """Generate Section 1 formulas that reference Section 3 data."""

        start_row = self.section1_start_row
        test_start_col = 4
        section3_data_end = section3_data_start + num_leaders - 1

        # For each test column, add formulas that reference Section 3
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(test_start_col + i)

            # GC Score (weighted) - references Section 3 column for this test
            ws[f'{col_letter}{start_row + 2}'] = f'=COUNTIF({col_letter}{section3_data_start}:{col_letter}{section3_data_end},"GC")*5'
            ws[f'{col_letter}{start_row + 2}'].number_format = self.number_format

            # PC Score (weighted)
            ws[f'{col_letter}{start_row + 3}'] = f'=COUNTIF({col_letter}{section3_data_start}:{col_letter}{section3_data_end},"PC")*3'
            ws[f'{col_letter}{start_row + 3}'].number_format = self.number_format

            # DNC Score (weighted)
            ws[f'{col_letter}{start_row + 4}'] = f'=COUNTIF({col_letter}{section3_data_start}:{col_letter}{section3_data_end},"DNC")*1'
            ws[f'{col_letter}{start_row + 4}'].number_format = self.number_format

            # Total applicable count
            ws[f'{col_letter}{start_row + 5}'] = f'=COUNTIF({col_letter}{section3_data_start}:{col_letter}{section3_data_end},"<>N/A")'
            ws[f'{col_letter}{start_row + 5}'].number_format = self.number_format

            # Weighted score
            ws[f'{col_letter}{start_row + 6}'] = f'=IF({col_letter}{start_row + 5}=0,"N/A",({col_letter}{start_row + 2}+{col_letter}{start_row + 3}+{col_letter}{start_row + 4})/({col_letter}{start_row + 5}*5))'
            ws[f'{col_letter}{start_row + 6}'].number_format = self.decimal_format

            # Rating
            ws[f'{col_letter}{start_row + 7}'] = f'=IF({col_letter}{start_row + 6}="N/A","N/A",IF({col_letter}{start_row + 6}>=4,"GC",IF({col_letter}{start_row + 6}>=2.5,"PC","DNC")))'

        # Summary formulas (dynamically reference all test columns)
        summary_start_col = test_start_col + len(rule_names)
        weighted_score_col = get_column_letter(summary_start_col)
        rating_col = get_column_letter(summary_start_col + 1)

        # Dynamic range for all test columns
        first_test_col = get_column_letter(test_start_col)
        last_test_col = get_column_letter(test_start_col + len(rule_names) - 1)
        test_range = f"{first_test_col}{start_row + 6}:{last_test_col}{start_row + 6}"

        ws[f'{weighted_score_col}{start_row + 6}'] = f'=IF(COUNT({test_range})=0,"N/A",SUM({test_range})/COUNT({test_range}))'
        ws[f'{weighted_score_col}{start_row + 6}'].number_format = self.decimal_format

        ws[f'{rating_col}{start_row + 6}'] = f'=IF({weighted_score_col}{start_row + 6}="N/A","N/A",IF({weighted_score_col}{start_row + 6}>=4,"GC",IF({weighted_score_col}{start_row + 6}>=2.5,"PC","DNC")))'

    def _generate_section2_formulas(self, ws, rule_names: List[str], audit_leaders: List[str],
                                   section2_start: int, section3_data_start: int):
        """Generate Section 2 formulas that reference Section 3 data."""

        data_start_row = section2_start + 2  # Section title + header
        test_start_col = 4
        summary_start_col = test_start_col + len(rule_names)
        section3_data_end = section3_data_start + len(audit_leaders) - 1

        for i, leader in enumerate(audit_leaders):
            row = data_start_row + i
            section3_row = section3_data_start + i  # Corresponding row in Section 3

            # Test scores - reference Section 3 aggregate columns for consistency
            agg_start_col_s3 = test_start_col + len(rule_names)  # Start of aggregate columns in Section 3
            gc_col_s3 = get_column_letter(agg_start_col_s3)      # GC Count column
            pc_col_s3 = get_column_letter(agg_start_col_s3 + 1)  # PC Count column
            dnc_col_s3 = get_column_letter(agg_start_col_s3 + 2) # DNC Count column

            for j, rule_name in enumerate(rule_names):
                col_letter = get_column_letter(test_start_col + j)

                # Formula to calculate test-specific weighted score for this leader
                # This references the Section 3 status for this leader/test combination
                section3_test_col = get_column_letter(test_start_col + j)
                ws[f'{col_letter}{row}'] = f'=IF({section3_test_col}{section3_row}="GC",5,IF({section3_test_col}{section3_row}="PC",3,IF({section3_test_col}{section3_row}="DNC",1,0)))'
                ws[f'{col_letter}{row}'].number_format = self.number_format

            # Summary calculations
            weighted_score_col = get_column_letter(summary_start_col)
            rating_col = get_column_letter(summary_start_col + 1)

            # Total weighted score - sum all test columns for this leader
            first_test_col = get_column_letter(test_start_col)
            last_test_col = get_column_letter(test_start_col + len(rule_names) - 1)
            test_range = f"{first_test_col}{row}:{last_test_col}{row}"

            ws[f'{weighted_score_col}{row}'] = f'=SUM({test_range})'
            ws[f'{weighted_score_col}{row}'].number_format = self.decimal_format

            # Average rating - reference Section 3 aggregate data for this leader
            total_applicable_col_s3 = get_column_letter(agg_start_col_s3 + 3)  # Total Applicable Count
            ws[f'{rating_col}{row}'] = f'=IF({total_applicable_col_s3}{section3_row}=0,"N/A",{weighted_score_col}{row}/({total_applicable_col_s3}{section3_row}*5))'
            ws[f'{rating_col}{row}'].number_format = self.decimal_format

    def _cleanup_worksheet(self, ws):
        """Final cleanup and formatting of the worksheet."""
        # Set column widths properly
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = self.leader_column_width
        ws.column_dimensions['C'].width = self.label_column_width

        # Set row heights
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 18

        # Replace any None values with empty strings
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    cell.value = ""


def main():
    """Example usage of the DynamicQAReportGenerator."""

    # Sample data structure - replace with actual data
    sample_rule_results = {
        "QA-ID-1": type('MockResult', (), {
            'compliance_metrics': {
                'compliant_count': 15,
                'partially_compliant_count': 3,
                'non_compliant_count': 2,
                'not_applicable_count': 0
            },
            'party_results': {
                'Leader A': {
                    'metrics': {
                        'gc_count': 8,
                        'pc_count': 1,
                        'dnc_count': 1,
                        'na_count': 0,
                        'total_count': 10
                    },
                    'status': 'GC'
                },
                'Leader B': {
                    'metrics': {
                        'gc_count': 7,
                        'pc_count': 2,
                        'dnc_count': 1,
                        'na_count': 0,
                        'total_count': 10
                    },
                    'status': 'GC'
                }
            },
            'threshold': 0.02
        })(),
        "QA-ID-2": type('MockResult', (), {
            'compliance_metrics': {
                'compliant_count': 12,
                'partially_compliant_count': 4,
                'non_compliant_count': 4,
                'not_applicable_count': 0
            },
            'party_results': {
                'Leader A': {
                    'metrics': {
                        'gc_count': 6,
                        'pc_count': 2,
                        'dnc_count': 2,
                        'na_count': 0,
                        'total_count': 10
                    },
                    'status': 'PC'
                },
                'Leader B': {
                    'metrics': {
                        'gc_count': 6,
                        'pc_count': 2,
                        'dnc_count': 2,
                        'na_count': 0,
                        'total_count': 10
                    },
                    'status': 'PC'
                }
            },
            'threshold': 0.02
        })()
    }

    # Generate report
    generator = DynamicQAReportGenerator()
    output_path = generator.generate_report(
        rule_results=sample_rule_results,
        output_path="qa_summary_report.xlsx",
        review_year="2025"
    )

    print(f"Generated report: {output_path}")


if __name__ == "__main__":
    main()