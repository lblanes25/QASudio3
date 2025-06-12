"""
Dynamic Summary Template Processor for IAG and AL Results and Ratings

Reorganized into logical sections:
- Common utilities and data processing
- Section 1: IAG Overall Results and Ratings
- Section 2: Audit Leader Overall Results and Ratings
- Section 3: Audit Leader Average Test Results

Each section handles both data calculation and Excel formatting separately.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.cell import MergedCell
import pandas as pd
from pathlib import Path
import logging
import datetime
from typing import Dict, Any, Optional, List, Tuple
import shutil

logger = logging.getLogger(__name__)


class DynamicSummaryTemplateProcessor:
    """
    Processes the IAG and AL Results and Ratings summary template dynamically.
    Handles variable numbers of rules and audit leaders.
    """

    def __init__(self, template_path: str):
        """Initialize the template processor."""
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")

        # Define common styles
        self._setup_styles()

    def _setup_styles(self):
        """Define common Excel styles used throughout the template."""
        self.header_font = Font(bold=True, size=11)
        self.data_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.number_format = '#,##0'
        self.percentage_format = '0.0%'

        # Define color fills for compliance statuses
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ==========================================
    # COMMON UTILITIES AND DATA PROCESSING
    # ==========================================

    def _safe_write_cell(self, ws, cell_ref: str, value: Any, **kwargs):
        """Safely write to a cell, handling merged cells properly."""
        cell = ws[cell_ref]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    ws[top_left.coordinate].value = value
                    for attr, attr_value in kwargs.items():
                        setattr(ws[top_left.coordinate], attr, attr_value)
                    return

        cell.value = value
        for attr, attr_value in kwargs.items():
            setattr(cell, attr, attr_value)

    def _safe_read_cell(self, ws, cell_ref: str):
        """Safely read from a cell, handling merged cells properly."""
        cell = ws[cell_ref]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = merged_range.start_cell
                    return ws[top_left.coordinate].value

        return cell.value

    def _extract_dynamic_structure(self, rule_results: Dict[str, Any],
                                  results: Dict[str, Any],
                                  responsible_party_column: str) -> Tuple[List[str], List[str], Dict]:
        """Extract the dynamic structure from validation results."""

        logger.info(f"Extracting dynamic structure with responsible_party_column: '{responsible_party_column}'")

        rule_names = list(rule_results.keys())
        audit_leaders = set()
        leader_rule_matrix = {}

        for rule_id, rule_result in rule_results.items():
            logger.debug(f"Processing rule: {rule_id}")

            if hasattr(rule_result, 'party_results') and rule_result.party_results:
                logger.info(f"Rule {rule_id}: party_results found with {len(rule_result.party_results)} leaders")
                for leader, party_data in rule_result.party_results.items():
                    audit_leaders.add(leader)
                    if leader not in leader_rule_matrix:
                        leader_rule_matrix[leader] = {}

                    if isinstance(party_data, dict):
                        if 'metrics' in party_data and isinstance(party_data['metrics'], dict):
                            metrics = party_data['metrics']
                            leader_rule_matrix[leader][rule_id] = {
                                'gc_count': metrics.get('gc_count', 0),
                                'pc_count': metrics.get('pc_count', 0),
                                'dnc_count': metrics.get('dnc_count', 0),
                                'na_count': metrics.get('na_count', 0),
                                'total': metrics.get('total_count', 0),
                                'status': party_data.get('status', 'N/A'),
                                'error_rate': metrics.get('dnc_rate', 0),
                                'threshold': getattr(rule_result.rule, 'threshold', 0.02)
                            }
                        elif 'gc_count' in party_data:
                            leader_rule_matrix[leader][rule_id] = party_data
                        else:
                            leader_rule_matrix[leader][rule_id] = self._calculate_leader_rule_result(None, rule_result)

            elif hasattr(rule_result, 'result_df') and rule_result.result_df is not None:
                df = rule_result.result_df
                if responsible_party_column in df.columns:
                    unique_leaders = df[responsible_party_column].dropna().unique()
                    for leader in unique_leaders:
                        audit_leaders.add(str(leader))
                        if str(leader) not in leader_rule_matrix:
                            leader_rule_matrix[str(leader)] = {}

                        leader_df = df[df[responsible_party_column] == leader]
                        leader_rule_matrix[str(leader)][rule_id] = self._calculate_leader_rule_result(leader_df, rule_result)

        return sorted(rule_names), sorted(list(audit_leaders)), leader_rule_matrix

    def _calculate_leader_rule_result(self, df: pd.DataFrame, rule_result: Any = None) -> Dict[str, Any]:
        """Calculate results for a leader's data in a rule."""

        # First try to use compliance_metrics if available
        if rule_result and hasattr(rule_result, 'compliance_metrics'):
            metrics = rule_result.compliance_metrics
            gc_count = metrics.get('compliant_count', 0)
            pc_count = metrics.get('partially_compliant_count', 0)
            dnc_count = metrics.get('non_compliant_count', 0)
            na_count = metrics.get('not_applicable_count', 0)
            total = gc_count + pc_count + dnc_count + na_count

            applicable_total = gc_count + pc_count + dnc_count
            if applicable_total > 0:
                error_rate = (pc_count + dnc_count) / applicable_total
            else:
                error_rate = 0

            threshold = getattr(rule_result, 'threshold', 0.02) if rule_result else 0.02

            if applicable_total == 0:
                status = 'N/A'
            elif error_rate > threshold:
                status = 'DNC'
            else:
                status = 'GC'

            return {
                'total': total,
                'gc_count': gc_count,
                'pc_count': pc_count,
                'dnc_count': dnc_count,
                'na_count': na_count,
                'applicable_total': applicable_total,
                'error_rate': error_rate,
                'threshold': threshold,
                'status': status
            }

        # Fallback to DataFrame analysis
        if df is None or df.empty:
            return {
                'total': 0,
                'gc_count': 0,
                'pc_count': 0,
                'dnc_count': 0,
                'na_count': 0,
                'applicable_total': 0,
                'error_rate': 0,
                'threshold': 0.02,
                'status': 'N/A'
            }

        total = len(df)
        result_col = None

        if rule_result and hasattr(rule_result, 'result_column'):
            result_col = rule_result.result_column
        else:
            for col in ['Result', 'Status', 'Compliance', 'Overall Test Result']:
                if col in df.columns:
                    result_col = col
                    break

        if result_col and result_col in df.columns:
            gc_count = df[result_col].isin(['GC', 'PASS', 'Pass', 'TRUE', True, 1]).sum()
            pc_count = df[result_col].isin(['PC', 'PARTIAL', 'Partial']).sum()
            dnc_count = df[result_col].isin(['DNC', 'FAIL', 'Fail', 'FALSE', False, 0]).sum()
            na_count = df[result_col].isin(['N/A', 'NA', None, '']).sum() + df[result_col].isna().sum()
        else:
            gc_count = pc_count = dnc_count = na_count = 0

        applicable_total = gc_count + pc_count + dnc_count
        if applicable_total > 0:
            error_rate = (pc_count + dnc_count) / applicable_total
        else:
            error_rate = 0

        threshold = 0.02
        if rule_result and hasattr(rule_result, 'threshold'):
            threshold = rule_result.threshold

        if applicable_total == 0:
            status = 'N/A'
        elif error_rate > threshold:
            status = 'DNC'
        else:
            status = 'GC'

        return {
            'total': total,
            'gc_count': gc_count,
            'pc_count': pc_count,
            'dnc_count': dnc_count,
            'na_count': na_count,
            'applicable_total': applicable_total,
            'error_rate': error_rate,
            'threshold': threshold,
            'status': status
        }

    def _cleanup_none_values(self, ws):
        """Replace any None values with empty strings to avoid Excel errors."""
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                if cell.value is None:
                    cell.value = ""
                elif cell.value == "None":
                    cell.value = ""

    # ==========================================
    # MAIN GENERATION METHOD
    # ==========================================

    def generate_summary_report(self,
                               results: Dict[str, Any],
                               rule_results: Dict[str, Any],
                               output_path: str,
                               responsible_party_column: str = "Audit Leader",
                               review_year: Optional[str] = None,
                               **kwargs) -> str:
        """Generate a dynamic summary report using the template."""

        logger.info(f"Generating dynamic summary report for {len(rule_results)} rules")

        # Copy template and load workbook
        shutil.copy2(self.template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Extract dynamic structure from validation data
        rule_names, audit_leaders, leader_rule_matrix = self._extract_dynamic_structure(
            rule_results, results, responsible_party_column
        )

        # Update header with review year
        self._update_header(ws, review_year)

        # Process each section
        self._process_section1_iag_overall(ws, rule_results, rule_names, leader_rule_matrix)
        self._process_section2_audit_leader_overall(ws, rule_names, audit_leaders, leader_rule_matrix)
        self._process_section3_detailed_test_results(ws, rule_names, audit_leaders, leader_rule_matrix)

        # Final cleanup and save
        self._cleanup_none_values(ws)
        wb.save(output_path)
        wb.close()

        logger.info(f"Generated dynamic summary report: {output_path}")
        return output_path

    def _update_header(self, ws, review_year: Optional[str]):
        """Update the header section with review year."""
        if not review_year:
            review_year = f"{datetime.datetime.now().strftime('%Y')} QA Review"

        b2_cell = ws['B2']
        if isinstance(b2_cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if b2_cell.coordinate in merged_range:
                    b2_cell = ws[merged_range.start_cell.coordinate]
                    break

        if b2_cell.value:
            cell_value = str(b2_cell.value)
            cell_value = cell_value.replace('{{Review_Year}}', review_year)
            cell_value = cell_value.replace('{{Review_Name}}', '')
            self._safe_write_cell(ws, 'B2', cell_value.strip())

    # ==========================================
    # SECTION 1: IAG OVERALL RESULTS AND RATINGS
    # ==========================================

    def _process_section1_iag_overall(self, ws, rule_results: Dict[str, Any],
                                     rule_names: List[str], leader_rule_matrix: Dict[str, Dict]):
        """Process Section 1: IAG Overall Results and Ratings."""
        logger.info("Processing Section 1: IAG Overall Results and Ratings")

        # Calculate data for Section 1
        section1_data = self._calculate_section1_data(rule_results, leader_rule_matrix)

        # Update Section 1 with calculated data
        self._update_section1_cells(ws, section1_data)

        # Add dynamic test columns to Section 1
        self._add_section1_test_columns(ws, rule_names, leader_rule_matrix)

        # Add summary columns to Section 1
        self._add_section1_summary_columns(ws, rule_names)

    def _calculate_section1_data(self, rule_results: Dict[str, Any],
                                leader_rule_matrix: Dict[str, Dict]) -> Dict[str, int]:
        """Calculate aggregate data for Section 1."""
        total_gc = total_pc = total_dnc = total_na = 0

        # First try to use rule_results directly for accurate counts
        for rule_id, rule_result in rule_results.items():
            if hasattr(rule_result, 'compliance_metrics'):
                metrics = rule_result.compliance_metrics
                total_gc += metrics.get('compliant_count', 0)
                total_pc += metrics.get('partially_compliant_count', 0)
                total_dnc += metrics.get('non_compliant_count', 0)
                total_na += metrics.get('not_applicable_count', 0)
            else:
                # Fallback to leader_rule_matrix
                for leader_data in leader_rule_matrix.values():
                    if rule_id in leader_data:
                        rule_data = leader_data[rule_id]
                        total_gc += rule_data.get('gc_count', 0)
                        total_pc += rule_data.get('pc_count', 0)
                        total_dnc += rule_data.get('dnc_count', 0)
                        total_na += rule_data.get('na_count', 0)
                        break

        return {
            'total_gc': total_gc,
            'total_pc': total_pc,
            'total_dnc': total_dnc,
            'total_na': total_na,
            'total_applicable': total_gc + total_pc + total_dnc
        }

    def _update_section1_cells(self, ws, section1_data: Dict[str, int]):
        """Update Section 1 cells with calculated values."""
        # Update cells with calculated totals (looking for placeholders)
        if self._safe_read_cell(ws, 'D6') and '{{IAG_COUNT_OF_GC}}' in str(self._safe_read_cell(ws, 'D6')):
            self._safe_write_cell(ws, 'D6', section1_data['total_gc'])

        if self._safe_read_cell(ws, 'D7') and '{{IAG_COUNT_OF_PC}}' in str(self._safe_read_cell(ws, 'D7')):
            self._safe_write_cell(ws, 'D7', section1_data['total_pc'])

        if self._safe_read_cell(ws, 'D8') and '{{IAG_COUNT_OF_DNC}}' in str(self._safe_read_cell(ws, 'D8')):
            self._safe_write_cell(ws, 'D8', section1_data['total_dnc'])

        if self._safe_read_cell(ws, 'D9') and '{{IAG_TOTAL_COUNT}}' in str(self._safe_read_cell(ws, 'D9')):
            self._safe_write_cell(ws, 'D9', section1_data['total_applicable'])

    def _add_section1_test_columns(self, ws, rule_names: List[str], leader_rule_matrix: Dict[str, Dict]):
        """Add dynamic test columns to Section 1."""
        # Find Section 3 data range for formulas
        section3_header_row = self._find_section3_header_row(ws)
        section3_start_row = section3_header_row + 1
        section3_end_row = section3_start_row + len(leader_rule_matrix.keys()) - 1

        start_col = 4  # Column D
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(start_col + i)

            # Add header for this test column
            ws[f'{col_letter}5'] = rule_name
            ws[f'{col_letter}5'].font = self.header_font
            ws[f'{col_letter}5'].alignment = self.center_align
            ws[f'{col_letter}5'].border = self.border
            ws.column_dimensions[col_letter].width = 15

            # Add formulas for each row based on Section 3 data
            # Row 6: Count of GC values * 5 (weighted)
            ws[f'{col_letter}6'] = f'=COUNTIF({col_letter}${section3_start_row}:{col_letter}${section3_end_row},"GC")*5'
            ws[f'{col_letter}6'].number_format = self.number_format
            ws[f'{col_letter}6'].border = self.border

            # Row 7: Count of PC values * 3 (weighted)
            ws[f'{col_letter}7'] = f'=COUNTIF({col_letter}${section3_start_row}:{col_letter}${section3_end_row},"PC")*3'
            ws[f'{col_letter}7'].number_format = self.number_format
            ws[f'{col_letter}7'].border = self.border

            # Row 8: Count of DNC values * 1 (weighted)
            ws[f'{col_letter}8'] = f'=COUNTIF({col_letter}${section3_start_row}:{col_letter}${section3_end_row},"DNC")*1'
            ws[f'{col_letter}8'].number_format = self.number_format
            ws[f'{col_letter}8'].border = self.border

            # Row 9: Count of applicable (non-N/A) values
            ws[f'{col_letter}9'] = f'=COUNTIF({col_letter}{section3_start_row}:{col_letter}{section3_end_row},"<>N/A")'
            ws[f'{col_letter}9'].number_format = self.number_format
            ws[f'{col_letter}9'].border = self.border

            # Row 10: Weighted average score (sum of weighted scores / total applicable * 5)
            ws[f'{col_letter}10'] = f'=IF({col_letter}9=0,"N/A",({col_letter}6+{col_letter}7+{col_letter}8)/({col_letter}9*5))'
            ws[f'{col_letter}10'].number_format = '0.00'
            ws[f'{col_letter}10'].border = self.border

            # Row 11: Rating based on score
            ws[f'{col_letter}11'] = f'=IF({col_letter}10="N/A","N/A",IF({col_letter}10>=4,"GC",IF({col_letter}10>=2.5,"PC","DNC")))'
            ws[f'{col_letter}11'].border = self.border
            ws[f'{col_letter}11'].alignment = self.center_align

    def _add_section1_summary_columns(self, ws, rule_names: List[str]):
        """Add summary columns to Section 1."""
        summary_start_col = 4 + len(rule_names)

        summary_headers = [
            "Weighted Score across Audit Leaders",
            "Weighted Average Rating: 4 Point Scale",
            "Volume of Sampled Audit Entities for IAG",
            "Overridden IAG Rating, Where Applicable",
            "Rating Override Rationale, Where Applicable"
        ]

        # Add headers
        for i, header in enumerate(summary_headers):
            col_letter = get_column_letter(summary_start_col + i)
            ws[f'{col_letter}5'] = header  # Same row as test headers
            ws[f'{col_letter}5'].font = self.header_font
            ws[f'{col_letter}5'].alignment = self.center_align
            ws[f'{col_letter}5'].border = self.border
            ws.column_dimensions[col_letter].width = 20

        # Add formulas in row 10 (Weighted Score across Audit Leaders)
        weighted_score_col = get_column_letter(summary_start_col)
        rating_col = get_column_letter(summary_start_col + 1)

        # Weighted Score: Average of all test scores
        test_range = f"D10:{get_column_letter(summary_start_col - 1)}10"
        ws[f'{weighted_score_col}10'] = f'=IFERROR(SUM({test_range})/COUNT({test_range}),"N/A")'
        ws[f'{weighted_score_col}10'].number_format = '0.00'
        ws[f'{weighted_score_col}10'].border = self.border

        # Rating: Based on weighted score
        ws[f'{rating_col}10'] = f'=IF({weighted_score_col}10="N/A", "N/A", IF({weighted_score_col}10>=4, "GC", IF({weighted_score_col}10<2.5, "DNC", "PC")))'
        ws[f'{rating_col}10'].border = self.border
        ws[f'{rating_col}10'].alignment = self.center_align

    # ==========================================
    # SECTION 2: AUDIT LEADER OVERALL RESULTS AND RATINGS
    # ==========================================

    def _process_section2_audit_leader_overall(self, ws, rule_names: List[str],
                                              audit_leaders: List[str], leader_rule_matrix: Dict[str, Dict]):
        """Process Section 2: Audit Leader Overall Results and Ratings."""
        logger.info("Processing Section 2: Audit Leader Overall Results and Ratings")

        # Setup Section 2 structure
        self._setup_section2_structure(ws, rule_names, audit_leaders)

        # Populate Section 2 with data and formulas
        self._populate_section2_data(ws, rule_names, audit_leaders, leader_rule_matrix)

        # Add summary columns to Section 2
        self._add_section2_summary_columns(ws, rule_names, audit_leaders)

    def _setup_section2_structure(self, ws, rule_names: List[str], audit_leaders: List[str]):
        """Setup the structure for Section 2."""
        leader_start_row = 15  # Typical starting row for audit leader section

        # Add rule headers
        start_col = 4  # Column D
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(start_col + i)
            ws[f'{col_letter}14'] = rule_name
            ws[f'{col_letter}14'].font = self.header_font
            ws[f'{col_letter}14'].alignment = self.center_align
            ws[f'{col_letter}14'].border = self.border
            ws.column_dimensions[col_letter].width = 15

        # Add audit leader rows
        for i, leader in enumerate(audit_leaders):
            row = leader_start_row + i
            ws[f'B{row}'] = leader
            ws[f'B{row}'].font = self.data_font
            ws[f'B{row}'].border = self.border

            ws[f'C{row}'] = "Total Weighted Score"
            ws[f'C{row}'].font = self.data_font
            ws[f'C{row}'].border = self.border

    def _populate_section2_data(self, ws, rule_names: List[str], audit_leaders: List[str],
                               leader_rule_matrix: Dict[str, Dict]):
        """Populate Section 2 with calculated data."""
        leader_start_row = 15
        start_col = 4

        for i, leader in enumerate(audit_leaders):
            row = leader_start_row + i

            for j, rule_name in enumerate(rule_names):
                col_letter = get_column_letter(start_col + j)

                # Calculate weighted score for this leader and rule
                if leader in leader_rule_matrix and rule_name in leader_rule_matrix[leader]:
                    rule_data = leader_rule_matrix[leader][rule_name]
                    
                    # Calculate weighted score WITHOUT risk level multiplication
                    # This matches how Section 3 calculates scores
                    weighted_score = ((rule_data.get('gc_count', 0) * 5 +
                                     rule_data.get('pc_count', 0) * 3 +
                                     rule_data.get('dnc_count', 0) * 1))
                else:
                    weighted_score = 0

                ws[f'{col_letter}{row}'] = weighted_score
                ws[f'{col_letter}{row}'].font = self.data_font
                ws[f'{col_letter}{row}'].border = self.border
                ws[f'{col_letter}{row}'].number_format = self.number_format

    def _add_section2_summary_columns(self, ws, rule_names: List[str], audit_leaders: List[str]):
        """Add summary columns to Section 2."""
        summary_start_col = 4 + len(rule_names)
        leader_start_row = 15

        summary_headers = [
            "Weighted Score",
            "Weighted Average Rating: 4 Point Scale",
            "Volume of Sampled Audit Entities by AL",
            "Overridden AL Rating, Where Applicable",
            "Rating Override Rationale, Where Applicable"
        ]

        # Add headers
        for i, header in enumerate(summary_headers):
            col_letter = get_column_letter(summary_start_col + i)
            ws[f'{col_letter}14'] = header
            ws[f'{col_letter}14'].font = self.header_font
            ws[f'{col_letter}14'].alignment = self.center_align
            ws[f'{col_letter}14'].border = self.border
            ws.column_dimensions[col_letter].width = 20

        # Add formulas for each audit leader
        section3_start_row = self._find_section3_data_start(ws)
        section3_end_row = section3_start_row + len(audit_leaders) - 1

        for i, leader in enumerate(audit_leaders):
            row = leader_start_row + i

            weighted_score_col = get_column_letter(summary_start_col)
            weighted_avg_col = get_column_letter(summary_start_col + 1)

            # Sum of all weighted scores for this leader
            test_range = f"D{row}:{get_column_letter(summary_start_col - 1)}{row}"
            ws[f'{weighted_score_col}{row}'] = f'=SUM({test_range})'
            ws[f'{weighted_score_col}{row}'].number_format = '0.00'
            ws[f'{weighted_score_col}{row}'].border = self.border

            # Weighted average calculation
            total_count_formula = f'=COUNTIFS($B${section3_start_row}:$B${section3_end_row},$B{row},D${section3_start_row}:D${section3_end_row},"<>N/A")'
            ws[f'{weighted_avg_col}{row}'] = f'=IF({total_count_formula}=0,"N/A",{weighted_score_col}{row}/({total_count_formula}*5))'
            ws[f'{weighted_avg_col}{row}'].number_format = '0.00'
            ws[f'{weighted_avg_col}{row}'].border = self.border

    # ==========================================
    # SECTION 3: AUDIT LEADER AVERAGE TEST RESULTS
    # ==========================================

    def _process_section3_detailed_test_results(self, ws, rule_names: List[str],
                                               audit_leaders: List[str], leader_rule_matrix: Dict[str, Dict]):
        """Process Section 3: Detailed Test Results."""
        logger.info("Processing Section 3: Audit Leader Average Test Results")

        # Setup Section 3 structure
        section3_start_row = self._setup_section3_structure(ws, rule_names, audit_leaders)

        # Add blue analytics section
        self._add_section3_analytics_rows(ws, rule_names, section3_start_row)

        # Populate detailed grid with results
        self._populate_section3_detailed_grid(ws, rule_names, audit_leaders, leader_rule_matrix, section3_start_row)

    def _setup_section3_structure(self, ws, rule_names: List[str], audit_leaders: List[str]) -> int:
        """Setup the structure for Section 3 and return starting row."""
        # Find where Section 3 should start (after Section 2)
        section3_start_row = 35 + len(audit_leaders)

        # Add section title
        ws[f'A{section3_start_row}'] = "Audit Leader Average Test Results"
        ws[f'A{section3_start_row}'].font = self.header_font

        # Add headers
        header_row = section3_start_row + 2
        ws[f'B{header_row}'] = "Audit Leader"
        ws[f'C{header_row}'] = "Samples Tested for Audit Leader"

        # Add rule headers
        start_col = 4
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(start_col + i)
            ws[f'{col_letter}{header_row}'] = rule_name
            ws[f'{col_letter}{header_row}'].font = self.header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.border
            ws.column_dimensions[col_letter].width = 20

        # Add aggregate headers
        agg_start_col = start_col + len(rule_names)
        agg_headers = ["GC Count", "PC Count", "DNC Count", "Total Applicable Count", "Average Score", "Average Rating: 4 Point Scale"]

        for i, header in enumerate(agg_headers):
            col_letter = get_column_letter(agg_start_col + i)
            ws[f'{col_letter}{header_row}'] = header
            ws[f'{col_letter}{header_row}'].font = self.header_font
            ws[f'{col_letter}{header_row}'].alignment = self.center_align
            ws[f'{col_letter}{header_row}'].border = self.border
            ws.column_dimensions[col_letter].width = 15

        return header_row + 1  # Return data starting row

    def _add_section3_analytics_rows(self, ws, rule_names: List[str], data_start_row: int):
        """Add the blue analytics section above the data."""
        analytics_start_row = data_start_row - 10  # Place analytics section above data

        # Add Analytics rows
        for i, rule_name in enumerate(rule_names):
            col_letter = get_column_letter(4 + i)  # Start at column D

            # Row 1: "Analytics"
            ws[f'{col_letter}{analytics_start_row}'] = "Analytics"
            ws[f'{col_letter}{analytics_start_row}'].font = self.header_font
            ws[f'{col_letter}{analytics_start_row}'].alignment = self.center_align

            # Row 2: Blank
            ws[f'{col_letter}{analytics_start_row + 1}'] = ""

            # Row 3: Error threshold (from rule data)
            ws[f'{col_letter}{analytics_start_row + 2}'] = "2%"  # Default, should come from rule
            ws[f'{col_letter}{analytics_start_row + 2}'].alignment = self.center_align

            # Row 4: Risk level (from rule data)
            ws[f'{col_letter}{analytics_start_row + 3}'] = "3"  # Default, should come from rule
            ws[f'{col_letter}{analytics_start_row + 3}'].alignment = self.center_align

            # Rows 5-7: "Not Applicable"
            for j in range(3):
                ws[f'{col_letter}{analytics_start_row + 4 + j}'] = "Not Applicable"
                ws[f'{col_letter}{analytics_start_row + 4 + j}'].alignment = self.center_align

            # Row 8: Rule ID
            ws[f'{col_letter}{analytics_start_row + 7}'] = rule_name
            ws[f'{col_letter}{analytics_start_row + 7}'].alignment = self.center_align

            # Apply blue background to this section
            for row_offset in range(8):
                ws[f'{col_letter}{analytics_start_row + row_offset}'].fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                ws[f'{col_letter}{analytics_start_row + row_offset}'].font = Font(color="FFFFFF")

    def _populate_section3_detailed_grid(self, ws, rule_names: List[str], audit_leaders: List[str],
                                        leader_rule_matrix: Dict[str, Dict], data_start_row: int):
        """Populate the detailed grid with compliance status and calculations."""

        start_col = 4
        rule_cols = {}
        for i, rule_name in enumerate(rule_names):
            rule_cols[rule_name] = get_column_letter(start_col + i)

        # Calculate aggregate column positions
        agg_start_col = start_col + len(rule_names)
        gc_col = get_column_letter(agg_start_col)
        pc_col = get_column_letter(agg_start_col + 1)
        dnc_col = get_column_letter(agg_start_col + 2)
        total_col = get_column_letter(agg_start_col + 3)
        avg_col = get_column_letter(agg_start_col + 4)
        rating_col = get_column_letter(agg_start_col + 5)

        # Populate data for each audit leader
        for i, leader in enumerate(audit_leaders):
            row = data_start_row + i

            # Leader name and blank samples column
            ws[f'B{row}'] = leader
            ws[f'B{row}'].border = self.border
            ws[f'C{row}'] = ""  # Manual entry field
            ws[f'C{row}'].border = self.border

            # Track totals for this leader
            total_gc = total_pc = total_dnc = 0

            # Add status for each rule
            for rule_name in rule_names:
                col_letter = rule_cols[rule_name]

                if leader in leader_rule_matrix and rule_name in leader_rule_matrix[leader]:
                    rule_data = leader_rule_matrix[leader][rule_name]
                    status = rule_data.get('status', 'N/A')

                    # Apply color based on status
                    if status == 'GC':
                        fill = self.green_fill
                    elif status == 'PC':
                        fill = self.yellow_fill
                    elif status == 'DNC':
                        fill = self.red_fill
                    else:
                        fill = None

                    ws[f'{col_letter}{row}'] = status
                    if fill:
                        ws[f'{col_letter}{row}'].fill = fill

                    # Add to totals
                    total_gc += rule_data.get('gc_count', 0)
                    total_pc += rule_data.get('pc_count', 0)
                    total_dnc += rule_data.get('dnc_count', 0)
                else:
                    ws[f'{col_letter}{row}'] = 'N/A'

                ws[f'{col_letter}{row}'].alignment = self.center_align
                ws[f'{col_letter}{row}'].border = self.border

            # Populate aggregate columns
            ws[f'{gc_col}{row}'] = total_gc
            ws[f'{pc_col}{row}'] = total_pc
            ws[f'{dnc_col}{row}'] = total_dnc
            ws[f'{total_col}{row}'] = total_gc + total_pc + total_dnc

            # Calculate average score and rating
            applicable_total = total_gc + total_pc + total_dnc
            if applicable_total > 0:
                avg_score = (total_gc * 5 + total_pc * 3 + total_dnc * 1) / applicable_total
                ws[f'{avg_col}{row}'] = avg_score
                ws[f'{avg_col}{row}'].number_format = '0.00'

                if avg_score >= 4:
                    rating = "GC"
                    fill = self.green_fill
                elif avg_score >= 2.5:
                    rating = "PC"
                    fill = self.yellow_fill
                else:
                    rating = "DNC"
                    fill = self.red_fill

                ws[f'{rating_col}{row}'] = rating
                ws[f'{rating_col}{row}'].fill = fill
            else:
                ws[f'{avg_col}{row}'] = 0
                ws[f'{rating_col}{row}'] = "N/A"

            # Apply formatting to all aggregate cells
            for col in [gc_col, pc_col, dnc_col, total_col, avg_col, rating_col]:
                ws[f'{col}{row}'].alignment = self.center_align
                ws[f'{col}{row}'].border = self.border

    # ==========================================
    # HELPER METHODS
    # ==========================================

    def _find_section3_data_start(self, ws) -> int:
        """Find the starting row of Section 3 data."""
        # This returns the data row, not the header row
        header_row = self._find_section3_header_row(ws)
        return header_row + 1 if header_row else 31

    def _find_section3_header_row(self, ws) -> int:
        """Find the header row of Section 3."""
        for row in range(20, 100):
            cell_b = ws[f'B{row}'].value
            cell_c = ws[f'C{row}'].value  
            if (cell_b and "Audit Leader" in str(cell_b) and 
                cell_c and "Samples Tested for Audit Leader" in str(cell_c)):
                return row

        logger.warning("Could not find Section 3 header row, using fallback row 30")
        return 30


def create_summary_placeholder_template():
    """Create a template showing the dynamic structure with placeholders."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IAG and AL Results Summary"

    # Header
    ws['B2'] = "QA {{Review_Year}} and AL Summary Ratings and Results Report"

    # IAG Overall section
    ws['B4'] = "IAG Overall Results and Rating"
    ws['B5'] = "Total Score"
    ws['B6'] = "GC Score"
    ws['B7'] = "PC Score"
    ws['B8'] = "DNC Score"
    ws['B9'] = "Total Count of Applications"
    ws['B10'] = "Weighted Score Across All Tests"
    ws['B11'] = "Weighted Rating Across All Tests"

    # Placeholders
    ws['D6'] = "{{IAG_COUNT_OF_GC}}"
    ws['D7'] = "{{IAG_COUNT_OF_PC}}"
    ws['D8'] = "{{IAG_COUNT_OF_DNC}}"
    ws['D9'] = "{{IAG_TOTAL_COUNT}}"

    # Audit Leader section
    ws['B13'] = "Audit Leader Overall Results and Ratings"
    ws['B14'] = "Audit Leader"
    ws['C14'] = "Measurement Description"
    ws['D14'] = "{{RULE_1_NAME}}"
    ws['E14'] = "{{RULE_2_NAME}}"
    ws['F14'] = "..."

    # Leader placeholders
    ws['B15'] = "{{AUDIT_LEADER_1_NAME}}"
    ws['B16'] = "{{AUDIT_LEADER_2_NAME}}"
    ws['B17'] = "..."

    # Notes
    ws['A20'] = "NOTES:"
    ws['A21'] = "1. Rules are added dynamically as columns"
    ws['A22'] = "2. Audit Leaders are added dynamically as rows"
    ws['A23'] = "3. Detailed analytics section populates automatically"
    ws['A24'] = "4. All formulas adjust to dynamic ranges"

    wb.save('qa_summary_dynamic_template_placeholders.xlsx')
    wb.close()

    print("Created dynamic summary placeholder template")


if __name__ == "__main__":
    create_summary_placeholder_template()