#!/usr/bin/env python3
"""
Script to extract and compare data from Excel files, specifically the 
"IAG and AL Results and Ratings" tab.
"""

import sys
import argparse
from pathlib import Path
try:
    from openpyxl import load_workbook
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)


def extract_sheet_data(file_path, sheet_name=None):
    """Extract data from an Excel sheet."""
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # List all sheet names
        print(f"\nAvailable sheets in {Path(file_path).name}:")
        for idx, name in enumerate(wb.sheetnames):
            print(f"  {idx+1}. {name}")
        
        # Find the IAG sheet
        target_sheet = None
        if sheet_name:
            if sheet_name in wb.sheetnames:
                target_sheet = wb[sheet_name]
            else:
                print(f"\nWarning: Sheet '{sheet_name}' not found")
        else:
            # Look for IAG sheet by partial name match
            for name in wb.sheetnames:
                if "IAG" in name and ("Results" in name or "Ratings" in name):
                    target_sheet = wb[name]
                    sheet_name = name
                    break
        
        if not target_sheet:
            print("\nNo IAG and AL Results sheet found")
            return None
            
        print(f"\nExtracting data from sheet: {sheet_name}")
        print("-" * 60)
        
        # Extract all non-empty cells
        data = {}
        max_row = target_sheet.max_row
        max_col = target_sheet.max_column
        
        print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
        
        # Extract headers and key sections
        print("\nSheet content preview:")
        print("-" * 60)
        
        # Show first 30 rows to understand structure
        for row in range(1, min(31, max_row + 1)):
            row_data = []
            has_content = False
            
            for col in range(1, min(11, max_col + 1)):  # First 10 columns
                cell = target_sheet.cell(row=row, column=col)
                value = cell.value
                
                # Store cell reference with value
                cell_ref = f"{chr(64+col)}{row}"
                
                if value is not None:
                    has_content = True
                    data[cell_ref] = {
                        'value': value,
                        'formula': cell.formula if hasattr(cell, 'formula') else None,
                        'type': type(value).__name__
                    }
                
                # Format for display
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(f"{value:.2f}" if isinstance(value, float) else str(value))
                else:
                    row_data.append(str(value)[:20])  # Truncate long strings
            
            if has_content:
                print(f"Row {row:2d}: " + " | ".join(f"{val:>12}" for val in row_data))
        
        # Look for specific sections
        print("\n\nKey sections found:")
        print("-" * 60)
        
        # Find IAG Overall Results section
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = target_sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "IAG Overall" in cell_value or "Overall Results" in cell_value:
                        print(f"\nFound IAG section at {chr(64+col)}{row}: {cell_value}")
                        # Show next few rows
                        for r in range(row, min(row+10, max_row+1)):
                            row_vals = []
                            for c in range(col, min(col+5, max_col+1)):
                                val = target_sheet.cell(row=r, column=c).value
                                row_vals.append(str(val) if val else "")
                            if any(row_vals):
                                print(f"  {r}: " + " | ".join(row_vals))
                    
                    elif "Audit Leader" in cell_value or "Leader" in cell_value:
                        print(f"\nFound Leader section at {chr(64+col)}{row}: {cell_value}")
                        # Show next few rows
                        for r in range(row, min(row+10, max_row+1)):
                            row_vals = []
                            for c in range(col, min(col+5, max_col+1)):
                                val = target_sheet.cell(row=r, column=c).value
                                row_vals.append(str(val) if val else "")
                            if any(row_vals):
                                print(f"  {r}: " + " | ".join(row_vals))
        
        # Extract formulas if any
        print("\n\nFormulas found:")
        print("-" * 60)
        formula_count = 0
        for cell_ref, cell_data in data.items():
            if cell_data.get('formula'):
                print(f"{cell_ref}: {cell_data['formula']}")
                formula_count += 1
                if formula_count > 20:  # Limit output
                    print("... (more formulas exist)")
                    break
        
        if formula_count == 0:
            print("No formulas found (all values are static)")
        
        return data
        
    except Exception as e:
        print(f"Error reading file: {e}")
        import traceback
        traceback.print_exc()
        return None


def compare_files(file1, file2, sheet_name=None):
    """Compare two Excel files."""
    print("="*80)
    print("EXCEL FILE COMPARISON")
    print("="*80)
    
    data1 = extract_sheet_data(file1, sheet_name)
    print("\n" + "="*80)
    data2 = extract_sheet_data(file2, sheet_name)
    
    if data1 and data2:
        print("\n" + "="*80)
        print("DIFFERENCES SUMMARY")
        print("="*80)
        
        # Compare keys
        keys1 = set(data1.keys())
        keys2 = set(data2.keys())
        
        only_in_file1 = keys1 - keys2
        only_in_file2 = keys2 - keys1
        common_keys = keys1 & keys2
        
        if only_in_file1:
            print(f"\nCells only in {Path(file1).name}: {len(only_in_file1)}")
            for key in sorted(list(only_in_file1))[:10]:
                print(f"  {key}: {data1[key]['value']}")
        
        if only_in_file2:
            print(f"\nCells only in {Path(file2).name}: {len(only_in_file2)}")
            for key in sorted(list(only_in_file2))[:10]:
                print(f"  {key}: {data2[key]['value']}")
        
        # Compare values in common cells
        differences = []
        for key in common_keys:
            if data1[key]['value'] != data2[key]['value']:
                differences.append((key, data1[key]['value'], data2[key]['value']))
        
        if differences:
            print(f"\nCells with different values: {len(differences)}")
            for key, val1, val2 in differences[:20]:
                print(f"  {key}: '{val1}' vs '{val2}'")


def main():
    parser = argparse.ArgumentParser(description="Extract and compare Excel file data")
    parser.add_argument("file1", help="First Excel file path")
    parser.add_argument("file2", nargs="?", help="Second Excel file path (for comparison)")
    parser.add_argument("-s", "--sheet", help="Specific sheet name to examine")
    
    args = parser.parse_args()
    
    if args.file2:
        compare_files(args.file1, args.file2, args.sheet)
    else:
        extract_sheet_data(args.file1, args.sheet)


if __name__ == "__main__":
    if len(sys.argv) == 1:
        print("Usage examples:")
        print("  python compare_excel_tabs.py report.xlsx")
        print("  python compare_excel_tabs.py report.xlsx template.xlsx")
        print("  python compare_excel_tabs.py report.xlsx -s 'IAG and AL Results and Ratings'")
    else:
        main()