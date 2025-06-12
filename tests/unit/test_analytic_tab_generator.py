"""
Unit tests for AnalyticTabGenerator class.

Tests the generation of individual analytic tabs matching QA-ID format.
"""

import unittest
from unittest.mock import Mock, MagicMock, patch
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import tempfile
import os

from openpyxl import Workbook, load_workbook

from core.rule_engine.rule_manager import ValidationRule
from core.rule_engine.rule_evaluator import RuleEvaluationResult
from core.rule_engine.compliance_determiner import ComplianceStatus


class TestAnalyticTabGenerator(unittest.TestCase):
    """Test cases for AnalyticTabGenerator."""
    
    def setUp(self):
        """Set up test fixtures."""
        # Create a mock rule
        self.mock_rule = ValidationRule(
            rule_id="70",
            name="IT/IS Risk High with No Applications",
            title="Audit Entity IT or IS Level 1 Risk is High or Critical with 0 Applications Mapped",
            description="Identifies audit entities where the IT or IS risk categories are rated 'high' or 'critical' and 0 applications are mapped to the AE.",
            formula="=IF(AND(OR([IT Risk]=\"High\", [IT Risk]=\"Critical\", [IS Risk]=\"High\", [IS Risk]=\"Critical\"), [Applications]=\"\"), \"DNC\", \"GC\")",
            threshold=0.98,
            category="Direct ACT",
            metadata={
                'observation_criteria': 'IT / IS Risk Ratings and application risk ratings are generally expected to be in alignment.',
                'threshold_rationale': 'The 2% threshold is deemed appropriate as it accounts for manual tagging errors and inappropriately overriding the inherent risks.',
                'exclusions': 'x, y and z from the population for ABC Reason',
                'data_source_method': 'QA generated the population directly from the system of record, Archer, using Archer\'s reporting functionality.',
                'responsible_party_column': 'Audit Leader'
            }
        )
        
        # Create test data
        self.test_data = pd.DataFrame({
            'Audit Leader': ['Leader1', 'Leader1', 'Leader2', 'Leader2', 'Leader3'],
            'Audit Entity ID': ['AE-1', 'AE-2', 'AE-3', 'AE-4', 'AE-5'],
            'Audit Entity Name': ['Entity One', 'Entity Two', 'Entity Three', 'Entity Four', 'Entity Five'],
            'IT Risk': ['High', 'Low', 'Critical', 'Medium', 'High'],
            'IS Risk': ['Medium', 'Low', 'High', 'Low', 'Critical'],
            'Applications': ['APP-1', '', 'APP-2,APP-3', '', '']
        })
        
        # Create result dataframe with evaluation results
        self.result_df = self.test_data.copy()
        self.result_df['Result_Rule_70'] = [True, True, True, True, False]  # Last one fails
        
        # Create mock rule evaluation result
        self.mock_rule_result = Mock(spec=RuleEvaluationResult)
        self.mock_rule_result.rule = self.mock_rule
        self.mock_rule_result.result_df = self.result_df
        self.mock_rule_result.result_column = 'Result_Rule_70'
        self.mock_rule_result.compliance_status = ComplianceStatus.COMPLIANT
        self.mock_rule_result.summary = {
            'rule_id': '70',
            'rule_name': 'IT/IS Risk High with No Applications',
            'compliance_status': ComplianceStatus.COMPLIANT,
            'compliance_rate': 0.8,
            'total_items': 5,
            'gc_count': 4,
            'pc_count': 0,
            'dnc_count': 1,
            'na_count': 0,
            'error_count': 0
        }
        
        # Mock party results
        self.mock_rule_result.party_results = {
            'Leader1': {
                'status': ComplianceStatus.COMPLIANT,
                'metrics': {
                    'total_count': 2,
                    'gc_count': 2,
                    'pc_count': 0,
                    'dnc_count': 0,
                    'na_count': 0,
                    'dnc_rate': 0.0,
                    'error_count': 0
                }
            },
            'Leader2': {
                'status': ComplianceStatus.COMPLIANT,
                'metrics': {
                    'total_count': 2,
                    'gc_count': 2,
                    'pc_count': 0,
                    'dnc_count': 0,
                    'na_count': 0,
                    'dnc_rate': 0.0,
                    'error_count': 0
                }
            },
            'Leader3': {
                'status': ComplianceStatus.NOT_COMPLIANT,
                'metrics': {
                    'total_count': 1,
                    'gc_count': 0,
                    'pc_count': 0,
                    'dnc_count': 1,
                    'na_count': 0,
                    'dnc_rate': 1.0,
                    'error_count': 0
                }
            }
        }
        
        # Mock get_failing_items method
        self.mock_rule_result.get_failing_items = Mock(return_value=self.result_df[self.result_df['Result_Rule_70'] == False])
        
    def test_initialization(self):
        """Test AnalyticTabGenerator initialization."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        self.assertEqual(generator.rule, self.mock_rule)
        self.assertEqual(generator.rule_result, self.mock_rule_result)
        self.assertTrue(generator.source_data.equals(self.test_data))
        self.assertEqual(generator.responsible_party_column, 'Audit Leader')
    
    def test_worksheet_name_creation(self):
        """Test worksheet name generation."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        ws_name = generator._create_worksheet_name()
        self.assertEqual(ws_name, "QA-ID-70")
        
        # Test with long rule ID
        self.mock_rule.rule_id = "very_long_rule_id_that_exceeds_31_chars"
        ws_name = generator._create_worksheet_name()
        self.assertLessEqual(len(ws_name), 31)
        self.assertTrue(ws_name.startswith("QA-ID-"))
    
    def test_metadata_getters(self):
        """Test metadata getter methods."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        # Test analytic title
        title = generator._get_analytic_title()
        self.assertEqual(title, "Audit Entity IT or IS Level 1 Risk is High or Critical with 0 Applications Mapped")
        
        # Test analytic type
        analytic_type = generator._get_analytic_type()
        self.assertEqual(analytic_type, "Direct ACT")
        
        # Test observation criteria
        obs_criteria = generator._get_observation_criteria()
        self.assertEqual(obs_criteria, "IT / IS Risk Ratings and application risk ratings are generally expected to be in alignment.")
        
        # Test population criteria
        pop_criteria = generator._get_population_criteria()
        self.assertIn("5 total records", pop_criteria)
        self.assertIn("QA manually excluded x, y and z", pop_criteria)
        
        # Test threshold rationale
        threshold_rationale = generator._get_threshold_rationale()
        self.assertEqual(threshold_rationale, "The 2% threshold is deemed appropriate as it accounts for manual tagging errors and inappropriately overriding the inherent risks.")
    
    def test_overall_test_result_determination(self):
        """Test determination of overall test results."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        # Test GC result (error rate <= 2%)
        self.assertEqual(generator._determine_overall_test_result(0.01), "GC")
        self.assertEqual(generator._determine_overall_test_result(0.02), "GC")
        
        # Test PC result (2% < error rate <= 50%)
        self.assertEqual(generator._determine_overall_test_result(0.03), "PC")
        self.assertEqual(generator._determine_overall_test_result(0.50), "PC")
        
        # Test DNC result (error rate > 50%)
        self.assertEqual(generator._determine_overall_test_result(0.51), "DNC")
        self.assertEqual(generator._determine_overall_test_result(1.0), "DNC")
    
    def test_generate_worksheet_structure(self):
        """Test that generated worksheet has correct structure."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        # Create workbook and generate worksheet
        workbook = Workbook()
        ws_name = generator.generate_analytic_worksheet(workbook)
        
        self.assertIn(ws_name, workbook.sheetnames)
        worksheet = workbook[ws_name]
        
        # Test Section 1: Metadata (rows 1-10)
        self.assertEqual(worksheet['B1'].value, 'Analytic Title:')
        self.assertEqual(worksheet['C1'].value, self.mock_rule.title)
        self.assertEqual(worksheet['B2'].value, 'QA Analytic ID:')
        self.assertEqual(worksheet['C2'].value, '70')
        self.assertEqual(worksheet['B3'].value, 'Analytic Type')
        self.assertEqual(worksheet['C3'].value, 'Direct ACT')
        self.assertEqual(worksheet['B10'].value, 'Threshold:')
        self.assertEqual(worksheet['C10'].value, 0.98)
        
        # Test Section 2: IAG Summary Results (row 13)
        self.assertEqual(worksheet['B13'].value, 'IAG Summary Results')
        self.assertEqual(worksheet['B14'].value, 'Name for Aggregating Results')
        self.assertEqual(worksheet['C14'].value, 'Name')
        self.assertEqual(worksheet['D14'].value, 'Score and Rating')
        self.assertEqual(worksheet['E14'].value, 'Results')
        
        # Test IAG Overall row
        self.assertEqual(worksheet['B21'].value, 'IAG Overall')
        self.assertEqual(worksheet['D21'].value, 'Test Result')
        
        # Test Section 3: First audit leader section should start at row 24
        self.assertEqual(worksheet['B24'].value, 'Summary Results by Audit Leader')
        
        # Test Section 4: Detailed Results (row 58)
        self.assertEqual(worksheet['B58'].value, 'Detailed Results')
        self.assertEqual(worksheet['B59'].value, 'Archer Record Data')
        self.assertEqual(worksheet['I59'].value, 'QA Analytic Test(s) and Results')
    
    def test_party_results_calculation(self):
        """Test calculation of party results when not provided."""
        # Create generator without party results
        mock_result_no_parties = Mock(spec=RuleEvaluationResult)
        mock_result_no_parties.rule = self.mock_rule
        mock_result_no_parties.result_df = self.result_df
        mock_result_no_parties.result_column = 'Result_Rule_70'
        mock_result_no_parties.party_results = {}
        
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=mock_result_no_parties,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        party_results = generator._calculate_party_results()
        
        # Should have results for each leader
        self.assertIn('Leader1', party_results)
        self.assertIn('Leader2', party_results)
        self.assertIn('Leader3', party_results)
        
        # Check counts
        self.assertEqual(party_results['Leader1']['total_count'], 2)
        self.assertEqual(party_results['Leader3']['total_count'], 1)
    
    def test_detailed_headers_creation(self):
        """Test creation of detailed results headers."""
        generator = AnalyticTabGenerator(
            rule=self.mock_rule,
            rule_result=self.mock_rule_result,
            source_data=self.test_data,
            responsible_party_column='Audit Leader'
        )
        
        headers = generator._create_detailed_headers()
        
        # Check base headers
        self.assertEqual(headers[0], "Audit Leader")
        self.assertEqual(headers[1], "Audit Entity ID")
        self.assertEqual(headers[2], "Audit Entity Name")
        
        # Check test result headers
        self.assertIn(f"Test: {self.mock_rule.name}", headers)
        self.assertIn("Overall Test Result Override\n(If applicable)", headers)
        self.assertIn("Finding ID, If applicable", headers)
    
    def test_excel_output_generation(self):
        """Test actual Excel file generation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            generator = AnalyticTabGenerator(
                rule=self.mock_rule,
                rule_result=self.mock_rule_result,
                source_data=self.test_data,
                responsible_party_column='Audit Leader'
            )
            
            # Create workbook and generate worksheet
            workbook = Workbook()
            ws_name = generator.generate_analytic_worksheet(workbook)
            
            # Save to file
            output_path = os.path.join(temp_dir, "test_analytic_tab.xlsx")
            workbook.save(output_path)
            
            # Verify file exists
            self.assertTrue(os.path.exists(output_path))
            
            # Load and verify structure
            loaded_wb = load_workbook(output_path)
            self.assertIn(ws_name, loaded_wb.sheetnames)
            
            # Check some key cells
            ws = loaded_wb[ws_name]
            self.assertEqual(ws['B1'].value, 'Analytic Title:')
            self.assertEqual(ws['B13'].value, 'IAG Summary Results')
            self.assertEqual(ws['B58'].value, 'Detailed Results')


if __name__ == '__main__':
    unittest.main()