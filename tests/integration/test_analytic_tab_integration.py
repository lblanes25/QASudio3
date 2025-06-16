"""
Integration test for Individual Analytic Tab generation.

Tests the complete flow from validation through report generation,
verifying that the output matches the QA-ID format specification.
"""

import unittest
import tempfile
import os
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime

from openpyxl import load_workbook

from services.validation_service import ValidationPipeline
from core.rule_engine.rule_manager import ValidationRule, ValidationRuleManager
# from reporting.generation.report_generator import ReportGenerator  # Removed


class TestAnalyticTabIntegration(unittest.TestCase):
    """Integration tests for analytic tab generation."""
    
    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.output_dir = Path(self.temp_dir) / "output"
        self.output_dir.mkdir(exist_ok=True)
        
        # Create test data matching the QA-ID-70 example
        self.test_data = pd.DataFrame({
            'Audit Leader': ['Name1'] * 10 + ['Name2'] * 10 + ['Name3'] * 10,
            'Audit Entity ID': [f'AE-{i}' for i in range(1, 31)],
            'Audit Entity Name': [f'Entity {i}' for i in range(1, 31)],
            'Audit Entity Status': ['Active'] * 30,
            'IT Risk': ['High', 'Critical', 'Low', 'Medium', 'High'] * 6,
            'IS Risk': ['Medium', 'High', 'Low', 'Low', 'Critical'] * 6,
            'Key Primary and Secondary IT Applications': [
                'ARA-6', '', 'ARA-7', 'ARA-8', '',
                'ARA-9', 'ARA-10', '', 'ARA-11', 'ARA-12',
                '', 'ARA-13', 'ARA-14', '', 'ARA-15',
                'ARA-16', '', 'ARA-17', 'ARA-18', '',
                'ARA-19', 'ARA-20', '', 'ARA-21', 'ARA-22',
                '', 'ARA-23', 'ARA-24', '', 'ARA-25'
            ]
        })
        
        # Save test data
        self.data_file = self.output_dir / "test_data.csv"
        self.test_data.to_csv(self.data_file, index=False)
        
        # Create validation rule matching QA-ID-70
        self.rule = ValidationRule(
            rule_id="70",
            name="IT/IS Risk High with No Applications",
            title="Audit Entity IT or IS Level 1 Risk is High or Critical with 0 Applications Mapped",
            description="Identifies audit entities where the IT or IS risk categories are rated 'high' or 'critical' and 0 applications are mapped to the AE.",
            formula='=IF(AND(OR([IT Risk]="High", [IT Risk]="Critical", [IS Risk]="High", [IS Risk]="Critical"), [Key Primary and Secondary IT Applications]=""), FALSE, TRUE)',
            threshold=0.98,
            category="Direct ACT",
            metadata={
                'observation_criteria': 'IT / IS Risk Ratings and application risk ratings are generally expected to be in alignment.',
                'threshold_rationale': 'The 2% threshold is deemed appropriate as it accounts for manual tagging errors and inappropriately overriding the inherent risks.',
                'data_source_method': 'QA generated the population directly from the system of record, Archer, using Archer\'s reporting functionality.',
                'responsible_party_column': 'Audit Leader'
            }
        )
        
        # Initialize pipeline
        self.pipeline = ValidationPipeline(
            output_dir=str(self.output_dir)
        )
        
        # Add rule to rule manager
        self.pipeline.rule_manager.add_rule(self.rule)
    
    def tearDown(self):
        """Clean up test environment."""
        import shutil
        shutil.rmtree(self.temp_dir)
    
    def test_comprehensive_report_generation(self):
        """Test generation of comprehensive IAG report with individual tabs."""
        # Run validation
        results = self.pipeline.validate_data_source(
            data_source=str(self.data_file),
            rule_ids=["70"],
            analytic_id="TEST_ANALYTIC",
            responsible_party_column="Audit Leader"
        )
        
        # Verify validation succeeded
        self.assertTrue(results['success'])
        self.assertIn('rule_results', results)
        
        # Generate comprehensive report
        report_path = self.pipeline.generate_comprehensive_iag_report(
            results=results,
            rule_results=results['rule_results'],
            responsible_party_column="Audit Leader"
        )
        
        # Verify report was created
        self.assertIsInstance(report_path, Path)
        self.assertTrue(report_path.exists())
        
        # Load and verify workbook structure
        workbook = load_workbook(report_path)
        
        # Check that QA-ID-70 tab exists
        self.assertIn("QA-ID-70", workbook.sheetnames)
        
        # Verify QA-ID-70 tab structure
        ws = workbook["QA-ID-70"]
        
        # Section 1: Metadata
        self.assertEqual(ws['B1'].value, 'Analytic Title:')
        self.assertEqual(ws['C1'].value, self.rule.title)
        self.assertEqual(ws['B2'].value, 'QA Analytic ID:')
        self.assertEqual(ws['C2'].value, '70')
        self.assertEqual(ws['B3'].value, 'Analytic Type')
        self.assertEqual(ws['C3'].value, 'Direct ACT')
        self.assertEqual(ws['B10'].value, 'Threshold:')
        self.assertAlmostEqual(ws['C10'].value, 0.98, places=2)
        
        # Section 2: IAG Summary Results
        self.assertEqual(ws['B13'].value, 'IAG Summary Results')
        self.assertEqual(ws['B14'].value, 'Name for Aggregating Results')
        self.assertEqual(ws['D15'].value, 'Count of "GC"')
        self.assertEqual(ws['B21'].value, 'IAG Overall')
        
        # Section 3: Audit Leader sections
        self.assertEqual(ws['B24'].value, 'Summary Results by Audit Leader')
        
        # Section 4: Detailed Results
        self.assertEqual(ws['B58'].value, 'Detailed Results')
        self.assertEqual(ws['B59'].value, 'Archer Record Data')
        self.assertEqual(ws['I59'].value, 'QA Analytic Test(s) and Results')
        
        # Check headers in row 60
        self.assertEqual(ws['B60'].value, 'Audit Leader')
        self.assertEqual(ws['C60'].value, 'Audit Entity ID')
        self.assertEqual(ws['D60'].value, 'Audit Entity Name')
        
        # Verify some data rows exist
        self.assertIsNotNone(ws['B61'].value)  # Should have data
        self.assertIsNotNone(ws['C61'].value)
    
    def test_multiple_rules_generation(self):
        """Test generation with multiple rules creates multiple tabs."""
        # Create additional rule
        rule2 = ValidationRule(
            rule_id="71",
            name="Test Rule 2",
            description="Another test rule",
            formula='=[IT Risk]="High"',
            threshold=0.95,
            metadata={'responsible_party_column': 'Audit Leader'}
        )
        
        self.pipeline.rule_manager.add_rule(rule2)
        
        # Run validation with both rules
        results = self.pipeline.validate_data_source(
            data_source=str(self.data_file),
            rule_ids=["70", "71"],
            analytic_id="MULTI_RULE_TEST",
            responsible_party_column="Audit Leader"
        )
        
        # Generate comprehensive report
        report_path = self.pipeline.generate_comprehensive_iag_report(
            results=results,
            rule_results=results['rule_results'],
            responsible_party_column="Audit Leader"
        )
        
        # Verify both tabs exist
        workbook = load_workbook(report_path)
        self.assertIn("QA-ID-70", workbook.sheetnames)
        self.assertIn("QA-ID-71", workbook.sheetnames)
    
    def test_format_matches_specification(self):
        """Test that output format exactly matches the specification."""
        # Run validation
        results = self.pipeline.validate_data_source(
            data_source=str(self.data_file),
            rule_ids=["70"],
            analytic_id="FORMAT_TEST",
            responsible_party_column="Audit Leader"
        )
        
        # Generate report
        report_path = self.pipeline.generate_comprehensive_iag_report(
            results=results,
            rule_results=results['rule_results'],
            responsible_party_column="Audit Leader"
        )
        
        # Load workbook
        workbook = load_workbook(report_path)
        ws = workbook["QA-ID-70"]
        
        # Verify merged cells match specification
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        
        # Check key merges from specification
        self.assertIn('C1:H1', merged_ranges)  # Analytic Title
        self.assertIn('C3:H3', merged_ranges)  # Analytic Type
        self.assertIn('C4:H4', merged_ranges)  # Description
        self.assertIn('C5:H5', merged_ranges)  # Observation Criteria
        self.assertIn('C6:H6', merged_ranges)  # Population Criteria
        self.assertIn('C7:H7', merged_ranges)  # Population Completeness
        self.assertIn('C8:H8', merged_ranges)  # Sample Size
        self.assertIn('C9:H9', merged_ranges)  # Threshold Rationale
        self.assertIn('B59:H59', merged_ranges)  # Archer Record Data
        self.assertIn('I59:N59', merged_ranges)  # QA Analytic Test(s) and Results
        
        # Verify formulas in IAG summary section
        self.assertEqual(ws['C15'].value, '=$B$21')  # Reference to IAG Overall
        self.assertEqual(ws['C16'].value, '=$B$21')
        self.assertEqual(ws['C17'].value, '=$B$21')
        
        # Verify data types
        self.assertIsInstance(ws['E15'].value, (int, float))  # GC count should be numeric
        self.assertIsInstance(ws['E19'].value, (int, float))  # Error rate should be numeric


if __name__ == '__main__':
    unittest.main()