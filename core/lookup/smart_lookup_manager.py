"""
Smart Lookup Manager for zero-configuration LOOKUP function support.
Implements intelligent file discovery, lazy loading, and column indexing.
"""

import os
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
import pandas as pd

logger = logging.getLogger(__name__)


class SmartLookupManager:
    """
    Manages lookup operations across multiple data sources with zero configuration.
    
    Features:
    - Automatic file discovery and column indexing
    - Lazy loading for large files (>50MB by default)
    - Smart column matching and disambiguation
    - Caching for performance optimization
    - File aliasing for friendly names
    """
    
    def __init__(self):
        """Initialize the Smart Lookup Manager with default settings."""
        self.loaded_files = {}      # filepath -> DataFrame
        self.file_metadata = {}     # filepath -> metadata (columns, rows, size)
        self.column_index = {}      # column -> [filepaths]
        self.file_aliases = {}      # friendly name -> filepath
        self.lookup_cache = {}      # cache for performance
        self.lazy_threshold_mb = 50 # Files > 50MB use lazy loading
        self.value_indices = {}     # filepath -> {column -> value_index}
        self.cache_hits = 0         # Track cache hit statistics
        self.cache_misses = 0       # Track cache miss statistics
        self.max_cache_size = 10000 # Maximum cache entries
        
        # Tracking for report integration
        self.tracking_enabled = False
        self.tracked_operations = []
        self.current_row_index = None
        
        logger.info("SmartLookupManager initialized with lazy threshold: %sMB", 
                   self.lazy_threshold_mb)
    
    def add_file(self, filepath: str, df: pd.DataFrame = None, 
                 alias: str = None, lazy: bool = None):
        """
        Register file for lookups with optional lazy loading.
        
        Args:
            filepath: Path to the file
            df: Optional pre-loaded DataFrame
            alias: Optional friendly name for the file
            lazy: Optional flag to force lazy loading behavior
        """
        try:
            file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
            
            # Auto-determine lazy loading based on file size
            if lazy is None:
                lazy = file_size_mb > self.lazy_threshold_mb
            
            if lazy or df is None:
                # Lazy loading - just store metadata
                self.file_metadata[filepath] = {
                    'columns': self._peek_columns(filepath),
                    'row_count': self._count_rows(filepath),
                    'size_mb': file_size_mb,
                    'lazy': True
                }
                logger.info(f"Registered {alias or Path(filepath).stem} for lazy loading "
                           f"({file_size_mb:.1f}MB, {self.file_metadata[filepath]['row_count']:,} rows)")
            else:
                # Full loading
                self.loaded_files[filepath] = df
                self.file_metadata[filepath] = {
                    'columns': list(df.columns),
                    'row_count': len(df),
                    'size_mb': file_size_mb,
                    'lazy': False
                }
                logger.info(f"Loaded {alias or Path(filepath).stem} "
                           f"({file_size_mb:.1f}MB, {len(df):,} rows)")
            
            # Create friendly alias
            if alias:
                self.file_aliases[alias] = filepath
            else:
                alias = Path(filepath).stem
                self.file_aliases[alias] = filepath
            
            # Index columns for fast discovery
            for col in self.file_metadata[filepath]['columns']:
                if col not in self.column_index:
                    self.column_index[col] = []
                self.column_index[col].append(filepath)
            
            # Create indices for loaded data
            if not lazy and df is not None:
                self._create_lookup_indices(filepath, df)
                
        except Exception as e:
            logger.error(f"Error adding file {filepath}: {e}")
            raise
    
    def _peek_columns(self, filepath: str) -> List[str]:
        """Get column names without loading full file."""
        ext = Path(filepath).suffix.lower()
        try:
            if ext in ['.csv', '.tsv']:
                # Read just first line
                with open(filepath, 'r', encoding='utf-8') as f:
                    header = f.readline().strip()
                    return header.split(',' if ext == '.csv' else '\t')
            elif ext in ['.xlsx', '.xls']:
                # Use openpyxl to read just headers
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                columns = []
                for cell in ws[1]:
                    if cell.value:
                        columns.append(str(cell.value))
                wb.close()
                return columns
        except Exception as e:
            logger.warning(f"Error peeking columns from {filepath}: {e}")
        return []
    
    def _count_rows(self, filepath: str) -> int:
        """Count rows without loading full file."""
        ext = Path(filepath).suffix.lower()
        try:
            if ext in ['.csv', '.tsv']:
                with open(filepath, 'r', encoding='utf-8') as f:
                    return sum(1 for _ in f) - 1  # Subtract header
            elif ext in ['.xlsx', '.xls']:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                count = wb.active.max_row - 1  # Subtract header
                wb.close()
                return count
        except Exception as e:
            logger.warning(f"Error counting rows in {filepath}: {e}")
        return 0
    
    def _ensure_loaded(self, filepath: str) -> pd.DataFrame:
        """Load file data if it was lazy-loaded."""
        if filepath not in self.loaded_files:
            if filepath in self.file_metadata and self.file_metadata[filepath].get('lazy'):
                logger.info(f"Loading {Path(filepath).name} for first lookup...")
                # Import here to avoid circular imports
                from data_integration.io.importer import DataImporter
                importer = DataImporter()
                df = importer.load_file(filepath)
                self.loaded_files[filepath] = df
                self._create_lookup_indices(filepath, df)
                self.file_metadata[filepath]['lazy'] = False
                return df
        return self.loaded_files.get(filepath)
    
    def smart_lookup(self, lookup_value: Any, 
                    search_column: str = None,
                    return_column: str = None,
                    source_hint: str = None) -> Any:
        """
        Intelligent lookup with multiple resolution strategies.
        
        Args:
            lookup_value: The value to search for
            search_column: Column to search in (optional)
            return_column: Column to return value from
            source_hint: Optional hint for which file to use
            
        Returns:
            The found value or None if not found
        """
        # Strategy 1: If source hint provided, try that first
        if source_hint:
            result = self._try_lookup_in_source(
                lookup_value, search_column, return_column, source_hint
            )
            if result is not None:
                return result
        
        # Strategy 2: Find file with both columns
        if search_column and return_column:
            for filepath in self._find_files_with_columns(search_column, return_column):
                df = self._ensure_loaded(filepath)  # Load if needed
                if df is not None:
                    result = self._perform_lookup(
                        df, lookup_value, search_column, return_column, filepath
                    )
                    if result is not None:
                        return result
        
        # Strategy 3: Smart single-column lookup
        if not search_column and return_column:
            # Find lookup_value in any unique column, return return_column
            return self._smart_value_lookup(lookup_value, return_column)
        
        return None
    
    def _generate_cache_key(self, lookup_value: Any, search_col: str, 
                           return_col: str, filepath: str) -> str:
        """Generate a unique cache key for the lookup."""
        return f"{filepath}|{search_col}|{return_col}|{lookup_value}"
    
    def _perform_lookup(self, df: pd.DataFrame, lookup_value: Any, 
                       search_col: str, return_col: str, filepath: str) -> Any:
        """
        Perform actual lookup with clear missing value handling.
        Returns None if not found - this will make comparisons fail appropriately.
        """
        # Check cache first
        cache_key = self._generate_cache_key(lookup_value, search_col, return_col, filepath)
        if cache_key in self.lookup_cache:
            self.cache_hits += 1
            logger.debug(f"Cache hit for lookup: {lookup_value} in {search_col}")
            cached_result = self.lookup_cache[cache_key]
            
            # Track the operation if tracking is enabled
            if self.tracking_enabled:
                self._track_lookup_operation(lookup_value, search_col, return_col, 
                                           filepath, cached_result, from_cache=True)
            
            return cached_result
        
        self.cache_misses += 1
        result = None
        success = False
        
        try:
            # Check if we have an index for this column
            if filepath in self.value_indices and search_col in self.value_indices[filepath]:
                index = self.value_indices[filepath][search_col]
                if lookup_value in index.index:
                    result = index.loc[lookup_value, return_col]
                    success = True
            else:
                # Fallback to DataFrame search
                matches = df[df[search_col] == lookup_value]
                if not matches.empty:
                    result = matches.iloc[0][return_col]
                    success = True
            
            # Store in cache (including None results to avoid repeated searches)
            self._add_to_cache(cache_key, result)
            
            if result is None:
                # Not found - log for debugging but don't crash
                logger.debug(f"Lookup not found: {lookup_value} in {search_col} "
                            f"(file: {Path(filepath).name})")
            
            # Track the operation if tracking is enabled
            if self.tracking_enabled:
                self._track_lookup_operation(lookup_value, search_col, return_col, 
                                           filepath, result, from_cache=False, success=success)
            
            return result
            
        except Exception as e:
            logger.warning(f"Lookup error in {Path(filepath).name}: {e}")
            
            # Track the error if tracking is enabled
            if self.tracking_enabled:
                self._track_lookup_operation(lookup_value, search_col, return_col, 
                                           filepath, None, from_cache=False, 
                                           success=False, error=str(e))
            
            return None
    
    def _add_to_cache(self, cache_key: str, value: Any):
        """Add a value to the cache with size management."""
        # Implement simple LRU by clearing oldest entries if cache is too large
        if len(self.lookup_cache) >= self.max_cache_size:
            # Remove oldest 10% of cache entries
            entries_to_remove = int(self.max_cache_size * 0.1)
            for key in list(self.lookup_cache.keys())[:entries_to_remove]:
                del self.lookup_cache[key]
            logger.debug(f"Cache pruned: removed {entries_to_remove} oldest entries")
        
        self.lookup_cache[cache_key] = value
    
    def _find_files_with_columns(self, *columns) -> List[str]:
        """Find files that have all specified columns."""
        if not columns:
            return []
        
        # Start with files that have the first column
        candidates = set(self.column_index.get(columns[0], []))
        
        # Intersect with files that have other columns
        for col in columns[1:]:
            candidates &= set(self.column_index.get(col, []))
        
        return list(candidates)
    
    def _create_lookup_indices(self, filepath: str, df: pd.DataFrame):
        """Create indices for efficient lookups."""
        if filepath not in self.value_indices:
            self.value_indices[filepath] = {}
        
        # Create indices for columns that appear to be keys (high uniqueness)
        for col in df.columns:
            try:
                uniqueness_ratio = df[col].nunique() / len(df)
                if uniqueness_ratio > 0.9:  # 90% unique values
                    # Create index for fast lookups
                    self.value_indices[filepath][col] = df.set_index(col)
                    logger.debug(f"Created index for {col} in {Path(filepath).name} "
                                f"(uniqueness: {uniqueness_ratio:.1%})")
            except Exception as e:
                logger.warning(f"Could not create index for {col}: {e}")
    
    def _try_lookup_in_source(self, lookup_value: Any, search_column: str,
                             return_column: str, source_hint: str) -> Any:
        """Try to perform lookup in a specific source file."""
        # Check if source_hint is an alias
        filepath = self.file_aliases.get(source_hint)
        
        # If not an alias, check if it's a partial filename
        if not filepath:
            for fp in self.file_metadata:
                if source_hint in Path(fp).name:
                    filepath = fp
                    break
        
        if filepath:
            df = self._ensure_loaded(filepath)
            if df is not None:
                return self._perform_lookup(df, lookup_value, search_column, 
                                          return_column, filepath)
        
        return None
    
    def _smart_value_lookup(self, lookup_value: Any, return_column: str) -> Any:
        """
        Smart lookup when only value and return column are specified.
        Searches for the value in any unique column that exists in the same
        file as the return column.
        """
        # Create a cache key for smart lookups
        cache_key = f"smart|{return_column}|{lookup_value}"
        if cache_key in self.lookup_cache:
            self.cache_hits += 1
            logger.debug(f"Cache hit for smart lookup: {lookup_value} -> {return_column}")
            return self.lookup_cache[cache_key]
        
        self.cache_misses += 1
        result = None
        
        # Find files that have the return column
        files_with_return_col = self.column_index.get(return_column, [])
        
        for filepath in files_with_return_col:
            df = self._ensure_loaded(filepath)
            if df is None:
                continue
                
            # Look for the value in any indexed column
            if filepath in self.value_indices:
                for col, index in self.value_indices[filepath].items():
                    if lookup_value in index.index:
                        try:
                            result = index.loc[lookup_value, return_column]
                            self._add_to_cache(cache_key, result)
                            return result
                        except:
                            pass
            
            # Fallback: search in all columns
            for col in df.columns:
                if col == return_column:
                    continue
                try:
                    matches = df[df[col] == lookup_value]
                    if not matches.empty:
                        result = matches.iloc[0][return_column]
                        self._add_to_cache(cache_key, result)
                        return result
                except:
                    pass
        
        # Cache None result to avoid repeated searches
        self._add_to_cache(cache_key, None)
        return None
    
    def clear_cache(self):
        """Clear the lookup cache and reset statistics."""
        self.lookup_cache.clear()
        self.cache_hits = 0
        self.cache_misses = 0
        logger.info("Lookup cache cleared and statistics reset")
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics about loaded files and lookups."""
        cache_hit_rate = 0
        if self.cache_hits + self.cache_misses > 0:
            cache_hit_rate = self.cache_hits / (self.cache_hits + self.cache_misses)
        
        return {
            'files_loaded': len(self.loaded_files),
            'files_lazy': sum(1 for m in self.file_metadata.values() if m.get('lazy')),
            'total_columns': len(self.column_index),
            'total_rows': sum(m.get('row_count', 0) for m in self.file_metadata.values()),
            'cache_size': len(self.lookup_cache),
            'cache_hits': self.cache_hits,
            'cache_misses': self.cache_misses,
            'cache_hit_rate': f"{cache_hit_rate:.1%}",
            'indices_created': sum(len(idx) for idx in self.value_indices.values())
        }
    
    def unload_file(self, filepath: str):
        """Unload a file from memory and update indices."""
        # Remove from loaded files
        if filepath in self.loaded_files:
            del self.loaded_files[filepath]
        
        # Remove from metadata
        if filepath in self.file_metadata:
            columns = self.file_metadata[filepath].get('columns', [])
            del self.file_metadata[filepath]
            
            # Update column index
            for col in columns:
                if col in self.column_index:
                    self.column_index[col] = [
                        fp for fp in self.column_index[col] if fp != filepath
                    ]
                    if not self.column_index[col]:
                        del self.column_index[col]
        
        # Remove from value indices
        if filepath in self.value_indices:
            del self.value_indices[filepath]
        
        # Remove aliases
        aliases_to_remove = [alias for alias, fp in self.file_aliases.items() 
                           if fp == filepath]
        for alias in aliases_to_remove:
            del self.file_aliases[alias]
        
        # Clear related cache entries
        self._clear_file_from_cache(filepath)
        
        logger.info(f"Unloaded file: {Path(filepath).name}")
    
    def _clear_file_from_cache(self, filepath: str):
        """Clear cache entries related to a specific file."""
        keys_to_remove = [key for key in self.lookup_cache 
                         if filepath in str(key)]
        for key in keys_to_remove:
            del self.lookup_cache[key]
    
    # Report integration methods
    
    def enable_tracking(self):
        """Enable tracking of lookup operations for report generation."""
        self.tracking_enabled = True
        self.tracked_operations = []
        logger.debug("Lookup tracking enabled")
    
    def disable_tracking(self):
        """Disable tracking of lookup operations."""
        self.tracking_enabled = False
        logger.debug(f"Lookup tracking disabled. Captured {len(self.tracked_operations)} operations")
    
    def set_current_row(self, row_index: int):
        """Set the current row being processed for tracking purposes."""
        self.current_row_index = row_index
    
    def get_tracked_operations(self) -> List[Dict[str, Any]]:
        """Get all tracked lookup operations."""
        return self.tracked_operations.copy()
    
    def clear_tracked_operations(self):
        """Clear tracked operations."""
        self.tracked_operations = []
        self.current_row_index = None
    
    def _track_lookup_operation(self, lookup_value: Any, search_col: str, 
                               return_col: str, filepath: str, result: Any,
                               from_cache: bool = False, success: bool = True,
                               error: str = None):
        """Track a single lookup operation for reporting."""
        if not self.tracking_enabled:
            return
        
        operation = {
            'row_index': self.current_row_index,
            'lookup_value': lookup_value,
            'search_column': search_col,
            'return_column': return_col,
            'source_file': filepath,
            'source_alias': self.file_aliases.get(filepath, Path(filepath).stem),
            'result': result,
            'success': success,
            'from_cache': from_cache,
            'error': error,
            'timestamp': pd.Timestamp.now()
        }
        
        self.tracked_operations.append(operation)