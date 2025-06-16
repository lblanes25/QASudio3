#!/usr/bin/env python3
"""
Test IAG report generation with severity weighting
This runs a fresh validation with responsible_party_column to generate party_results
"""

import os
from services.validation_service import ValidationPipeline
from openpyxl import load_workbook

def test_with_severity_weighting():
    """Run validation and generate report with severity weighting"""
    
    # Step 1: Run validation with responsible_party_column
    print("Step 1: Running fresh validation with responsible_party_column...")
    pipeline = ValidationPipeline()
    
    # Use the business monitoring data
    data_file = "business_monitoring_dummy_data.xlsx"
    if not os.path.exists(data_file):
        print(f"ERROR: {data_file} not found!")
        return False
    
    # Run validation specifying responsible_party_column
    results = pipeline.validate_data_source(
        data_source=data_file,
        responsible_party_column='AuditLeader',  # This is critical!
        output_formats=['json', 'excel_template'],  # Generate both JSON and Excel
        use_all_rules=True  # Use all available rules
    )
    
    print(f"Validation status: {results.get('status')}")
    print(f"Output files: {results.get('output_files')}")
    
    # Check if party_results were generated
    if 'rule_results' in results:
        party_results_count = 0
        for rule_id, rule_result in results['rule_results'].items():
            if 'party_results' in rule_result:
                party_results_count += 1
                print(f"✓ Rule {rule_id} has party_results")
        
        print(f"\nTotal rules with party_results: {party_results_count}/{len(results['rule_results'])}")
    
    # Step 2: Generate Excel report using the results
    if results.get('output_files'):
        json_file = next((f for f in results['output_files'] if f.endswith('.json')), None)
        if json_file:
            print(f"\nStep 2: Generating Excel report from {json_file}")
            output_path = "test_iag_severity_weighted.xlsx"
            pipeline.generate_excel_report(json_file, output_path)
            
            # Verify the report
            print(f"\nStep 3: Verifying report...")
            wb = load_workbook(output_path)
            ws = wb["IAG Summary Report"]
            
            print(f"Overall Compliance Rate: {ws['B7'].value}")
            print(f"Overall Rating: {ws['B8'].value}")
            
            # Check the note to see which calculation was used
            note = ws['D7'].value
            if "severity weighting" in note and "Critical=3x" in note:
                print("✓ Severity weighting was used!")
            else:
                print("✗ Standard IAG scoring was used (no severity weighting)")
                
            print(f"\nReport saved as: {output_path}")
            return True
    
    return False

if __name__ == "__main__":
    success = test_with_severity_weighting()