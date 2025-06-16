#!/usr/bin/env python3
"""
Check what tabs exist in the master report
"""

from openpyxl import load_workbook

def check_master_report():
    """Check tabs in master report"""
    
    wb = load_workbook("test_complete_iag_report.xlsx")
    
    print("Sheets in master report:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    if "Guide" not in wb.sheetnames:
        print("\n⚠️  Guide tab is missing from master report!")
        print("This is why the split is failing.")
        print("\nTo fix: Generate a new master report using the latest code.")

if __name__ == "__main__":
    check_master_report()